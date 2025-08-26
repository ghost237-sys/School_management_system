from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from django.contrib.auth import login, authenticate, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Avg, Count, DecimalField
from django.urls import reverse
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.conf import settings
from landing.forms import SiteSettingsForm, GalleryImageForm, CategoryForm, CategoryMediaForm
from landing.models import SiteSettings, GalleryImage, Category, CategoryMedia

from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.decorators.cache import cache_page
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.views.decorators.http import require_POST
from django.utils.dateparse import parse_datetime
from django.db import transaction
from .messaging_utils import _normalize_msisdn, send_sms
from .mpesa_utils import initiate_stk_push, query_stk_status
from django.utils.http import url_has_allowed_host_and_scheme

import datetime
import openpyxl
import json
import pandas as pd

from .models import (
    User, Student, Teacher, Class, Subject, Exam, Term, Grade, 
    FeeCategory, FeeAssignment, FeePayment, Event, Deadline, 
    TeacherClassAssignment, Department, AcademicYear, Notification,
    PeriodSlot, DefaultTimetable, Attendance, MpesaTransaction, MpesaC2BLedger,
    GradeCommentTemplate,
)
from .forms import (
    AddStudentForm, StudentContactUpdateForm, EditStudentClassForm, FeeCategoryForm,
    GradeUploadForm, TeacherProfileUpdateForm, EditTermDatesForm
)
from django.contrib.auth.decorators import user_passes_test
from django.db.models.functions import Length
from django.db.models import Sum
import csv
import os
from django.conf import settings
from django.http import FileResponse, Http404
from io import BytesIO
from .pdf_utils import pdf_response_from_rows

def is_admin(user):
    return user.is_authenticated and getattr(user, 'role', None) == 'admin'

def is_admin_or_clerk(user):
    return user.is_authenticated and getattr(user, 'role', None) in ('admin', 'clerk')

@login_required(login_url='login')
@user_passes_test(is_admin)
def admin_website_settings(request):
    """Allow admins to manage public website settings and categories."""
    settings = SiteSettings.objects.first()
    if settings is None:
        settings = SiteSettings.objects.create()
    category_form = CategoryForm()
    media_form = CategoryMediaForm()
    edit_category_obj = None
    edit_id = request.GET.get('edit')
    if edit_id:
        try:
            edit_category_obj = Category.objects.get(pk=edit_id)
            category_form = CategoryForm(instance=edit_category_obj)
        except Category.DoesNotExist:
            edit_category_obj = None
    if request.method == 'POST':
        # Distinguish which form was submitted
        if 'save_settings' in request.POST:
            form = SiteSettingsForm(request.POST, request.FILES, instance=settings)
            if form.is_valid():
                form.save()
                messages.success(request, 'Website settings updated successfully.')
                return redirect('admin_website_settings')
            else:
                messages.error(request, 'Please correct the errors below.')
        elif 'add_category' in request.POST:
            category_form = CategoryForm(request.POST)
            if category_form.is_valid():
                category_form.save()
                messages.success(request, 'Category added.')
                return redirect('admin_website_settings')
            else:
                messages.error(request, 'Please correct the errors in category form.')
        elif 'save_category' in request.POST:
            cat_id = request.POST.get('category_id')
            if not cat_id:
                messages.error(request, 'Missing category ID.')
            else:
                instance = get_object_or_404(Category, pk=cat_id)
                category_form = CategoryForm(request.POST, instance=instance)
                if category_form.is_valid():
                    category_form.save()
                    messages.success(request, 'Category updated.')
                    return redirect('admin_website_settings')
                else:
                    messages.error(request, 'Please correct the errors in category form.')
        elif 'delete_category' in request.POST:
            cat_id = request.POST.get('category_id')
            if not cat_id:
                messages.error(request, 'Missing category ID.')
            else:
                try:
                    Category.objects.get(pk=cat_id).delete()
                    messages.success(request, 'Category deleted.')
                    return redirect('admin_website_settings')
                except Category.DoesNotExist:
                    messages.error(request, 'Category not found.')
        elif 'add_category_media' in request.POST:
            media_form = CategoryMediaForm(request.POST, request.FILES)
            if media_form.is_valid():
                media_form.save()
                messages.success(request, 'Category media uploaded.')
                return redirect('admin_website_settings')
            else:
                messages.error(request, 'Please correct the errors in media form.')
        elif 'add_gallery_images' in request.POST:
            # Multi-upload for Gallery template
            try:
                cat_id = int(request.POST.get('category'))
                category = get_object_or_404(Category, pk=cat_id)
            except (TypeError, ValueError):
                category = None
            images = request.FILES.getlist('images')
            caption = request.POST.get('caption', '')
            try:
                base_order = int(request.POST.get('order', '0'))
            except ValueError:
                base_order = 0
            if not category:
                messages.error(request, 'Invalid category for gallery upload.')
            elif not images:
                messages.error(request, 'Please select one or more images.')
            else:
                created = 0
                for idx, img in enumerate(images):
                    CategoryMedia.objects.create(
                        category=category,
                        kind='image',
                        image=img,
                        caption=caption,
                        order=base_order + idx,
                    )
                    created += 1
                messages.success(request, f'Uploaded {created} image(s) to gallery.')
                return redirect('admin_website_settings')
        elif 'add_single_photo' in request.POST:
            # Single image for Single Photo template
            try:
                cat_id = int(request.POST.get('category'))
                category = get_object_or_404(Category, pk=cat_id)
            except (TypeError, ValueError):
                category = None
            img = request.FILES.get('image')
            caption = request.POST.get('caption', '')
            try:
                order_val = int(request.POST.get('order', '0'))
            except ValueError:
                order_val = 0
            if not category or not img:
                messages.error(request, 'Please select an image to upload.')
            else:
                CategoryMedia.objects.create(
                    category=category,
                    kind='image',
                    image=img,
                    caption=caption,
                    order=order_val,
                )
                messages.success(request, 'Photo uploaded.')
                return redirect('admin_website_settings')
        elif 'add_video_file' in request.POST:
            # Single video file for Video template
            try:
                cat_id = int(request.POST.get('category'))
                category = get_object_or_404(Category, pk=cat_id)
            except (TypeError, ValueError):
                category = None
            vid = request.FILES.get('video_file')
            caption = request.POST.get('caption', '')
            try:
                order_val = int(request.POST.get('order', '0'))
            except ValueError:
                order_val = 0
            if not category or not vid:
                messages.error(request, 'Please select a video file to upload.')
            else:
                CategoryMedia.objects.create(
                    category=category,
                    kind='video',
                    file=vid,
                    caption=caption,
                    order=order_val,
                )
                messages.success(request, 'Video uploaded.')
                return redirect('admin_website_settings')
        elif 'add_pdf_files' in request.POST:
            # Multi-upload for File template (PDF documents)
            try:
                cat_id = int(request.POST.get('category'))
                category = get_object_or_404(Category, pk=cat_id)
            except (TypeError, ValueError):
                category = None
            pdfs = request.FILES.getlist('pdfs')
            caption = request.POST.get('caption', '')
            try:
                base_order = int(request.POST.get('order', '0'))
            except ValueError:
                base_order = 0
            if not category:
                messages.error(request, 'Invalid category for PDF upload.')
            elif not pdfs:
                messages.error(request, 'Please select one or more PDF files.')
            else:
                created = 0
                for idx, f in enumerate(pdfs):
                    name_ok = f.name.lower().endswith('.pdf')
                    content_ok = getattr(f, 'content_type', '') == 'application/pdf'
                    if not (name_ok or content_ok):
                        continue
                    CategoryMedia.objects.create(
                        category=category,
                        kind='file',
                        file=f,
                        caption=caption,
                        order=base_order + idx,
                    )
                    created += 1
                if created:
                    messages.success(request, f'Uploaded {created} PDF file(s).')
                else:
                    messages.error(request, 'No valid PDF files were uploaded.')
                return redirect('admin_website_settings')
        elif 'save_video_side_text' in request.POST:
            # Update only the support text for video template from the preview column
            cat_id = request.POST.get('category_id')
            if not cat_id:
                messages.error(request, 'Missing category ID.')
            else:
                instance = get_object_or_404(Category, pk=cat_id)
                instance.video_side_text = request.POST.get('video_side_text', '')
                instance.save(update_fields=['video_side_text'])
                messages.success(request, 'Support text updated.')
                return redirect('admin_website_settings')
        elif 'delete_media' in request.POST:
            media_id = request.POST.get('media_id')
            try:
                obj = CategoryMedia.objects.get(pk=media_id)
                obj.delete()
                messages.success(request, 'Media deleted.')
            except CategoryMedia.DoesNotExist:
                messages.error(request, 'Media not found.')
            return redirect('admin_website_settings')
        else:
            # Fallback to settings form
            form = SiteSettingsForm(request.POST, request.FILES, instance=settings)
            if form.is_valid():
                form.save()
                messages.success(request, 'Website settings updated successfully.')
                return redirect('admin_website_settings')
            else:
                messages.error(request, 'Please correct the errors below.')
    else:
        form = SiteSettingsForm(instance=settings)
    categories = Category.objects.all()
    media_items = CategoryMedia.objects.select_related('category').all()
    return render(request, 'dashboards/admin_website_settings.html', { 'form': form, 'category_form': category_form, 'media_form': media_form, 'categories': categories, 'media_items': media_items, 'edit_category': edit_category_obj })


@login_required(login_url='login')
@user_passes_test(is_admin)
def admin_gallery(request):
    """Upload and manage gallery images."""
    if request.method == 'POST':
        form = GalleryImageForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Image uploaded to gallery.')
            return redirect('admin_gallery')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = GalleryImageForm()

    images = GalleryImage.objects.all()
    return render(request, 'dashboards/admin_gallery.html', { 'form': form, 'images': images })


@login_required(login_url='login')
@user_passes_test(is_admin)
def admin_categories(request):
    """Add and manage navbar categories."""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category saved.')
            return redirect('admin_categories')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CategoryForm()

    categories = Category.objects.all()
    return render(request, 'dashboards/admin_categories.html', { 'form': form, 'categories': categories })

@login_required(login_url='login')
@user_passes_test(is_admin)
def downloads(request):
    """
    Simple Downloads hub for admins.
    Provides links to download templates and exports already implemented in the system.
    """
    classes = Class.objects.all().order_by('level', 'name')
    users_preview = User.objects.all().order_by('-date_joined')[:10]
    teachers_preview = Teacher.objects.select_related('user', 'department').order_by('id')[:10]
    students_preview = Student.objects.select_related('user', 'class_group').order_by('id')[:10]
    departments = Department.objects.all().order_by('name')
    subjects = Subject.objects.all().order_by('name')
    levels = Class.objects.values_list('level', flat=True).distinct().order_by('level')
    fee_categories = FeeCategory.objects.all().order_by('name')
    terms = Term.objects.select_related('academic_year').all().order_by('-academic_year__year', 'name')
    exams = Exam.objects.select_related('term').all().order_by('-term__academic_year__year', 'term__name', 'name')
    return render(request, 'dashboards/downloads.html', {
        'classes': classes,
        'users_preview': users_preview,
        'teachers_preview': teachers_preview,
        'students_preview': students_preview,
        'users_total': User.objects.count(),
        'teachers_total': Teacher.objects.count(),
        'students_total': Student.objects.count(),
        'departments': departments,
        'subjects': subjects,
        'levels': levels,
        'fee_categories': fee_categories,
        'terms': terms,
        'exams': exams,
    })

# =============================
# CSV/Export Endpoints (Admin)
# =============================

def _csv_response(filename: str):
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def export_students_without_arrears_pdf(request):
    class_id = request.GET.get('class_id')
    term_id = request.GET.get('term_id')
    order = request.GET.get('order', 'asc')
    sort_by = request.GET.get('sort_by', 'name')
    min_balance = request.GET.get('min_balance')
    max_balance = request.GET.get('max_balance')

    students = Student.objects.select_related('user', 'class_group').all()
    if class_id:
        students = students.filter(class_group_id=class_id)

    rows_data = []
    for s in students:
        fa_q = FeeAssignment.objects.filter(class_group=s.class_group)
        if term_id:
            fa_q = fa_q.filter(term_id=term_id)
        owed = fa_q.aggregate(total=Sum('amount'))['total'] or 0
        fp_q = FeePayment.objects.filter(student=s, status='approved')
        if term_id:
            fp_q = fp_q.filter(fee_assignment__term_id=term_id)
        paid = fp_q.aggregate(total=Sum('amount_paid'))['total'] or 0
        balance = (owed - paid)
        if balance <= 0:
            rows_data.append((s, balance, owed, paid))

    try:
        if min_balance not in (None, ''):
            mb = float(min_balance)
            rows_data = [r for r in rows_data if r[1] >= mb]
        if max_balance not in (None, ''):
            xb = float(max_balance)
            rows_data = [r for r in rows_data if r[1] <= xb]
    except ValueError:
        pass

    is_desc = order in ['desc', 'za']
    if sort_by == 'admission':
        rows_data.sort(key=lambda r: (len(r[0].admission_no), r[0].admission_no), reverse=is_desc)
    elif sort_by == 'balance':
        rows_data.sort(key=lambda r: r[1], reverse=is_desc)
    else:
        rows_data.sort(key=lambda r: ((r[0].user.first_name or '') + ' ' + (r[0].user.last_name or '')).strip().lower(), reverse=is_desc)

    columns = ['Admission No', 'Full Name', 'Class', 'Level', 'Owed', 'Paid', 'Balance']
    rows = []
    for s, balance, owed, paid in rows_data:
        u = s.user
        rows.append([
            s.admission_no,
            u.get_full_name() if u else '',
            getattr(s.class_group, 'name', ''),
            getattr(s.class_group, 'level', ''),
            owed,
            paid,
            balance,
        ])
    return pdf_response_from_rows('students_without_arrears.pdf', 'Students without Arrears', _site_header_rows(), columns, rows)

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def export_students_with_arrears_pdf(request):
    class_id = request.GET.get('class_id')
    term_id = request.GET.get('term_id')
    order = request.GET.get('order', 'desc')
    sort_by = request.GET.get('sort_by', 'balance')
    min_balance = request.GET.get('min_balance')
    max_balance = request.GET.get('max_balance')

    students = Student.objects.select_related('user', 'class_group').all()
    if class_id:
        students = students.filter(class_group_id=class_id)

    rows_data = []
    for s in students:
        fa_q = FeeAssignment.objects.filter(class_group=s.class_group)
        if term_id:
            fa_q = fa_q.filter(term_id=term_id)
        owed = fa_q.aggregate(total=Sum('amount'))["total"] or 0
        fp_q = FeePayment.objects.filter(student=s, status='approved')
        if term_id:
            fp_q = fp_q.filter(fee_assignment__term_id=term_id)
        paid = fp_q.aggregate(total=Sum('amount_paid'))["total"] or 0
        balance = (owed - paid)
        if balance > 0:
            rows_data.append((s, balance, owed, paid))

    try:
        if min_balance not in (None, ''):
            mb = float(min_balance)
            rows_data = [r for r in rows_data if r[1] >= mb]
        if max_balance not in (None, ''):
            xb = float(max_balance)
            rows_data = [r for r in rows_data if r[1] <= xb]
    except ValueError:
        pass

    is_desc = order in ['desc', 'za']
    if sort_by == 'name':
        rows_data.sort(key=lambda r: ((r[0].user.first_name or '') + ' ' + (r[0].user.last_name or '')).strip().lower(), reverse=is_desc)
    elif sort_by == 'admission':
        rows_data.sort(key=lambda r: (len(r[0].admission_no), r[0].admission_no), reverse=is_desc)
    else:
        rows_data.sort(key=lambda r: r[1], reverse=is_desc)

    columns = ['Admission No', 'Full Name', 'Class', 'Level', 'Owed', 'Paid', 'Balance']
    rows = []
    for s, balance, owed, paid in rows_data:
        u = s.user
        rows.append([
            s.admission_no,
            u.get_full_name() if u else '',
            getattr(s.class_group, 'name', ''),
            getattr(s.class_group, 'level', ''),
            owed,
            paid,
            balance,
        ])
    return pdf_response_from_rows('students_with_arrears.pdf', 'Students with Arrears', _site_header_rows(), columns, rows)

def _site_header_rows():
    """Return a list of single-cell rows for school header: name, motto, address, and a spacer."""
    rows = []
    try:
        from landing.models import SiteSettings
        ss = SiteSettings.objects.first()
        if ss:
            rows.append([getattr(ss, 'school_name', '')])
            motto = getattr(ss, 'school_motto', '')
            addr = getattr(ss, 'contact_address', '')
            if motto:
                rows.append([motto])
            if addr:
                rows.append([addr])
            rows.append([])
    except Exception:
        pass
    return rows

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def export_students_with_arrears_csv(request):
    class_id = request.GET.get('class_id')
    term_id = request.GET.get('term_id')
    order = request.GET.get('order', 'desc')  # asc/desc
    sort_by = request.GET.get('sort_by', 'balance')  # balance|name|admission
    min_balance = request.GET.get('min_balance')
    max_balance = request.GET.get('max_balance')

    students = Student.objects.select_related('user', 'class_group').all()
    if class_id:
        students = students.filter(class_group_id=class_id)

    # Build balances
    rows = []
    for s in students:
        # Owed: sum of fee assignments for student's class (optionally by term)
        fa_q = FeeAssignment.objects.filter(class_group=s.class_group)
        if term_id:
            fa_q = fa_q.filter(term_id=term_id)
        owed = fa_q.aggregate(total=Sum('amount'))['total'] or 0
        # Paid: sum of approved payments for this student (optionally by term via fee_assignment)
        fp_q = FeePayment.objects.filter(student=s, status='approved')
        if term_id:
            fp_q = fp_q.filter(fee_assignment__term_id=term_id)
        paid = fp_q.aggregate(total=Sum('amount_paid'))['total'] or 0
        balance = (owed - paid)
        if balance > 0:
            rows.append((s, balance, owed, paid))

    # Optional balance range filter
    try:
        if min_balance is not None and min_balance != '':
            mb = float(min_balance)
            rows = [r for r in rows if r[1] >= mb]
        if max_balance is not None and max_balance != '':
            xb = float(max_balance)
            rows = [r for r in rows if r[1] <= xb]
    except ValueError:
        pass

    # Sorting
    is_desc = order in ['desc', 'za']
    if sort_by == 'name':
        rows.sort(key=lambda r: ((r[0].user.first_name or '') + ' ' + (r[0].user.last_name or '')).strip().lower(), reverse=is_desc)
    elif sort_by == 'admission':
        rows.sort(key=lambda r: (len(r[0].admission_no), r[0].admission_no), reverse=is_desc)
    else:  # balance
        rows.sort(key=lambda r: r[1], reverse=is_desc)

    resp = _csv_response('students_with_arrears.csv')
    writer = csv.writer(resp)
    # Prepend school header
    try:
        from landing.models import SiteSettings
        ss = SiteSettings.objects.first()
        if ss:
            writer.writerow([getattr(ss, 'school_name', '')])
            if getattr(ss, 'school_motto', ''):
                writer.writerow([ss.school_motto])
            if getattr(ss, 'contact_address', ''):
                writer.writerow([ss.contact_address])
            writer.writerow([])
    except Exception:
        pass
    writer.writerow(['Admission No', 'Full Name', 'Class', 'Level', 'Owed', 'Paid', 'Balance'])
    for s, balance, owed, paid in rows:
        user = s.user
        writer.writerow([
            s.admission_no,
            user.get_full_name() if user else '',
            getattr(s.class_group, 'name', ''),
            getattr(s.class_group, 'level', ''),
            owed,
            paid,
            balance,
        ])
    return resp

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def export_students_without_arrears_csv(request):
    class_id = request.GET.get('class_id')
    term_id = request.GET.get('term_id')
    order = request.GET.get('order', 'asc')  # asc/desc
    sort_by = request.GET.get('sort_by', 'name')  # name|admission|balance
    min_balance = request.GET.get('min_balance')
    max_balance = request.GET.get('max_balance')

    students = Student.objects.select_related('user', 'class_group').all()
    if class_id:
        students = students.filter(class_group_id=class_id)

    rows = []
    for s in students:
        fa_q = FeeAssignment.objects.filter(class_group=s.class_group)
        if term_id:
            fa_q = fa_q.filter(term_id=term_id)
        owed = fa_q.aggregate(total=Sum('amount'))['total'] or 0
        fp_q = FeePayment.objects.filter(student=s, status='approved')
        if term_id:
            fp_q = fp_q.filter(fee_assignment__term_id=term_id)
        paid = fp_q.aggregate(total=Sum('amount_paid'))['total'] or 0
        balance = (owed - paid)
        if balance <= 0:
            rows.append((s, balance, owed, paid))

    # Optional balance range filter (allow negatives too)
    try:
        if min_balance is not None and min_balance != '':
            mb = float(min_balance)
            rows = [r for r in rows if r[1] >= mb]
        if max_balance is not None and max_balance != '':
            xb = float(max_balance)
            rows = [r for r in rows if r[1] <= xb]
    except ValueError:
        pass

    is_desc = order in ['desc', 'za']
    if sort_by == 'admission':
        rows.sort(key=lambda r: (len(r[0].admission_no), r[0].admission_no), reverse=is_desc)
    elif sort_by == 'balance':
        rows.sort(key=lambda r: r[1], reverse=is_desc)
    else:
        rows.sort(key=lambda r: ((r[0].user.first_name or '') + ' ' + (r[0].user.last_name or '')).strip().lower(), reverse=is_desc)

    resp = _csv_response('students_without_arrears.csv')
    writer = csv.writer(resp)
    # Prepend school header
    try:
        from landing.models import SiteSettings
        ss = SiteSettings.objects.first()
        if ss:
            writer.writerow([getattr(ss, 'school_name', '')])
            if getattr(ss, 'school_motto', ''):
                writer.writerow([ss.school_motto])
            if getattr(ss, 'contact_address', ''):
                writer.writerow([ss.contact_address])
            writer.writerow([])
    except Exception:
        pass
    writer.writerow(['Admission No', 'Full Name', 'Class', 'Level', 'Owed', 'Paid', 'Balance'])
    for s, balance, owed, paid in rows:
        user = s.user
        writer.writerow([
            s.admission_no,
            user.get_full_name() if user else '',
            getattr(s.class_group, 'name', ''),
            getattr(s.class_group, 'level', ''),
            owed,
            paid,
            balance,
        ])
    return resp

@login_required(login_url='login')
@user_passes_test(is_admin)
def export_users_csv(request):
    qs = User.objects.all().order_by('id')
    resp = _csv_response('users.csv')
    writer = csv.writer(resp)
    try:
        from landing.models import SiteSettings
        ss = SiteSettings.objects.first()
        if ss:
            writer.writerow([getattr(ss, 'school_name', '')])
            if getattr(ss, 'school_motto', ''):
                writer.writerow([ss.school_motto])
            if getattr(ss, 'contact_address', ''):
                writer.writerow([ss.contact_address])
            writer.writerow([])
    except Exception:
        pass
    writer.writerow(['ID', 'Username', 'First Name', 'Last Name', 'Email', 'Role', 'Is Active', 'Date Joined'])
    for u in qs:
        writer.writerow([u.id, u.username, u.first_name, u.last_name, u.email, getattr(u, 'role', ''), u.is_active, u.date_joined])
    return resp

@login_required(login_url='login')
@user_passes_test(is_admin)
def export_users_pdf(request):
    qs = User.objects.all().order_by('id')
    columns = ['ID', 'Username', 'First Name', 'Last Name', 'Email', 'Role', 'Is Active', 'Date Joined']
    rows = []
    for u in qs:
        rows.append([
            u.id, u.username, u.first_name, u.last_name, u.email,
            getattr(u, 'role', ''), u.is_active,
            u.date_joined.strftime('%Y-%m-%d %H:%M') if getattr(u, 'date_joined', None) else ''
        ])
    return pdf_response_from_rows('users.pdf', 'Users', _site_header_rows(), columns, rows)

@login_required(login_url='login')
@user_passes_test(is_admin)
def export_teachers_csv(request):
    qs = Teacher.objects.select_related('user', 'department').all()
    # Filters
    dept_id = request.GET.get('department_id')
    subj_id = request.GET.get('subject_id')
    order = request.GET.get('order', 'az')  # az or za
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    if subj_id:
        qs = qs.filter(subjects__id=subj_id)
    # Ordering by teacher name
    if order == 'za':
        qs = qs.order_by('-user__first_name', '-user__last_name', '-id')
    else:
        qs = qs.order_by('user__first_name', 'user__last_name', 'id')
    resp = _csv_response('teachers.csv')
    writer = csv.writer(resp)
    try:
        from landing.models import SiteSettings
        ss = SiteSettings.objects.first()
        if ss:
            writer.writerow([getattr(ss, 'school_name', '')])
            if getattr(ss, 'school_motto', ''):
                writer.writerow([ss.school_motto])
            if getattr(ss, 'contact_address', ''):
                writer.writerow([ss.contact_address])
            writer.writerow([])
    except Exception:
        pass
    writer.writerow(['ID', 'Name', 'Username', 'Department', 'TSC Number', 'Staff ID', 'Phone', 'Gender', 'Email'])
    for t in qs:
        user = t.user
        name = user.get_full_name() if user else ''
        writer.writerow([t.id, name, user.username if user else '', getattr(t.department, 'name', ''), t.tsc_number or '', t.staff_id or '', t.phone or '', t.gender or '', (user.email if user else '')])
    return resp

@login_required(login_url='login')
@user_passes_test(is_admin)
def export_teachers_pdf(request):
    qs = Teacher.objects.select_related('user', 'department').all()
    dept_id = request.GET.get('department_id')
    subj_id = request.GET.get('subject_id')
    order = request.GET.get('order', 'az')
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    if subj_id:
        qs = qs.filter(subjects__id=subj_id)
    if order == 'za':
        qs = qs.order_by('-user__first_name', '-user__last_name', '-id')
    else:
        qs = qs.order_by('user__first_name', 'user__last_name', 'id')
    columns = ['ID', 'Name', 'Username', 'Department', 'TSC Number', 'Staff ID', 'Phone', 'Gender', 'Email']
    rows = []
    for t in qs:
        user = t.user
        name = user.get_full_name() if user else ''
        rows.append([t.id, name, user.username if user else '', getattr(t.department, 'name', ''), t.tsc_number or '', t.staff_id or '', t.phone or '', t.gender or '', (user.email if user else '')])
    return pdf_response_from_rows('teachers.pdf', 'Teachers', _site_header_rows(), columns, rows)

@login_required(login_url='login')
@user_passes_test(is_admin)
def export_students_csv(request):
    qs = Student.objects.select_related('user', 'class_group').all()
    # Filters
    class_id = request.GET.get('class_id')
    level = request.GET.get('level')
    order = request.GET.get('order', 'asc')  # asc/desc or az/za
    sort_by = request.GET.get('sort_by', 'admission')  # name or admission
    if class_id:
        qs = qs.filter(class_group_id=class_id)
    if level:
        qs = qs.filter(class_group__level=level)
    # Ordering by student name
    is_desc = order in ['desc', 'za']
    if sort_by == 'admission':
        # Length-aware lexicographic ordering to handle ADM2 < ADM10 numerically
        qs = qs.annotate(_adm_len=Length('admission_no'))
        if is_desc:
            qs = qs.order_by('-_adm_len', '-admission_no')
        else:
            qs = qs.order_by('_adm_len', 'admission_no')
    else:
        # name
        qs = qs.order_by('-user__first_name', '-user__last_name') if is_desc else qs.order_by('user__first_name', 'user__last_name')
    resp = _csv_response('students.csv')
    writer = csv.writer(resp)
    try:
        from landing.models import SiteSettings
        ss = SiteSettings.objects.first()
        if ss:
            writer.writerow([getattr(ss, 'school_name', '')])
            if getattr(ss, 'school_motto', ''):
                writer.writerow([ss.school_motto])
            if getattr(ss, 'contact_address', ''):
                writer.writerow([ss.contact_address])
            writer.writerow([])
    except Exception:
        pass
    writer.writerow(['Admission No', 'Full Name', 'Username', 'Class', 'Level', 'Gender', 'Phone', 'Graduated'])
    for s in qs:
        user = s.user
        full_name = user.get_full_name() if user else ''
        username = user.username if user else ''
        writer.writerow([s.admission_no, full_name, username, getattr(s.class_group, 'name', ''), getattr(s.class_group, 'level', ''), s.gender, s.phone or '', s.graduated])
    return resp

@login_required(login_url='login')
@user_passes_test(is_admin)
def export_students_pdf(request):
    qs = Student.objects.select_related('user', 'class_group').all()
    class_id = request.GET.get('class_id')
    level = request.GET.get('level')
    order = request.GET.get('order', 'asc')
    sort_by = request.GET.get('sort_by', 'admission')
    if class_id:
        qs = qs.filter(class_group_id=class_id)
    if level:
        qs = qs.filter(class_group__level=level)
    is_desc = order in ['desc', 'za']
    if sort_by == 'admission':
        qs = qs.annotate(_adm_len=Length('admission_no'))
        qs = qs.order_by('-_adm_len', '-admission_no') if is_desc else qs.order_by('_adm_len', 'admission_no')
    else:
        qs = qs.order_by('-user__first_name', '-user__last_name') if is_desc else qs.order_by('user__first_name', 'user__last_name')
    columns = ['Admission No', 'Full Name', 'Username', 'Class', 'Level', 'Gender', 'Phone', 'Graduated']
    rows = []
    for s in qs:
        user = s.user
        full_name = user.get_full_name() if user else ''
        username = user.username if user else ''
        rows.append([s.admission_no, full_name, username, getattr(s.class_group, 'name', ''), getattr(s.class_group, 'level', ''), s.gender, s.phone or '', s.graduated])
    return pdf_response_from_rows('students.pdf', 'Students', _site_header_rows(), columns, rows)

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def export_fee_assignments_csv(request):
    qs = FeeAssignment.objects.select_related('fee_category', 'class_group', 'term__academic_year').all()
    class_id = request.GET.get('class_id')
    term_id = request.GET.get('term_id')
    if class_id:
        qs = qs.filter(class_group_id=class_id)
    if term_id:
        qs = qs.filter(term_id=term_id)
    qs = qs.order_by('class_group__level', 'class_group__name', 'fee_category__name')
    resp = _csv_response('fee_assignments.csv')
    writer = csv.writer(resp)
    try:
        from landing.models import SiteSettings
        ss = SiteSettings.objects.first()
        if ss:
            writer.writerow([getattr(ss, 'school_name', '')])
            if getattr(ss, 'school_motto', ''):
                writer.writerow([ss.school_motto])
            if getattr(ss, 'contact_address', ''):
                writer.writerow([ss.contact_address])
            writer.writerow([])
    except Exception:
        pass
    writer.writerow(['ID', 'Fee Category', 'Class', 'Level', 'Term', 'Academic Year', 'Amount'])
    for fa in qs:
        writer.writerow([
            fa.id,
            fa.fee_category.name if fa.fee_category else '',
            fa.class_group.name if fa.class_group else '',
            fa.class_group.level if fa.class_group else '',
            fa.term.name if fa.term else '',
            fa.term.academic_year.year if getattr(fa.term, 'academic_year', None) else '',
            fa.amount,
        ])
    return resp

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def export_fee_assignments_pdf(request):
    qs = FeeAssignment.objects.select_related('fee_category', 'class_group', 'term__academic_year').all()
    class_id = request.GET.get('class_id')
    term_id = request.GET.get('term_id')
    if class_id:
        qs = qs.filter(class_group_id=class_id)
    if term_id:
        qs = qs.filter(term_id=term_id)
    qs = qs.order_by('class_group__level', 'class_group__name', 'fee_category__name')
    columns = ['ID', 'Fee Category', 'Class', 'Level', 'Term', 'Academic Year', 'Amount']
    rows = []
    for fa in qs:
        rows.append([
            fa.id,
            fa.fee_category.name if fa.fee_category else '',
            fa.class_group.name if fa.class_group else '',
            fa.class_group.level if fa.class_group else '',
            fa.term.name if fa.term else '',
            fa.term.academic_year.year if getattr(fa.term, 'academic_year', None) else '',
            fa.amount,
        ])
    return pdf_response_from_rows('fee_assignments.pdf', 'Fee Assignments', _site_header_rows(), columns, rows)

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def export_fee_payments_csv(request):
    qs = FeePayment.objects.select_related('student__user', 'fee_assignment__fee_category', 'fee_assignment__class_group', 'fee_assignment__term__academic_year').all()
    # Filters
    method = request.GET.get('method')  # mpesa, cash, bank
    category_id = request.GET.get('category_id')
    class_id = request.GET.get('class_id')
    term_id = request.GET.get('term_id')
    if method:
        qs = qs.filter(payment_method__iexact=method)
    if category_id:
        qs = qs.filter(fee_assignment__fee_category_id=category_id)
    if class_id:
        qs = qs.filter(fee_assignment__class_group_id=class_id)
    if term_id:
        qs = qs.filter(fee_assignment__term_id=term_id)
    qs = qs.order_by('-payment_date')
    resp = _csv_response('fee_payments.csv')
    writer = csv.writer(resp)
    try:
        from landing.models import SiteSettings
        ss = SiteSettings.objects.first()
        if ss:
            writer.writerow([getattr(ss, 'school_name', '')])
            if getattr(ss, 'school_motto', ''):
                writer.writerow([ss.school_motto])
            if getattr(ss, 'contact_address', ''):
                writer.writerow([ss.contact_address])
            writer.writerow([])
    except Exception:
        pass
    writer.writerow(['ID', 'Student', 'Admission No', 'Class', 'Level', 'Fee Category', 'Amount Paid', 'Payment Date', 'Method', 'Reference', 'Phone', 'Status'])
    for p in qs:
        student = p.student
        user = getattr(student, 'user', None)
        class_group = getattr(student, 'class_group', None)
        cat = p.fee_assignment.fee_category.name if p.fee_assignment and p.fee_assignment.fee_category else ''
        writer.writerow([
            p.id,
            (user.get_full_name() if user else str(student.id)),
            getattr(student, 'admission_no', ''),
            getattr(class_group, 'name', ''),
            getattr(class_group, 'level', ''),
            cat,
            p.amount_paid,
            p.payment_date.strftime('%Y-%m-%d %H:%M'),
            p.payment_method or '',
            p.reference or '',
            p.phone_number or '',
            p.status,
        ])
    return resp

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def export_fee_payments_pdf(request):
    qs = FeePayment.objects.select_related('student__user', 'fee_assignment__fee_category', 'fee_assignment__class_group', 'fee_assignment__term__academic_year').all()
    method = request.GET.get('method')
    category_id = request.GET.get('category_id')
    class_id = request.GET.get('class_id')
    term_id = request.GET.get('term_id')
    if method:
        qs = qs.filter(payment_method__iexact=method)
    if category_id:
        qs = qs.filter(fee_assignment__fee_category_id=category_id)
    if class_id:
        qs = qs.filter(fee_assignment__class_group_id=class_id)
    if term_id:
        qs = qs.filter(fee_assignment__term_id=term_id)
    qs = qs.order_by('-payment_date')
    columns = ['ID', 'Student', 'Admission No', 'Class', 'Level', 'Fee Category', 'Amount Paid', 'Payment Date', 'Method', 'Reference', 'Phone', 'Status']
    rows = []
    for p in qs:
        student = p.student
        user = getattr(student, 'user', None)
        class_group = getattr(student, 'class_group', None)
        cat = p.fee_assignment.fee_category.name if p.fee_assignment and p.fee_assignment.fee_category else ''
        rows.append([
            p.id,
            (user.get_full_name() if user else str(student.id)),
            getattr(student, 'admission_no', ''),
            getattr(class_group, 'name', ''),
            getattr(class_group, 'level', ''),
            cat,
            p.amount_paid,
            p.payment_date.strftime('%Y-%m-%d %H:%M'),
            p.payment_method or '',
            p.reference or '',
            p.phone_number or '',
            p.status,
        ])
    return pdf_response_from_rows('fee_payments.pdf', 'Fee Payments', _site_header_rows(), columns, rows)

@login_required(login_url='login')
@user_passes_test(is_admin)
def admin_payment_callback_logs(request):
    import os, json
    log_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'payment_callback_logs.txt')
    logs = []
    try:
        with open(log_path, 'r') as logf:
            logs = json.load(logf)
    except Exception:
        logs = []
    return render(request, 'dashboards/admin_payment_callback_logs.html', {'logs': logs})

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def admin_mpesa_reconcile(request):
    """Admin page to reconcile M-Pesa transactions and pending FeePayments."""
    action = request.POST.get('action') if request.method == 'POST' else None
    if action == 'query_tx':
        checkout_id = request.POST.get('checkout_request_id')
        if checkout_id:
            tx = MpesaTransaction.objects.filter(checkout_request_id=checkout_id).first()
            if tx:
                resp = query_stk_status(tx.checkout_request_id)
                if isinstance(resp, dict) and not resp.get('error'):
                    rc = resp.get('ResultCode')
                    rd = resp.get('ResultDesc')
                    if rc is not None:
                        try:
                            tx.result_code = int(rc)
                        except Exception:
                            pass
                    if rd:
                        tx.result_desc = rd
                    if str(rc) == '0':
                        tx.status = 'success'
                    elif str(rc) not in (None, '', '0'):
                        tx.status = 'failed'
                    tx.save()
                    messages.success(request, f"Queried STK status for {checkout_id}: {rd}")
                else:
                    # Avoid dumping raw HTML or oversized payloads into messages
                    err_msg = None
                    if isinstance(resp, dict):
                        err_msg = resp.get('error') or resp.get('ResultDesc') or resp.get('errorMessage')
                    if not err_msg:
                        err_msg = str(resp)
                    # Truncate very long messages
                    if len(err_msg) > 240:
                        err_msg = err_msg[:240] + '…'
                    messages.error(request, f"Query failed for {checkout_id}: {err_msg}")
            else:
                messages.error(request, "Transaction not found")
        return redirect('admin_mpesa_reconcile')
    elif action == 'approve_payment':
        fp_id = request.POST.get('fee_payment_id')
        if fp_id:
            fp = FeePayment.objects.filter(id=fp_id, status='pending').first()
            if fp:
                fp.status = 'approved'
                fp.save(update_fields=['status'])
                messages.success(request, f"Payment {fp.reference or fp.id} approved.")
            else:
                messages.error(request, "Pending payment not found")
        return redirect('admin_mpesa_reconcile')
    elif action == 'link_payment':
        fp_id = request.POST.get('fee_payment_id')
        ref = request.POST.get('reference')
        if fp_id and ref:
            fp = FeePayment.objects.filter(id=fp_id).first()
            tx = MpesaTransaction.objects.filter(mpesa_receipt=ref).first()
            if fp and tx:
                fp.mpesa_transaction = tx
                if not fp.reference:
                    fp.reference = ref
                fp.save(update_fields=['mpesa_transaction','reference'])
                messages.success(request, f"Linked payment {fp.id} to transaction {tx.checkout_request_id}.")
            else:
                messages.error(request, "Could not find payment/transaction to link")
        return redirect('admin_mpesa_reconcile')
    elif action == 'record_external':
        # Manually record a PayBill payment that was not initiated via the site
        from decimal import Decimal
        trans_id = (request.POST.get('trans_id') or '').strip()
        amount = request.POST.get('amount')
        phone = _normalize_msisdn(request.POST.get('phone') or '')
        account_ref = (request.POST.get('account_ref') or '').strip()
        if not trans_id or not amount:
            messages.error(request, 'TransID and Amount are required.')
            return redirect('admin_mpesa_reconcile')
        # Idempotency: if FeePayment already exists by reference, skip
        if FeePayment.objects.filter(reference=trans_id).exists():
            messages.info(request, f'Payment with TransID {trans_id} already recorded.')
            return redirect('admin_mpesa_reconcile')
        # Resolve student from Account#ADM or phone
        student = None
        if account_ref.startswith('Account#'):
            adm = account_ref.split('#', 1)[1].strip()
            student = Student.objects.filter(admission_no=adm).first()
        if not student and phone:
            student = Student.objects.filter(phone=phone).first()
        # Ensure a MpesaTransaction exists
        tx = MpesaTransaction.objects.filter(mpesa_receipt=trans_id).first()
        if not tx:
            try:
                tx = MpesaTransaction.objects.create(
                    student=student,
                    phone_number=phone or '',
                    amount=Decimal(str(amount)),
                    account_reference=account_ref or None,
                    checkout_request_id=f'manual-{trans_id}',
                    status='success',
                    mpesa_receipt=trans_id,
                    result_desc='Manual entry (external PayBill)'
                )
            except Exception as e:
                messages.error(request, f'Error creating transaction: {e}')
                return redirect('admin_mpesa_reconcile')
        # Choose fee assignment (current term or any)
        from django.db.models import Sum as _Sum
        today = timezone.now().date()
        current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()
        outstanding_assignment = None
        if student:
            q = FeeAssignment.objects.filter(class_group=student.class_group)
            if current_term:
                q = q.filter(term=current_term)
            for fa in q:
                paid = FeePayment.objects.filter(student=student, fee_assignment=fa).aggregate(total=_Sum('amount_paid'))['total'] or 0
                if float(paid) < float(fa.amount):
                    outstanding_assignment = fa
                    break
            if not outstanding_assignment:
                outstanding_assignment = q.first()
        # Create approved FeePayment
        try:
            FeePayment.objects.create(
                student=student,
                fee_assignment=outstanding_assignment,
                amount_paid=Decimal(str(amount)),
                payment_method='mpesa',
                reference=trans_id,
                phone_number=phone,
                status='approved',
                mpesa_transaction=tx
            )
            messages.success(request, f'Recorded external payment {trans_id}.')
        except Exception as e:
            messages.error(request, f'Error creating FeePayment: {e}')
        return redirect('admin_mpesa_reconcile')

    pending_txs = MpesaTransaction.objects.filter(status='pending').order_by('-created_at')[:100]
    pending_payments = FeePayment.objects.filter(status='pending').order_by('-payment_date')[:100]
    return render(request, 'admin/mpesa_reconcile.html', {
        'pending_txs': pending_txs,
        'pending_payments': pending_payments,
    })

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def admin_c2b_ledger(request):
    """List PayBill (C2B) ledger entries for admins/clerks with basic search.

    Supports GET filters: q (search in trans_id, msisdn, bill_ref), since (ISO date), until (ISO date).
    """
    q = request.GET.get('q', '').strip()
    since = request.GET.get('since', '').strip()
    until = request.GET.get('until', '').strip()

    entries = MpesaC2BLedger.objects.all().order_by('-created_at')
    if q:
        from django.db.models import Q, Case, When, Value, IntegerField
        entries = entries.filter(
            Q(trans_id__icontains=q) |
            Q(msisdn__icontains=q) |
            Q(bill_ref__icontains=q)
        )
    if since:
        try:
            from datetime import datetime
            dt = datetime.fromisoformat(since)
            entries = entries.filter(created_at__gte=dt)
        except Exception:
            pass
    if until:
        try:
            from datetime import datetime
            dt = datetime.fromisoformat(until)
            entries = entries.filter(created_at__lte=dt)
        except Exception:
            pass

    entries = entries[:200]
    return render(request, 'admin/mpesa_ledger.html', {
        'entries': entries,
        'q': q,
        'since': since,
        'until': until,
    })

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def admin_mpesa_all(request):
    """Admin/Clerk: View all M-Pesa transactions (C2B ledger and STK push).

    GET filters: q (TransID/MSISDN/BillRef), since (ISO), until (ISO)
    """
    q = request.GET.get('q', '').strip()
    since = request.GET.get('since', '').strip()
    until = request.GET.get('until', '').strip()

    from django.db.models import Q
    from datetime import datetime

    c2b = MpesaC2BLedger.objects.all().order_by('-created_at')
    if q:
        c2b = c2b.filter(Q(trans_id__icontains=q) | Q(msisdn__icontains=q) | Q(bill_ref__icontains=q))
    if since:
        try:
            c2b = c2b.filter(created_at__gte=datetime.fromisoformat(since))
        except Exception:
            pass
    if until:
        try:
            c2b = c2b.filter(created_at__lte=datetime.fromisoformat(until))
        except Exception:
            pass
    c2b = c2b[:300]

    stk = MpesaTransaction.objects.all().order_by('-created_at')
    if q:
        stk = stk.filter(
            Q(merchant_request_id__icontains=q) |
            Q(checkout_request_id__icontains=q) |
            Q(phone_number__icontains=q) |
            Q(account_reference__icontains=q) |
            Q(mpesa_receipt__icontains=q)
        )
    if since:
        try:
            stk = stk.filter(created_at__gte=datetime.fromisoformat(since))
        except Exception:
            pass
    if until:
        try:
            stk = stk.filter(created_at__lte=datetime.fromisoformat(until))
        except Exception:
            pass
    stk = stk[:300]

    return render(request, 'admin/mpesa.html', {
        'q': q,
        'since': since,
        'until': until,
        'c2b': c2b,
        'stk': stk,
    })

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def admin_payment_log_file(request):
    """Serve the raw payment callback log file to admin/clerk."""
    log_path = os.path.join(settings.BASE_DIR, 'payment_callback_logs.txt')
    if not os.path.exists(log_path):
        raise Http404('Log file not found')
    try:
        f = open(log_path, 'rb')
    except OSError:
        raise Http404('Unable to open log file')
    response = FileResponse(f, content_type='text/plain')
    response["Content-Disposition"] = 'inline; filename="payment_callback_logs.txt"'
    return response

from django.http import JsonResponse

@login_required(login_url='login')
@user_passes_test(is_admin)
def export_result_slip_csv(request):
    """
    Block-style result slip grouped by class.
    Each class renders a table: rows are students, columns are subjects, cell is grade letter (or score), and last column is Total Points.
    GET: exam_id (required), class_id (optional), level (optional)
    """
    exam_id = request.GET.get('exam_id')
    class_id = request.GET.get('class_id')
    level = request.GET.get('level')
    if not exam_id:
        return HttpResponse("Missing exam_id", status=400)
    # Pull all grades matching filters
    grades_qs = Grade.objects.select_related(
        'student__user', 'student__class_group', 'subject', 'exam__term__academic_year'
    ).filter(exam_id=exam_id)
    if class_id:
        grades_qs = grades_qs.filter(student__class_group_id=class_id)
    if level:
        grades_qs = grades_qs.filter(student__class_group__level=level)

    # Organize by class -> student -> subject
    from collections import defaultdict, OrderedDict
    class_students = defaultdict(OrderedDict)  # {Class: OrderedDict{student_id: Student}}
    class_subjects = defaultdict(set)          # {Class: set(Subject)}
    grade_map = {}                             # {(student_id, subject_id): Grade}

    for g in grades_qs:
        cls = g.student.class_group
        if not cls:
            continue
        class_subjects[cls].add(g.subject)
        if g.student.id not in class_students[cls]:
            class_students[cls][g.student.id] = g.student
        grade_map[(g.student.id, g.subject.id)] = g

    # 4-point scale mapping (including +/- variants)
    LETTER_POINTS = {
        'A+': 4, 'A': 4, 'A-': 4,
        'B+': 3, 'B': 3, 'B-': 3,
        'C+': 2, 'C': 2, 'C-': 2,
        'D+': 1, 'D': 1, 'D-': 1,
        'E': 0,
    }

    # Map average points back to an approximate mean grade
    def points_to_letter(avg: float) -> str:
        if avg is None:
            return ''
        if avg >= 3.5: return 'A'
        if avg >= 2.5: return 'B'
        if avg >= 1.5: return 'C'
        if avg >= 0.5: return 'D'
        return 'E'

    # Helper to compute letter possibly from grading scheme
    def compute_letter(g: Grade):
        if g.grade_letter:
            return g.grade_letter
        try:
            scheme = getattr(g.subject, 'grading_scheme', None)
            if scheme:
                return scheme.get_grade_letter(g.score) or ''
        except Exception:
            pass
        return ''

    # Build CSV
    resp = _csv_response('result_slip_block.csv')
    writer = csv.writer(resp)

    # Header block
    try:
        from landing.models import SiteSettings
        ss = SiteSettings.objects.first()
        if ss:
            writer.writerow([getattr(ss, 'school_name', '')])
            if getattr(ss, 'school_motto', ''):
                writer.writerow([ss.school_motto])
            if getattr(ss, 'contact_address', ''):
                writer.writerow([ss.contact_address])
            writer.writerow([])
    except Exception:
        pass

    # Stable class order
    classes_sorted = sorted(class_students.keys(), key=lambda c: (c.level, c.name))

    # If no classes (no grades), return empty table header
    if not classes_sorted:
        writer.writerow(['No data for the selected exam/filters'])
        return resp

    for cls in classes_sorted:
        # Class title row
        writer.writerow([f"Class: {cls.name} (Level {cls.level})"])
        # Column headers
        subjects_sorted = sorted(class_subjects[cls], key=lambda s: s.name.lower())
        columns = ['Position', 'Admission No', 'Full Name'] + [s.code for s in subjects_sorted] + ['Total Marks', 'Total Points', 'Avg Points', 'Mean Grade']
        writer.writerow(columns)

        # Prepare student metrics
        students_sorted = sorted(class_students[cls].values(), key=lambda s: (len(s.admission_no or ''), s.admission_no or ''))
        student_rows = []
        per_subject_points = {subj.id: [] for subj in subjects_sorted}
        for stu in students_sorted:
            u = stu.user
            row_prefix = [stu.admission_no, u.get_full_name() if u else '']
            total_points = 0
            taken = 0
            total_marks = 0.0
            subj_cells = []
            for subj in subjects_sorted:
                g = grade_map.get((stu.id, subj.id))
                if not g:
                    subj_cells.append('')
                    continue
                letter = compute_letter(g)
                # subject cell shows marks and letter when available
                if letter and g.score is not None:
                    cell = f"{g.score} ({letter})"
                elif g.score is not None:
                    cell = str(g.score)
                else:
                    cell = letter
                subj_cells.append(cell)
                if letter in LETTER_POINTS:
                    total_points += LETTER_POINTS[letter]
                    per_subject_points[subj.id].append(LETTER_POINTS[letter])
                    taken += 1
                # accumulate raw marks for tie-breaker
                try:
                    total_marks += float(g.score or 0)
                except Exception:
                    pass
            avg_points = round(total_points / taken, 2) if taken else 0
            mean_grade = points_to_letter(avg_points) if taken else ''
            # Store for ranking; position computed after sorting by total_points desc
            student_rows.append({
                'stu': stu,
                'row': row_prefix + subj_cells + [round(total_marks, 2), total_points, avg_points, mean_grade],
                'total_points': total_points,
                'avg_points': avg_points,
                'total_marks': total_marks,
            })

        # Rank by Total Points desc, then total Marks desc
        student_rows.sort(key=lambda r: (-r['total_points'], -r.get('total_marks', 0), len(r['stu'].admission_no or ''), r['stu'].admission_no or ''))
        position = 0
        last_points = None
        last_marks = None
        for idx, entry in enumerate(student_rows, start=1):
            pts = entry['total_points']
            marks = entry.get('total_marks', 0)
            if pts != last_points or marks != last_marks:
                position = idx
                last_points = pts
                last_marks = marks
            writer.writerow([position] + entry['row'])

        # Class averages row at bottom
        avg_row_prefix = ['', '', 'Class Averages']
        for subj in subjects_sorted:
            pts_list = per_subject_points.get(subj.id, [])
            if pts_list:
                subj_avg = round(sum(pts_list) / len(pts_list), 2)
                avg_row_prefix.append(f"{subj_avg} ({points_to_letter(subj_avg)})")
            else:
                avg_row_prefix.append('')
        # Overall class averages
        all_totals = [r['total_points'] for r in student_rows]
        all_avgs = [r['avg_points'] for r in student_rows if r['avg_points']]
        all_marks = [r['total_marks'] for r in student_rows]
        class_total_marks_avg = round(sum(all_marks) / len(all_marks), 2) if all_marks else 0
        class_total_avg = round(sum(all_totals) / len(all_totals), 2) if all_totals else 0
        class_avg_points = round(sum(all_avgs) / len(all_avgs), 2) if all_avgs else 0
        class_mean_grade = points_to_letter(class_avg_points) if all_avgs else ''
        writer.writerow(avg_row_prefix + [class_total_marks_avg, class_total_avg, class_avg_points, class_mean_grade])

        writer.writerow([])  # spacer between classes

    return resp

@login_required(login_url='login')
@user_passes_test(is_admin)
def export_result_slip_pdf(request):
    """PDF export in Block format: grouped by class with subject columns and Total Points."""
    exam_id = request.GET.get('exam_id')
    class_id = request.GET.get('class_id')
    level = request.GET.get('level')
    if not exam_id:
        return HttpResponse("Missing exam_id", status=400)

    # Fetch exam for title
    try:
        exam_obj = Exam.objects.select_related('term__academic_year').get(id=exam_id)
    except Exam.DoesNotExist:
        return HttpResponse("Invalid exam_id", status=400)

    grades_qs = Grade.objects.select_related(
        'student__user', 'student__class_group', 'subject', 'exam__term__academic_year'
    ).filter(exam_id=exam_id)
    if class_id:
        grades_qs = grades_qs.filter(student__class_group_id=class_id)
    if level:
        grades_qs = grades_qs.filter(student__class_group__level=level)

    from collections import defaultdict, OrderedDict
    class_students = defaultdict(OrderedDict)
    class_subjects = defaultdict(set)
    grade_map = {}

    for g in grades_qs:
        cls = g.student.class_group
        if not cls:
            continue
        class_subjects[cls].add(g.subject)
        if g.student.id not in class_students[cls]:
            class_students[cls][g.student.id] = g.student
        grade_map[(g.student.id, g.subject.id)] = g

    # 4-point scale mapping (including +/- variants)
    LETTER_POINTS = {
        'A+': 4, 'A': 4, 'A-': 4,
        'B+': 3, 'B': 3, 'B-': 3,
        'C+': 2, 'C': 2, 'C-': 2,
        'D+': 1, 'D': 1, 'D-': 1,
        'E': 0,
    }

    # Map average points back to an approximate mean grade (4-point bands)
    def points_to_letter(avg: float) -> str:
        if avg is None:
            return ''
        if avg >= 3.5: return 'A'
        if avg >= 2.5: return 'B'
        if avg >= 1.5: return 'C'
        if avg >= 0.5: return 'D'
        return 'E'

    def compute_letter(g: Grade):
        if g.grade_letter:
            return g.grade_letter
        try:
            scheme = getattr(g.subject, 'grading_scheme', None)
            if scheme:
                return scheme.get_grade_letter(g.score) or ''
        except Exception:
            pass
        return ''

    # Build PDF using reportlab directly to allow multiple tables per document
    try:
        from reportlab.lib.pagesizes import A4, landscape  # type: ignore
        from reportlab.lib import colors  # type: ignore
        from reportlab.lib.styles import getSampleStyleSheet  # type: ignore
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak  # type: ignore
    except Exception:
        return HttpResponse(
            "PDF generation is unavailable: 'reportlab' is not installed. Please install it (pip install reportlab).",
            content_type='text/plain',
            status=500,
        )

    buffer = BytesIO()
    pagesize = landscape(A4)
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = []

    # School header
    for line in _site_header_rows() or []:
        if not line:
            story.append(Spacer(1, 6))
        else:
            story.append(Paragraph(str(line[0]), styles['Heading4']))
    if _site_header_rows():
        story.append(Spacer(1, 6))

    # Global title
    term = exam_obj.term
    title = f"Block Result Slip — {exam_obj.name} — {term.name if term else ''} {term.academic_year.year if getattr(term, 'academic_year', None) else ''}"
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 10))

    classes_sorted = sorted(class_students.keys(), key=lambda c: (c.level, c.name))
    if not classes_sorted:
        story.append(Paragraph("No data for the selected exam/filters", styles['Normal']))
    else:
        for idx, cls in enumerate(classes_sorted):
            # Class subtitle
            story.append(Paragraph(f"Class: {cls.name} (Level {cls.level})", styles['Heading3']))
            story.append(Spacer(1, 6))

            subjects_sorted = sorted(class_subjects[cls], key=lambda s: s.name.lower())
            columns = ['Position', 'Admission No', 'Full Name'] + [s.code for s in subjects_sorted] + ['Total Marks', 'Total Points', 'Avg Points', 'Mean Grade']

            # Build rows with metrics
            rows = []
            students_sorted = sorted(class_students[cls].values(), key=lambda s: (len(s.admission_no or ''), s.admission_no or ''))
            per_subject_points = {subj.id: [] for subj in subjects_sorted}
            student_rows = []
            for stu in students_sorted:
                u = stu.user
                row_prefix = [stu.admission_no, u.get_full_name() if u else '']
                total_points = 0
                taken = 0
                total_marks = 0.0
                subj_cells = []
                for subj in subjects_sorted:
                    g = grade_map.get((stu.id, subj.id))
                    if not g:
                        subj_cells.append('')
                        continue
                    letter = compute_letter(g)
                    # subject cell shows marks and letter when available
                    if letter and getattr(g, 'score', None) is not None:
                        cell = f"{g.score} ({letter})"
                    elif getattr(g, 'score', None) is not None:
                        cell = str(g.score)
                    else:
                        cell = letter
                    subj_cells.append(cell)
                    if letter in LETTER_POINTS:
                        total_points += LETTER_POINTS[letter]
                        per_subject_points[subj.id].append(LETTER_POINTS[letter])
                        taken += 1
                    # accumulate raw marks for tie-breaker
                    try:
                        total_marks += float(g.score or 0)
                    except Exception:
                        pass
                avg_points = round(total_points / taken, 2) if taken else 0
                mean_grade = points_to_letter(avg_points) if taken else ''
                student_rows.append({
                    'stu': stu,
                    'row': row_prefix + subj_cells + [round(total_marks, 2), total_points, avg_points, mean_grade],
                    'total_points': total_points,
                    'avg_points': avg_points,
                    'total_marks': total_marks,
                })

            # Rank by Total Points desc, then total Marks desc and write with Position
            student_rows.sort(key=lambda r: (-r['total_points'], -r.get('total_marks', 0), len(r['stu'].admission_no or ''), r['stu'].admission_no or ''))
            position = 0
            last_points = None
            last_marks = None
            for idx, entry in enumerate(student_rows, start=1):
                pts = entry['total_points']
                marks = entry.get('total_marks', 0)
                if pts != last_points or marks != last_marks:
                    position = idx
                    last_points = pts
                    last_marks = marks
                rows.append([position] + entry['row'])

            # Class averages row (prefix blanks for Position and Admission No)
            avg_row = ['', '', 'Class Averages']
            for subj in subjects_sorted:
                pts_list = per_subject_points.get(subj.id, [])
                if pts_list:
                    subj_avg = round(sum(pts_list) / len(pts_list), 2)
                    avg_row.append(f"{subj_avg} ({points_to_letter(subj_avg)})")
                else:
                    avg_row.append('')
            all_totals = [r['total_points'] for r in student_rows]
            all_avgs = [r['avg_points'] for r in student_rows if r['avg_points']]
            all_marks = [r['total_marks'] for r in student_rows]
            class_total_marks_avg = round(sum(all_marks) / len(all_marks), 2) if all_marks else 0
            class_total_avg = round(sum(all_totals) / len(all_totals), 2) if all_totals else 0
            class_avg_points = round(sum(all_avgs) / len(all_avgs), 2) if all_avgs else 0
            class_mean_grade = points_to_letter(class_avg_points) if all_avgs else ''
            rows.append(avg_row + [class_total_marks_avg, class_total_avg, class_avg_points, class_mean_grade])

            data = [columns] + rows
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafafa')]),
            ]))

            story.append(table)

            # Page break between classes to avoid very wide stories
            if idx < len(classes_sorted) - 1:
                story.append(PageBreak())

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    filename = 'result_slip_block.pdf'
    resp = HttpResponse(content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    resp.write(pdf)
    return resp

from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
@login_required
@require_GET
def job_status(request, job_id):
    """Return JSON status for a `NotificationJob` for quick UI polling."""
    from .models import NotificationJob
    job = get_object_or_404(NotificationJob, pk=job_id)
    return JsonResponse({
        'id': job.id,
        'status': job.status,
        'total': job.total,
        'processed': job.processed,
        'success_count': job.success_count,
        'error_count': job.error_count,
        'meta': job.meta,
        'created_at': job.created_at.isoformat(),
        'updated_at': job.updated_at.isoformat(),
    })

@login_required(login_url='login')
@user_passes_test(is_admin)
def export_empty_subject_list_csv(request):
    """
    Empty mark entry template for a chosen class and subject.
    GET: class_id (required), subject_id (required)
    """
    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    if not class_id or not subject_id:
        return HttpResponse("Missing class_id or subject_id", status=400)
    try:
        cls = Class.objects.get(id=class_id)
        subj = Subject.objects.get(id=subject_id)
    except Class.DoesNotExist:
        return HttpResponse("Invalid class_id", status=400)
    except Subject.DoesNotExist:
        return HttpResponse("Invalid subject_id", status=400)
    students = Student.objects.select_related('user').filter(class_group=cls).order_by('user__first_name', 'user__last_name')
    resp = _csv_response('empty_subject_list.csv')
    writer = csv.writer(resp)
    try:
        from landing.models import SiteSettings
        ss = SiteSettings.objects.first()
        if ss:
            writer.writerow([getattr(ss, 'school_name', '')])
            if getattr(ss, 'school_motto', ''):
                writer.writerow([ss.school_motto])
            if getattr(ss, 'contact_address', ''):
                writer.writerow([ss.contact_address])
            writer.writerow([])
    except Exception:
        pass
    writer.writerow(['Admission No', 'Full Name', 'Class', 'Level', 'Subject', 'Score'])
    for s in students:
        user = s.user
        writer.writerow([
            s.admission_no,
            user.get_full_name() if user else '',
            getattr(cls, 'name', ''),
            getattr(cls, 'level', ''),
            subj.name,
            '',
        ])
    return resp

@login_required(login_url='login')
@user_passes_test(is_admin)
def export_empty_subject_list_pdf(request):
    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    if not class_id or not subject_id:
        return HttpResponse("Missing class_id or subject_id", status=400)
    try:
        cls = Class.objects.get(id=class_id)
        subj = Subject.objects.get(id=subject_id)
    except Class.DoesNotExist:
        return HttpResponse("Invalid class_id", status=400)
    except Subject.DoesNotExist:
        return HttpResponse("Invalid subject_id", status=400)
    students = Student.objects.select_related('user').filter(class_group=cls).order_by('user__first_name', 'user__last_name')
    columns = ['Admission No', 'Full Name', 'Class', 'Level', 'Subject', 'Score']
    rows = []
    for s in students:
        user = s.user
        rows.append([
            s.admission_no,
            user.get_full_name() if user else '',
            getattr(cls, 'name', ''),
            getattr(cls, 'level', ''),
            subj.name,
            '',
        ])
    return pdf_response_from_rows('empty_subject_list.pdf', f'Empty Subject List — {subj.name}', _site_header_rows(), columns, rows)

@login_required(login_url='login')
def manage_class_subjects(request, class_id):
    if not request.user.role == 'admin':
        return HttpResponseForbidden("You are not authorized to manage class subjects.")
    class_obj = get_object_or_404(Class, id=class_id)
    all_subjects = Subject.objects.all().order_by('name')
    if request.method == 'POST':
        selected_subject_ids = request.POST.getlist('subjects')
        class_obj.subjects.set(selected_subject_ids)
        messages.success(request, 'Subjects updated successfully!')
        return redirect('class_profile', class_id=class_obj.id)
    return render(request, 'dashboards/manage_class_subjects.html', {
        'class_obj': class_obj,
        'all_subjects': all_subjects,
    })

@login_required(login_url='login')
def assign_subject_teachers(request, class_id):
    """
    Admin UI to assign/reassign a teacher for each subject in the given class.
    Uses the TeacherClassAssignment model as the source of truth.
    """
    if not request.user.role == 'admin':
        return HttpResponseForbidden("You are not authorized to assign subject teachers.")
    class_obj = get_object_or_404(Class, id=class_id)

    # Subjects already attached to this class
    subjects = list(class_obj.subjects.all().order_by('name'))

    # Existing assignments for quick lookup: {(subject_id): teacher_id}
    existing = {
        a.subject_id: a.teacher_id
        for a in TeacherClassAssignment.objects.filter(class_group=class_obj)
    }

    # Build per-subject teacher choices: prefer teachers who teach that subject
    per_subject_teachers = {}
    for subj in subjects:
        per_subject_teachers[subj.id] = list(
            Teacher.objects.filter(subjects=subj)
            .select_related('user')
            .order_by('user__first_name', 'user__last_name', 'user__username')
        )

    if request.method == 'POST':
        # For each subject, read teacher selection input name: teacher_for_<subject_id>
        updates = 0
        for subj in subjects:
            key = f'teacher_for_{subj.id}'
            val = request.POST.get(key, '')
            teacher_id = int(val) if val.isdigit() else None

            current_teacher_id = existing.get(subj.id)
            # If no selection and there was an assignment -> delete it
            if not teacher_id and current_teacher_id:
                TeacherClassAssignment.objects.filter(class_group=class_obj, subject=subj).delete()
                updates += 1
            # If selected and changed -> upsert
            elif teacher_id and teacher_id != current_teacher_id:
                TeacherClassAssignment.objects.update_or_create(
                    class_group=class_obj,
                    subject=subj,
                    defaults={'teacher_id': teacher_id}
                )
                updates += 1
        if updates:
            messages.success(request, f"Saved {updates} assignment change(s).")
        else:
            messages.info(request, "No changes to save.")
        return redirect('assign_subject_teachers', class_id=class_obj.id)

    # Precompute rows for easy template rendering without custom filters
    rows = []
    for subj in subjects:
        rows.append({
            'subject': subj,
            'teachers': per_subject_teachers.get(subj.id, []),
            'selected_id': existing.get(subj.id),
        })

    context = {
        'class_obj': class_obj,
        'rows': rows,
    }
    return render(request, 'dashboards/assign_subject_teachers.html', context)

@login_required(login_url='login')
def manage_subject_grading(request, subject_id):
    from .models import Subject, SubjectGradingScheme
    subject = get_object_or_404(Subject, id=subject_id)
    error = None
    grading_scheme = getattr(subject, 'grading_scheme', None)
    grade_boundaries_json = ''
    last_updated = None
    updated_by = None
    if grading_scheme:
        import json
        grade_boundaries_json = json.dumps(grading_scheme.grade_boundaries, indent=2)
        last_updated = grading_scheme.updated_at
        updated_by = grading_scheme.updated_by
    if request.method == 'POST':
        import json
        try:
            boundaries = json.loads(request.POST.get('grade_boundaries', '{}'))
            # Validate boundaries format
            for k, v in boundaries.items():
                if not (isinstance(v, list) and len(v) == 2 and all(isinstance(x, (int, float)) for x in v)):
                    raise ValueError(f"Invalid boundary for {k}: {v}")
            if grading_scheme:
                grading_scheme.grade_boundaries = boundaries
                grading_scheme.updated_by = request.user
                grading_scheme.save()
            else:
                grading_scheme = SubjectGradingScheme.objects.create(
                    subject=subject,
                    grade_boundaries=boundaries,
                    updated_by=request.user
                )
            messages.success(request, 'Grading scheme updated successfully!')
            return redirect('admin_subjects')
        except Exception as e:
            error = f"Invalid input: {e}"
            grade_boundaries_json = request.POST.get('grade_boundaries', '')
    return render(request, 'dashboards/manage_subject_grading.html', {
        'subject': subject,
        'grade_boundaries_json': grade_boundaries_json,
        'error': error,
        'last_updated': last_updated,
        'updated_by': updated_by,
    })

@login_required(login_url='login')
def dashboard(request):
    user = request.user
    if user.role == 'admin':
        return redirect('admin_overview')
    elif user.role == 'clerk':
        return redirect('clerk_overview')
    elif user.role == 'teacher':
        teacher = get_object_or_404(Teacher, user=user)
        return redirect('teacher_dashboard', teacher_id=teacher.id)
    elif user.role == 'student':
        return redirect('student_profile', student_id=user.student.id)
    else:
        messages.error(request, "Invalid user role.")
        return redirect('login')


@login_required(login_url='login')
@user_passes_test(lambda u: u.role == 'admin', login_url='login')
def view_attendance(request):
    status_filter = request.GET.get('status', '')
    class_filter = request.GET.get('class_id', '')

    # Defer 'timestamp' to avoid selecting a non-existent DB column if migrations haven't been applied yet
    attendances = Attendance.objects.all().defer('timestamp').select_related(
        'student__user', 'subject', 'teacher__user', 'class_group'
    ).order_by('-date')

    if status_filter:
        attendances = attendances.filter(status=status_filter)
    
    if class_filter:
        attendances = attendances.filter(class_group__id=class_filter)

    classes = Class.objects.all()
    status_choices = Attendance.STATUS_CHOICES

    context = {
        'attendances': attendances,
        'classes': classes,
        'status_choices': status_choices,
        'current_status': status_filter,
        'current_class': int(class_filter) if class_filter else None
    }
    return render(request, 'attendance/view_attendance.html', context)


def can_view_results_for_student(request, student):
    """
    Centralized helper to decide if results access should be granted for a student
    based on fee arrears and SiteSettings. Admins and the student's class teacher bypass.

    Returns: (can_view: bool, message: str, restrict_results_by_fee: bool)
    """
    can_view_results = True
    fee_restriction_message = ""
    restrict_results_by_fee = False

    try:
        from landing.models import SiteSettings
        from .models import FeeAssignment, FeePayment, Teacher
        from django.db.models import Sum
        from django.core.cache import cache

        settings_obj = SiteSettings.objects.first()
        # Global toggle from settings page
        if settings_obj and settings_obj.restrict_results_by_fee:
            restrict_results_by_fee = True

        # Role bypasses
        is_admin = getattr(request.user, 'role', None) == 'admin'
        is_teacher = getattr(request.user, 'role', None) == 'teacher'
        is_class_teacher = False
        if is_teacher:
            try:
                teacher = Teacher.objects.filter(user=request.user).first()
                if teacher and student.class_group and student.class_group.class_teacher_id == teacher.id:
                    is_class_teacher = True
            except Exception:
                is_class_teacher = False

        if not (is_admin or is_class_teacher):
            total_billed = (
                FeeAssignment.objects.filter(class_group=student.class_group).aggregate(total=Sum('amount'))['total']
                or 0
            )
            total_paid = (
                FeePayment.objects.filter(student=student).aggregate(total=Sum('amount_paid'))['total']
                or 0
            )
            billed = float(total_billed or 0)
            paid = float(total_paid or 0)
            balance_amount = max(billed - paid, 0.0)
            balance_percent = (balance_amount / billed * 100.0) if billed > 0 else 0.0

            # 1) Enforce admin-activated absolute amount threshold from cache (immediate runtime control)
            bar_enabled = cache.get('results_bar_enabled')
            bar_threshold = cache.get('results_bar_balance_min')
            if bar_enabled and bar_threshold is not None:
                try:
                    amt_threshold = float(bar_threshold)
                except (TypeError, ValueError):
                    amt_threshold = None
                if amt_threshold is not None and balance_amount >= amt_threshold:
                    can_view_results = False
                    restrict_results_by_fee = True  # reflect active restriction for UI
                    fee_restriction_message = (
                        (settings_obj.fee_restriction_message if settings_obj else None)
                        or "Results are temporarily unavailable due to outstanding fees."
                    )
                    return can_view_results, fee_restriction_message, restrict_results_by_fee

            # 2) Fallback to percentage-based restriction from SiteSettings (if enabled)
            if settings_obj and settings_obj.restrict_results_by_fee:
                threshold = float(settings_obj.fee_restriction_threshold or 0)
                if balance_percent > threshold:
                    can_view_results = False
                    fee_restriction_message = (
                        settings_obj.fee_restriction_message
                        or "Results are temporarily unavailable due to outstanding fees."
                    )
    except Exception:
        # Fail-open for robustness; UI still provides hints
        pass

    return can_view_results, fee_restriction_message, restrict_results_by_fee


@login_required(login_url='login')
def student_results_blocked(request):
    """Dedicated page to show when a student's results are blocked by fees."""
    message = request.GET.get('msg') or "Results are temporarily unavailable due to outstanding fees."
    return render(request, 'dashboards/results_blocked.html', {
        'message': message,
    })


@login_required(login_url='login')
def student_profile(request, student_id):
    from django.utils import timezone
    student = get_object_or_404(Student, id=student_id)
    print(f"[DEBUG] Viewing profile for student ID: {student_id}, admission_no: {student.admission_no}, user: {student.user}, full_name: {getattr(student.user, 'get_full_name', lambda: student.user.username)() if student.user else 'NO USER'}")
    # Authorization: students can only view their own profile
    try:
        user_role = getattr(request.user, 'role', None)
        if user_role == 'student':
            # Find the logged-in student's record
            own_student = Student.objects.filter(user=request.user).first()
            if not own_student:
                from django.http import HttpResponseForbidden
                return HttpResponseForbidden('Student record not linked to your account.')
            if own_student.id != student.id:
                messages.warning(request, 'You can only view your own profile.')
                return redirect('student_profile', student_id=own_student.id)
    except Exception as e:
        # Fail safe: in case of unexpected issues, do not expose other students
        try:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('Not allowed.')
        except Exception:
            return redirect('student_dashboard')
    now = timezone.now()

    today = timezone.now().date()
    current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()

    # All fee assignments for the student's class (all time)
    fee_assignments = FeeAssignment.objects.filter(class_group=student.class_group)
    from django.db.models import Sum, Avg
    total_billed = fee_assignments.aggregate(total=Sum('amount'))['total'] or 0

    # All payments made by this student (all time)
    total_paid = FeePayment.objects.filter(student=student).aggregate(total=Sum('amount_paid'))['total'] or 0

    balance = total_billed - total_paid

    # --- Results access restriction logic (fee-based) ---
    can_view_results, fee_restriction_message, restrict_results_by_fee = can_view_results_for_student(request, student)

    # --- Exam selection and Performance/Grades Section ---
    # Build the base grades queryset
    grades_qs = Grade.objects.filter(student=student).select_related('subject', 'exam')

    # Exams this student has grades in
    student_exam_ids = list(grades_qs.values_list('exam_id', flat=True).distinct())
    student_exams = Exam.objects.filter(id__in=student_exam_ids).select_related('term').order_by('-start_date', '-id')
    # Only published exams are allowed for printing/downloading
    published_exams = student_exams.filter(results_published=True)

    selected_exam = None
    exam_id = request.GET.get('exam_id')
    if exam_id:
        try:
            ex = Exam.objects.select_related('term').get(id=exam_id)
            # If the selected exam is not published, do not show grades for it
            if ex.results_published:
                selected_exam = ex
                grades_qs = grades_qs.filter(exam=selected_exam)
            else:
                # Keep grades empty if an unpublished exam was requested
                grades_qs = grades_qs.none()
        except Exam.DoesNotExist:
            grades_qs = grades_qs.none()
    # If no exam explicitly selected, default to latest published exam the student has grades for
    if selected_exam is None:
        latest_published = published_exams.first()
        if latest_published:
            selected_exam = latest_published
            grades_qs = grades_qs.filter(exam=selected_exam)

    grades = grades_qs
    average_score = grades.aggregate(avg=Avg('score'))['avg'] if hasattr(grades, 'exists') and grades.exists() else None

    # --- Performance over time (all exams ever done, oldest first) ---
    # IMPORTANT: Use ALL grades for this student (not the filtered grades for selected exam)
    exam_performance = []
    # Only consider exams that are done/published
    all_grades_qs = Grade.objects.filter(student=student, exam__results_published=True)
    # 1) Preferred: aggregate directly from the student's ALL grades queryset
    exam_rows = (
        all_grades_qs
        .values('exam_id', 'exam__name', 'exam__start_date', 'exam__term__start_date')
        .annotate(average=Avg('score'))
        .order_by('exam__start_date', 'exam_id')
    )
    for row in exam_rows:
        ex_date = row.get('exam__start_date') or row.get('exam__term__start_date')
        exam_performance.append({
            'exam_id': row['exam_id'],
            'exam_name': row['exam__name'],
            'date': ex_date.strftime('%Y-%m-%d') if ex_date else '',
            'average_score': float(row['average']) if row['average'] is not None else None,
        })
    # 2) Fallback: if still empty, rebuild via Exam -> Grade join
    if not exam_performance:
        student_exam_ids = (
            Grade.objects
            .filter(student=student)
            .values_list('exam_id', flat=True)
            .distinct()
        )
        student_exams = Exam.objects.filter(id__in=student_exam_ids).order_by('start_date', 'id')
        for ex in student_exams:
            ex_grades = Grade.objects.filter(student=student, exam=ex)
            avg_val = ex_grades.aggregate(avg=Avg('score'))['avg'] if ex_grades.exists() else None
            ex_date = ex.start_date or (getattr(ex, 'term', None).start_date if getattr(ex, 'term', None) else None)
            exam_performance.append({
                'exam_id': ex.id,
                'exam_name': ex.name,
                'date': ex_date.strftime('%Y-%m-%d') if ex_date else '',
                'average_score': float(avg_val) if avg_val is not None else None,
            })
    # Debug trace (safe prints)
    try:
        print(f"[DEBUG] student_profile exam rows count={len(exam_performance)} student_id={student.id}")
    except Exception:
        pass

    # --- Ranking Logic ---
    # Class (stream) ranking
    class_students = Student.objects.filter(class_group=student.class_group)
    class_scores = []
    for s in class_students:
        grades = Grade.objects.filter(student=s)
        avg = grades.aggregate(avg=Avg('score'))['avg'] if grades.exists() else None
        class_scores.append({'student_id': s.id, 'avg': avg})
    class_scores = [x for x in class_scores if x['avg'] is not None]
    class_scores.sort(key=lambda x: x['avg'], reverse=True)
    class_rank = None
    stream_rank = None
    for idx, entry in enumerate(class_scores, start=1):
        if entry['student_id'] == student.id:
            class_rank = idx
            stream_rank = idx
            break
    class_total = len(class_scores)
    stream_total = class_total

    # Overall (whole level) ranking
    overall_level_students = Student.objects.filter(class_group__level=student.class_group.level)
    overall_level_scores = []
    for s in overall_level_students:
        grades = Grade.objects.filter(student=s)
        avg = grades.aggregate(avg=Avg('score'))['avg'] if grades.exists() else None
        overall_level_scores.append({'student_id': s.id, 'avg': avg})
    overall_level_scores = [x for x in overall_level_scores if x['avg'] is not None]
    overall_level_scores.sort(key=lambda x: x['avg'], reverse=True)
    overall_level_rank = None
    for idx, entry in enumerate(overall_level_scores, start=1):
        if entry['student_id'] == student.id:
            overall_level_rank = idx
            break
    overall_level_total = len(overall_level_scores)

    # Prepare overall level totals (already computed above)
    # overall_level_rank and overall_level_total are now correctly set

    # --- Total Points Calculation ---
    points_map = {'A': 4, 'B': 3, 'C': 2, 'D': 1}
    total_points = sum(points_map.get(g.grade_letter, 0) for g in grades if getattr(g, 'grade_letter', None))

    # Handle contact info update POST
    if request.method == 'POST' and 'update_contact' in request.POST:
        contact_form = StudentContactUpdateForm(request.POST, instance=student, user_instance=student.user)
        if contact_form.is_valid():
            contact_form.save()
            messages.success(request, 'Contact information updated successfully.')
            return redirect('student_profile', student_id=student.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        contact_form = StudentContactUpdateForm(instance=student, user_instance=student.user)

    context = {
        'student': student,
        'current_term': current_term,
        'term': (selected_exam.term if selected_exam else current_term),
        'total_billed': total_billed,
        'total_paid': total_paid,
        'balance': balance,
        'grades': grades,
        'average_score': average_score,
        'exam_performance': exam_performance,
        'class_rank': class_rank,
        'class_total': class_total,
        'overall_level_rank': overall_level_rank,
        'overall_level_total': overall_level_total,
        'stream_rank': stream_rank,
        'stream_total': stream_total,
        'fee_assignments': fee_assignments,
        'fee_payments': FeePayment.objects.filter(student=student, fee_assignment__in=fee_assignments),
        'total_points': total_points,
        'contact_form': contact_form,
        # Restriction flags for template/UI
        'can_view_results': can_view_results,
        'fee_restriction_message': fee_restriction_message,
        'restrict_results_by_fee': restrict_results_by_fee,
        # Exam selection in UI
        'published_exams': published_exams,
        'selected_exam': selected_exam,
    }
    return render(request, 'dashboards/student_profile.html', context)

@login_required(login_url='login')
def upload_grades(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        form = GradeUploadForm(request.POST, request.FILES, teacher=teacher)
        if form.is_valid():
            exam = form.cleaned_data['exam']
            subject = form.cleaned_data['subject']
            class_group = form.cleaned_data['class_group']
            file = request.FILES['file']

            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Create a mapping of student names to student objects for efficient lookup
                students_in_class = Student.objects.filter(class_group=class_group).select_related('user')
                student_map = {s.user.get_full_name().strip().lower(): s for s in students_in_class}

                # Check for duplicate names within the class, which would make lookups ambiguous
                from collections import Counter
                name_counts = Counter(student_map.keys())
                duplicates = [name for name, count in name_counts.items() if count > 1]
                if duplicates:
                    messages.error(request, f"Upload failed. The following student names are duplicated in the class: {', '.join(duplicates)}. Please resolve this issue before uploading.")
                    return redirect('upload_grades', teacher_id=teacher.id)

                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or row[0] is None:
                        continue  # Skip empty rows

                    student_name, score = row[0], row[1]
                    student_name = str(student_name).strip().lower()

                    student = student_map.get(student_name)

                    if not student:
                        messages.error(request, f"Error on row {i}: Student '{row[0]}' not found in this class.")
                        continue

                    try:
                        score_float = float(score)
                        grade_letter, remarks = 'F', 'NEEDS IMPROVEMENT'
                        if 80 <= score_float <= 100:
                            grade_letter, remarks = 'A', 'EXCELLENT'
                        elif 60 <= score_float < 80:
                            grade_letter, remarks = 'B', 'VERY GOOD'
                        elif 40 <= score_float < 60:
                            grade_letter, remarks = 'C', 'GOOD'
                        elif 20 <= score_float < 40:
                            grade_letter, remarks = 'D', 'AVERAGE'

                        Grade.objects.update_or_create(
                            student=student,
                            exam=exam,
                            subject=subject,
                            defaults={'score': score_float, 'grade_letter': grade_letter, 'remarks': remarks}
                        )
                    except (ValueError, TypeError):
                        messages.error(request, f"Error on row {i}: Invalid score '{score}' for student '{row[0]}'. Please enter a number.")
                        continue

                messages.success(request, 'Grades have been uploaded and processed.')
                # Redirect to the results page for the uploaded grades
                return redirect('exam_results', teacher_id=teacher.id, class_id=class_group.id, subject_id=subject.id, exam_id=exam.id)

            except Exception as e:
                messages.error(request, f"Error reading the Excel file: {e}")

    else:
        form = GradeUploadForm(teacher=teacher)
    
    return render(request, 'dashboards/upload_grades.html', {'form': form, 'teacher': teacher})

@login_required(login_url='login')
def teacher_profile(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.user.role == 'teacher' and request.user.teacher != teacher:
        return HttpResponseForbidden("You are not allowed to view this profile.")

    if request.method == 'POST':
        form = TeacherProfileUpdateForm(request.POST, instance=teacher, user_instance=teacher.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile has been updated successfully.')
            return redirect('teacher_profile', teacher_id=teacher.id)
    else:
        form = TeacherProfileUpdateForm(instance=teacher, user_instance=teacher.user)

    return render(request, 'dashboards/teacher_profile.html', {'teacher': teacher, 'form': form})
    if not (request.user.role == 'admin' or request.user.id == teacher.user.id):
        return HttpResponseForbidden('You do not have permission to view this profile.')

    return render(request, 'dashboards/teacher_profile.html', {'teacher': teacher})

@login_required(login_url='login')
def student_dashboard(request):
    """
    Student dashboard view: shows a welcome message and entry points for student actions.
    """
    # Optionally, you can add more context data here for performance, fees, notifications, etc.
    return render(request, 'dashboards/student_dashboard.html', {})

@login_required(login_url='login')
def teacher_dashboard(request, teacher_id):
    teacher = get_object_or_404(Teacher.objects.select_related('user'), id=teacher_id)
    # Access control: only the teacher themselves or an admin can view
    is_admin = getattr(request.user, 'role', None) == 'admin'
    if not is_admin and request.user != teacher.user:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('You are not allowed to view this dashboard.')

    # Get all classes where the teacher teaches (via TeacherClassAssignment) or is class_teacher
    assignments = TeacherClassAssignment.objects.filter(teacher=teacher).select_related('class_group', 'subject')
    assigned_classes = set()
    assigned_class_ids = set()
    # Prepare class_cards for dashboard cards (one per class+subject assignment)
    class_cards = []
    notifications = []
    greeting_name = f"Mr. {teacher.user.last_name}" if teacher.user.last_name else teacher.user.get_full_name() or teacher.user.username
    teacher_subjects = list(teacher.subjects.values_list('name', flat=True))
    # We'll collect classes from both assignments and class_teacher role, avoiding duplicates
    for assign in assignments:
        assigned_classes.add(assign.class_group)
        assigned_class_ids.add(assign.class_group.id)
        class_cards.append({
            'class': assign.class_group,
            'subject': assign.subject,
        })
    teacher_classes = list(assigned_classes)
    # Example: get upcoming deadlines (dummy)
    upcoming_deadlines = Deadline.objects.filter(teacher=teacher).order_by('due_date')[:5]
    from .models import TeacherResponsibility
    responsibilities = TeacherResponsibility.objects.filter(teacher=teacher).order_by('-assigned_at')

    # Ensure show_expired is defined if referenced in context/templates
    show_expired = request.GET.get('show_expired') == '1'

    context = {
        'teacher': teacher,
        'greeting_name': greeting_name,
        'teacher_subjects': teacher_subjects,
        'teacher_classes': teacher_classes,
        'notifications': notifications,
        'class_cards': class_cards,
        'upcoming_deadlines': upcoming_deadlines,
        'responsibilities': responsibilities,
        'show_expired': show_expired,
    }
    # Now add classes where teacher is class_teacher but not in TeacherClassAssignment
    from .models import Class
    extra_classes = Class.objects.filter(class_teacher=teacher).exclude(id__in=assigned_class_ids)
    for class_obj in extra_classes:
        students = Student.objects.filter(class_group=class_obj)
        student_count = students.count()
        class_cards.append({
            'class': class_obj,
            'subject': None,
            'is_class_teacher': True,
            'student_count': student_count,
            'avg_score': None,
            'exams_marked': 0,
            'total_exams': 0,
            'low_performers': 0,
        })
    # teacher_classes for summary panel: all unique class names
    teacher_classes = list(set([
        c['class_group'].name if 'class_group' in c else c['class'].name for c in class_cards
    ]))
    # Remove legacy/duplicate class_details/extra_classes logic (all cards come from class_cards now)


    # Fetch upcoming events (deadlines)
    from django.utils import timezone
    now = timezone.now()
    upcoming_deadlines = Event.objects.filter(is_done=False, start__gte=now).order_by('start')

    # Add days_remaining to each event for countdown
    for event in upcoming_deadlines:
        delta = event.start - now
        event.days_remaining = delta.days + (1 if delta.seconds > 0 else 0)

    # Fetch assigned responsibilities for this teacher
    from .models import TeacherResponsibility
    responsibilities = TeacherResponsibility.objects.filter(teacher=teacher).order_by('-assigned_at')

    context = {
        'teacher': teacher,
        'greeting_name': greeting_name,
        'teacher_subjects': teacher_subjects,
        'teacher_classes': teacher_classes,
        'notifications': notifications,
        'class_cards': class_cards,
        'upcoming_deadlines': upcoming_deadlines,
        'responsibilities': responsibilities,
    }
    return render(request, 'dashboards/teacher_dashboard.html', context)


@login_required(login_url='login')
def manage_attendance(request, teacher_id):
    from .models import Teacher, TeacherClassAssignment
    from django.shortcuts import get_object_or_404

    teacher = get_object_or_404(Teacher.objects.select_related('user'), id=teacher_id)
    assignments = TeacherClassAssignment.objects.filter(teacher=teacher).select_related('class_group', 'subject')

    context = {
        'teacher': teacher,
        'assignments': assignments,
    }
    return render(request, 'dashboards/manage_attendance.html', context)


@login_required(login_url='login')
def take_attendance(request, teacher_id, class_id, subject_id):
    from .models import Teacher, Class, Subject, Exam, Grade, Student, SubjectGradingScheme, Term, GradeCommentTemplate, User
    from django.shortcuts import get_object_or_404
    from django.utils import timezone
    import datetime

    teacher = get_object_or_404(Teacher, id=teacher_id)
    class_group = get_object_or_404(Class, id=class_id)
    subject = get_object_or_404(Subject, id=subject_id)
    students = Student.objects.filter(class_group=class_group).order_by('user__last_name', 'user__first_name')
    
    # Use today's date for attendance
    today = timezone.now().date()

    if request.method == 'POST':
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            if status:
                Attendance.objects.update_or_create(
                    student=student,
                    subject=subject,
                    date=today,
                    defaults={
                        'teacher': teacher,
                        'class_group': class_group,
                        'status': status
                    }
                )
        messages.success(request, f"Attendance for {class_group.name} - {subject.name} has been saved.")
        return redirect('manage_attendance', teacher_id=teacher.id)

    # Get existing attendance records for today to pre-fill the form
    existing_attendance = Attendance.objects.filter(class_group=class_group, subject=subject, date=today)
    attendance_map = {att.student_id: att.status for att in existing_attendance}

    context = {
        'teacher': teacher,
        'class_group': class_group,
        'subject': subject,
        'students': students,
        'attendance_map': attendance_map,
        'today': today,
    }
    return render(request, 'dashboards/take_attendance.html', context)


@login_required(login_url='login')
def manage_grades(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)

    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        subject_id = request.POST.get('subject_id')
        exam_id = request.POST.get('exam_id')

        if not all([class_id, subject_id, exam_id]):
            messages.error(request, 'Please select a class, subject, and exam.')
            return redirect('manage_grades', teacher_id=teacher.id)

        return redirect('input_grades', teacher_id=teacher.id, class_id=class_id, subject_id=subject_id, exam_id=exam_id)

    assignments = TeacherClassAssignment.objects.filter(teacher=teacher).select_related('class_group', 'subject')
    
    # Create a unique list of classes and subjects taught by the teacher
    teacher_classes = sorted(list(set(a.class_group for a in assignments if a.class_group)), key=lambda c: c.name)
    teacher_subjects = Subject.objects.filter(teacherclassassignment__teacher=teacher).distinct().order_by('name')
    
    # Show all unpublished exams that are done (can be in past or future)
    from django.utils import timezone
    today = timezone.now().date()
    exams = Exam.objects.filter(results_published=False).order_by('-start_date')

    context = {
        'teacher': teacher,
        'teacher_classes': teacher_classes,
        'teacher_subjects': teacher_subjects,
        'exams': exams,
    }
    return render(request, 'dashboards/manage_grades.html', context)


@login_required(login_url='login')
def admin_grade_comments(request):
    """Admin-only view to create/update grade comment templates by grade letter."""
    # Basic role check
    if getattr(request.user, 'role', None) != 'admin':
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('dashboard') if 'dashboard' in globals() else redirect('login')

    grade_letters = ['A', 'B', 'C', 'D', 'E', 'F']

    if request.method == 'POST':
        for letter in grade_letters:
            field_name = f'comment_{letter}'
            text = (request.POST.get(field_name) or '').strip()
            if not text:
                # Allow empty to clear
                obj, _ = GradeCommentTemplate.objects.get_or_create(grade_letter=letter)
                obj.comment = ''
                obj.updated_by = request.user
                obj.save()
                continue
            obj, _ = GradeCommentTemplate.objects.get_or_create(grade_letter=letter)
            obj.comment = text
            obj.updated_by = request.user
            obj.save()
        messages.success(request, 'Grade comments updated successfully.')
        return redirect('admin_grade_comments')

    # Build current comments map
    existing = {gct.grade_letter: gct.comment for gct in GradeCommentTemplate.objects.all()}
    comments = [(letter, existing.get(letter, '')) for letter in grade_letters]

    context = {
        'comments': comments,
        'updated_at_map': {gct.grade_letter: gct.updated_at for gct in GradeCommentTemplate.objects.all()},
    }
    return render(request, 'dashboards/admin_grade_comments.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin)
def admin_manage_grades_entry(request):
    """
    Admin entry point to manage grades: pick a teacher, then redirect to existing
    teacher-specific manage_grades flow.
    """
    teachers = Teacher.objects.select_related('user').order_by('user__username')
    if request.method == 'POST':
        teacher_id = request.POST.get('teacher_id')
        if not teacher_id:
            messages.error(request, 'Please select a teacher.')
        else:
            return redirect('manage_grades', teacher_id=teacher_id)
    return render(request, 'dashboards/admin_manage_grades_entry.html', {'teachers': teachers})


@login_required(login_url='login')
def input_grades(request, teacher_id, class_id, subject_id, exam_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    class_group = get_object_or_404(Class, id=class_id)
    subject = get_object_or_404(Subject, id=subject_id)
    exam = get_object_or_404(Exam, id=exam_id)

    # Block access if exam not done
    if not exam.is_done:
        messages.error(request, 'You can only input grades for exams that are done.')
        return redirect('manage_grades', teacher_id=teacher.id)

    # Disallow edits if results already published
    if request.method == 'POST' and getattr(exam, 'results_published', False):
        messages.error(request, 'This exam\'s results have been published. Editing grades is disabled.')
        return redirect('exam_results', teacher_id=teacher.id, class_id=class_group.id, subject_id=subject.id, exam_id=exam.id)

    if request.method == 'POST':
        students = Student.objects.filter(class_group=class_group)
        # Fetch grading scheme for this subject
        grading_scheme = getattr(subject, 'grading_scheme', None)
        max_score = 100  # Default max if no scheme
        if grading_scheme and grading_scheme.grade_boundaries:
            # Find the highest max value among all grade boundaries
            try:
                max_score = max(v[1] for v in grading_scheme.grade_boundaries.values())
            except Exception:
                max_score = 100
        any_error = False
        for student in students:
            score_str = request.POST.get(f'score_{student.id}')
            if score_str and score_str.strip():
                try:
                    score = float(score_str)
                    if score > max_score:
                        messages.error(request, f"Score for {student.user.get_full_name()} exceeds the maximum allowed ({max_score}) for this subject.")
                        any_error = True
                        continue
                    grade_letter, remarks = 'F', 'NEEDS IMPROVEMENT'
                    if 80 <= score <= 100:
                        grade_letter, remarks = 'A', 'EXCELLENT'
                    elif 60 <= score < 80:
                        grade_letter, remarks = 'B', 'VERY GOOD'
                    elif 40 <= score < 60:
                        grade_letter, remarks = 'C', 'GOOD'
                    elif 20 <= score < 40:
                        grade_letter, remarks = 'D', 'AVERAGE'

                    Grade.objects.update_or_create(
                        student=student,
                        exam=exam,
                        subject=subject,
                        defaults={'score': score, 'grade_letter': grade_letter, 'remarks': remarks}
                    )
                except (ValueError, TypeError):
                    messages.error(request, f"Invalid score for student {student.id}. Please enter a number.")
                    any_error = True
                    continue
        if not any_error:
            messages.success(request, 'Grades saved successfully!')
            return redirect('exam_results', teacher_id=teacher.id, class_id=class_group.id, subject_id=subject.id, exam_id=exam.id)
        # If there were errors, do not redirect, show the form again with errors.

    students = Student.objects.filter(class_group=class_group).order_by('user__last_name', 'user__first_name')
    grades = Grade.objects.filter(student__in=students, subject=subject, exam=exam).select_related('student')
    
    grades_map = {grade.student.id: grade for grade in grades}

    for student in students:
        student.grade = grades_map.get(student.id)

    context = {
        'teacher': teacher,
        'class_group': class_group,
        'subject': subject,
        'exam': exam,
        'students': students,
    }
    return render(request, 'dashboards/input_grades.html', context)

@login_required(login_url='login')
def publish_exam_results(request, exam_id):
    """Admin-only: publish results for an exam (only allowed after exam end date)."""
    exam = get_object_or_404(Exam, id=exam_id)
    user = request.user
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if not (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser):
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
        messages.error(request, 'You do not have permission to publish exam results.')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/admin_exams/'))
    from django.utils import timezone
    not_done_warning = None
    if not exam.is_done:
        # Proceed anyway but include a warning
        not_done_warning = 'Exam end date not reached; publishing anyway as requested.'

    if not exam.results_published:
        exam.results_published = True
        exam.published_at = timezone.now()
        exam.save(update_fields=['results_published', 'published_at'])
        # Queue background notifications and return fast with a job id
        try:
            from .models import NotificationJob
            from .tasks import send_exam_publish_notifications
            job = NotificationJob.objects.create(job_type='exam_publish', status='queued', meta={'exam_id': exam.id})
            base_url = request.build_absolute_uri('/')
            send_exam_publish_notifications.delay(job.id, exam.id, base_url)
            if is_ajax:
                resp = {'success': True, 'published': True, 'job_id': job.id}
                if not_done_warning:
                    resp['warning'] = not_done_warning
                return JsonResponse(resp)
            msg = f'Exam results published. Notifications are sending in background (Job #{job.id}).'
            if not_done_warning:
                msg = not_done_warning + ' ' + msg
            messages.success(request, msg)
        except Exception as e:
            # Do not block publishing if background job fails to enqueue
            import logging
            logging.getLogger(__name__).exception('Failed to enqueue exam publish notifications: %s', e)
            warn = 'Exam published, but failed to enqueue notifications. Check migrations/worker.'
            if not_done_warning:
                warn = not_done_warning + ' ' + warn
            if is_ajax:
                return JsonResponse({'success': True, 'published': True, 'job_id': None, 'warning': warn})
            messages.warning(request, warn)
    else:
        if is_ajax:
            return JsonResponse({'success': True, 'published': True, 'info': 'Exam results are already published.'})
        messages.info(request, 'Exam results are already published.')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/admin_exams/'))

@csrf_exempt
def download_result_slip_signed(request, token):
    """Public endpoint: serves a per-student result slip PDF using a signed token."""
    from django.core import signing
    from .models import ResultBlock
    try:
        data = signing.loads(token, salt='resultslip', max_age=getattr(settings, 'RESULT_SLIP_LINK_MAX_AGE', 60*60*24*14))
        student_id = data.get('sid')
        exam_id = data.get('eid')
        if not (student_id and exam_id):
            return HttpResponseForbidden('Invalid token')
    except signing.BadSignature:
        return HttpResponseForbidden('Invalid or expired token')

    # Enforce active block
    try:
        rb = ResultBlock.objects.filter(student_id=student_id, exam_id=exam_id, active=True).first()
        if rb:
            # Show a friendly blocked page
            msg = rb.reason or "Results are temporarily unavailable due to outstanding fees."
            return render(request, 'dashboards/results_blocked.html', { 'message': msg })
    except Exception:
        pass

    # Build rows similar to export_result_slip_pdf but limited to this student and exam
    grades = Grade.objects.select_related('student__user', 'student__class_group', 'subject', 'exam__term__academic_year')\
        .filter(exam_id=exam_id, student_id=student_id)\
        .order_by('subject__name')
    if not grades.exists():
        return Http404('No results available')

    s = grades.first().student
    cls = s.class_group
    columns = ['Admission No', 'Full Name', 'Class', 'Level', 'Subject', 'Score', 'Grade', 'Exam', 'Term', 'Year']
    rows = []
    for g in grades:
        term = g.exam.term if g.exam else None
        rows.append([
            s.admission_no,
            s.user.get_full_name() if s.user else '',
            getattr(cls, 'name', ''),
            getattr(cls, 'level', ''),
            g.subject.name if g.subject else '',
            g.score,
            g.grade_letter or '',
            g.exam.name if g.exam else '',
            term.name if term else '',
            term.academic_year.year if getattr(term, 'academic_year', None) else '',
        ])
    filename = f"result_slip_{s.admission_no or s.id}.pdf"
    return pdf_response_from_rows(filename, 'Result Slip', _site_header_rows(), columns, rows)

@login_required(login_url='login')
def unpublish_exam_results(request, exam_id):
    """Admin-only: unpublish results for an exam (re-enable grade editing)."""
    exam = get_object_or_404(Exam, id=exam_id)
    user = request.user
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if not (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser):
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
        messages.error(request, 'You do not have permission to unpublish exam results.')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/admin_exams/'))
    if exam.results_published:
        exam.results_published = False
        exam.save(update_fields=['results_published'])
        if is_ajax:
            return JsonResponse({'success': True, 'published': False})
        messages.success(request, 'Exam results unpublished. Teachers can edit grades again.')
    else:
        if is_ajax:
            return JsonResponse({'success': True, 'published': False, 'info': 'Exam results are not published yet.'})
        messages.info(request, 'Exam results are not published yet.')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/admin_exams/'))


@login_required(login_url='login')
def exam_results(request, teacher_id, class_id, subject_id, exam_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    class_group = get_object_or_404(Class, id=class_id)
    subject = get_object_or_404(Subject, id=subject_id)
    exam = get_object_or_404(Exam, id=exam_id)

    students = Student.objects.filter(class_group=class_group).order_by('user__last_name', 'user__first_name')
    grades = Grade.objects.filter(student__in=students, subject=subject, exam=exam).select_related('student')

    grades_map = {grade.student.id: grade for grade in grades}

    for student in students:
        student.grade = grades_map.get(student.id)

    context = {
        'teacher': teacher,
        'class_group': class_group,
        'subject': subject,
        'exam': exam,
        'students': students,
    }
    return render(request, 'dashboards/exam_results.html', context)


@login_required(login_url='login')
def gradebook(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)

    # Get all unique classes and subjects assigned to the teacher
    teacher_classes = Class.objects.filter(teacherclassassignment__teacher=teacher).distinct()
    # Subjects assigned via TeacherClassAssignment
    teacher_subjects = Subject.objects.filter(teacherclassassignment__teacher=teacher).distinct()

    selected_class = None
    selected_subject = None
    students_grades = []
    exams = []

    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        subject_id = request.POST.get('subject_id')

        if class_id and subject_id:
            selected_class = get_object_or_404(Class, id=class_id)
            selected_subject = get_object_or_404(Subject, id=subject_id)

            # After class selection, narrow subjects to those assigned to this class
            teacher_subjects = Subject.objects.filter(
                teacherclassassignment__teacher=teacher,
                teacherclassassignment__class_group=selected_class
            ).distinct()

            # Fetch students in the selected class
            students = Student.objects.filter(class_group=selected_class).order_by('user__last_name', 'user__first_name')
            
            # Fetch all exams that have grades for this subject and class
            grades_for_subject = Grade.objects.filter(subject=selected_subject, student__class_group=selected_class)
            exam_ids = grades_for_subject.values_list('exam_id', flat=True).distinct()
            exams = Exam.objects.filter(id__in=exam_ids).order_by('date')

            # Prepare a data structure for the template
            for student in students:
                grades = {grade.exam_id: grade for grade in grades_for_subject.filter(student=student)}
                students_grades.append({
                    'student': student,
                    'grades': [grades.get(exam.id) for exam in exams]
                })

    context = {
        'teacher': teacher,
        'teacher_classes': teacher_classes,
        'teacher_subjects': teacher_subjects,
        'selected_class': selected_class,
        'selected_subject': selected_subject,
        'students_grades': students_grades,
        'exams': exams,
    }
    return render(request, 'dashboards/gradebook.html', context)


# Admin Students View

from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.urls import reverse

@login_required(login_url='login')
def admin_exams_json(request):
    from .models import Exam
    if not hasattr(request.user, 'role') or request.user.role != 'admin':
        return JsonResponse({'error': 'Forbidden'}, status=403)
    exams = Exam.objects.select_related('term').all()
    events = []
    for exam in exams:
        events.append({
            'id': exam.id,
            'title': exam.name,
            'start': str(exam.date),
            'end': str(exam.date),
            'type': exam.get_type_display(),
            'term': str(exam.term),
            'level': exam.level,
            'category': 'exam',
        })
    return JsonResponse(events, safe=False)

@login_required(login_url='login')
def timetable_view(request):
    # Allow admin to trigger auto-generation via POST to this same view
    if request.method == 'POST' and getattr(request.user, 'role', None) == 'admin' and request.POST.get('auto_generate') == '1':
        try:
            report = generate_timetable(overwrite=True)
            messages.success(request, f"Timetable generated. Placed: {report.get('placed', 0)}, Skipped: {report.get('skipped', 0)}")
        except Exception as e:
            messages.error(request, f"Failed to generate timetable: {e}")
        from django.shortcuts import redirect
        return redirect('timetable_view')
    if not request.user.role == 'admin':
        return HttpResponseForbidden("You are not authorized to view this page.")
    
    # Dashboard context with correct variable names for the template
    from .models import FeeCategory, FeeAssignment, FeePayment, Event, Term, TeacherResponsibility
    from django.utils import timezone
    now = timezone.now()

    total_students = Student.objects.filter(graduated=False).count()
    students_boys = Student.objects.filter(graduated=False, gender__iexact='male').count()
    students_girls = Student.objects.filter(graduated=False, gender__iexact='female').count()
    total_teachers = Teacher.objects.count()
    teachers_male = Teacher.objects.filter(gender__iexact='male').count()
    teachers_female = Teacher.objects.filter(gender__iexact='female').count()
    total_classes = Class.objects.count()
    total_subjects = Subject.objects.count()
    total_fees = FeeCategory.objects.count()

    # Current term
    today = timezone.now().date()
    current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()

    # Upcoming events (next 5, not done, starting today or later)
    upcoming_events = Event.objects.filter(is_done=False, start__gte=now).order_by('start')[:5]

    # Assigned responsibilities (latest 8)
    responsibilities = TeacherResponsibility.objects.select_related('teacher', 'teacher__user', 'assigned_by').order_by('-assigned_at')[:8]

    context = {
        'total_students': total_students,
        'students_boys': students_boys,
        'students_girls': students_girls,
        'total_teachers': total_teachers,
        'teachers_male': teachers_male,
        'teachers_female': teachers_female,
        'total_classes': total_classes,
        'total_subjects': total_subjects,
        'total_fees': total_fees,
        'current_term': current_term,
        'upcoming_events': upcoming_events,
        'responsibilities': responsibilities,
    }
    return render(request, 'dashboards/admin_overview.html', context)

def edit_user(request, user_id):
    if request.user.role != 'admin':
        return redirect('login')
    from .models import User
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        role = request.POST.get('role', user.role)
        is_active = 'is_active' in request.POST
        is_staff = 'is_staff' in request.POST
        # Check for username/email uniqueness
        if User.objects.exclude(id=user.id).filter(username=username).exists():
            messages.error(request, 'Username already exists.')
        elif User.objects.exclude(id=user.id).filter(email=email).exists():
            messages.error(request, 'Email already exists.')
        else:
            user.username = username
            user.email = email
            user.first_name = first_name
            user.last_name = last_name
            user.role = role
            user.is_active = is_active
            user.is_staff = is_staff
            password = request.POST.get('password', '')
            if password:
                user.set_password(password)
            user.save()

            # If marked as clerk, grant finance-related model permissions automatically
            if role == 'clerk':
                try:
                    from django.contrib.auth.models import Permission
                    from django.contrib.contenttypes.models import ContentType
                    from .models import FeeCategory, FeeAssignment, FeePayment, MpesaTransaction
                    models_for_finance = [FeeCategory, FeeAssignment, FeePayment, MpesaTransaction]
                    perms_to_add = []
                    for mdl in models_for_finance:
                        ct = ContentType.objects.get_for_model(mdl)
                        perms_to_add.extend(list(Permission.objects.filter(content_type=ct)))
                    if perms_to_add:
                        user.user_permissions.add(*perms_to_add)
                except Exception as e:
                    # Non-fatal: log a message for admin UI
                    messages.warning(request, f"Clerk permissions could not be fully applied: {e}")
            messages.success(request, 'User updated successfully.')
            return redirect('admin_users')
    return render(request, 'dashboards/edit_user.html', {'user': user})

@login_required(login_url='login')
def delete_user(request, user_id):
    if request.user.role != 'admin':
        return redirect('login')
    from .models import User
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'User deleted successfully.')
        return redirect('admin_users')
    return render(request, 'dashboards/delete_user.html', {'user': user})


@login_required(login_url='login')
def admin_users(request):
    if request.user.role != 'admin':
        return redirect('login')
    from .models import User
    users = User.objects.all().order_by('username')
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '').strip()
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    if role_filter:
        users = users.filter(role=role_filter)
    users = users.order_by('username')
    context = {
        'users': users,
        'search_query': search_query,
        'role_filter': role_filter,
        'role_choices': User._meta.get_field('role').choices,
    }
    return render(request, 'dashboards/admin_users.html', context)

@login_required(login_url='login')
def admin_students(request):
    if request.user.role != 'admin':
        return redirect('login')
    from .models import Student, User, Class
    from datetime import date
    students = Student.objects.select_related('user', 'class_group').all()
    search_query = request.GET.get('search', '').strip()
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(admission_no__icontains=search_query)
        )
    # Extra filters
    gender = request.GET.get('gender', '').strip()
    class_id = request.GET.get('class_id', '').strip()
    age_min = request.GET.get('age_min', '').strip()
    age_max = request.GET.get('age_max', '').strip()

    if gender:
        students = students.filter(gender=gender)

    if class_id:
        students = students.filter(class_group_id=class_id)

    # Age filters (approximate):
    # age >= N  => birthdate <= today - N years
    # age <= M  => birthdate >= today - M years
    if age_min or age_max:
        today = date.today()

        def years_ago(n):
            if n is None:
                return None
            try:
                return today.replace(year=today.year - n)
            except ValueError:
                # handle Feb 29 -> Feb 28 on non-leap years
                return today.replace(month=2, day=28, year=today.year - n)

        if age_min.isdigit():
            students = students.filter(birthdate__lte=years_ago(int(age_min)))
        if age_max.isdigit():
            students = students.filter(birthdate__gte=years_ago(int(age_max)))

    add_student_form = AddStudentForm()
    if request.method == 'POST' and 'add_student' in request.POST:
        add_student_form = AddStudentForm(request.POST)
        if add_student_form.is_valid():
            from django.db import IntegrityError, transaction
            try:
                with transaction.atomic():
                    # Create user first
                    user = User.objects.create_user(
                        username=add_student_form.cleaned_data['username'],
                        email=add_student_form.cleaned_data['email'],
                        first_name=add_student_form.cleaned_data['first_name'],
                        last_name=add_student_form.cleaned_data['last_name'],
                        role='student',
                        password=add_student_form.cleaned_data['password']
                    )
                    
                    # Create student profile
                    student = Student(
                        user=user,
                        admission_no=add_student_form.cleaned_data['admission_no'],
                        gender=add_student_form.cleaned_data['gender'],
                        birthdate=add_student_form.cleaned_data['birthdate'],
                        class_group=add_student_form.cleaned_data['class_group']
                    )
                    student.save()
                    
                    messages.success(request, 'Student added successfully!')
                    return redirect('admin_students')
                    
            except IntegrityError as e:
                if 'username' in str(e):
                    messages.error(request, 'A user with that username already exists. Please choose a different username.')
                elif 'email' in str(e):
                    messages.error(request, 'A user with that email already exists. Please choose a different email.')
                else:
                    messages.error(request, f'An error occurred while adding the student: {str(e)}')
            except Exception as e:
                messages.error(request, f'An unexpected error occurred: {str(e)}')
        else:
            # Form is not valid, show form errors
            for field, errors in add_student_form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    context = {
        'students': students,
        'add_student_form': add_student_form,
        'classes': Class.objects.all().order_by('name'),
        'site_settings': SiteSettings.objects.first(),
    }
    return render(request, 'dashboards/admin_students.html', context)

# Admin Classes View
@login_required(login_url='login')
def admin_classes(request):
    if request.user.role != 'admin':
        return redirect('login')
    from .models import Class, Student, Subject, Teacher, TeacherClassAssignment, Grade
    from django.contrib import messages
    from .forms import AddClassForm
    add_class_form = AddClassForm()
    # Handle assignment POST and AddClassForm
    if request.method == 'POST':
        if 'assign_teacher' in request.POST:
            class_id = request.POST.get('class_id')
            subject_id = request.POST.get('subject_id')
            teacher_id = request.POST.get('teacher_id')  # may be '' to indicate unassign
            if class_id and subject_id:
                try:
                    class_obj = Class.objects.get(id=class_id)
                    subject_obj = Subject.objects.get(id=subject_id)
                    # Enforce: allow parent or atom; block component (child) subjects only
                    is_child = Subject.objects.filter(id=subject_obj.id, part_of__isnull=False).exists()
                    if is_child:
                        messages.error(request, f"Cannot assign a teacher to '{subject_obj.name}' because it is a component subject. Please select a parent or atomic subject.")
                        return redirect('admin_classes')

                    # Handle unassign when teacher_id is blank
                    if not teacher_id:
                        existing = TeacherClassAssignment.objects.filter(class_group=class_obj, subject=subject_obj).first()
                        if existing:
                            existing.delete()
                            messages.success(request, f"Unassigned {subject_obj.name} from {class_obj.name}.")
                        else:
                            messages.info(request, f"{subject_obj.name} was not assigned for {class_obj.name}.")
                        return redirect('admin_classes')

                    # Else, assign/reassign to a selected teacher
                    teacher_obj = Teacher.objects.get(id=teacher_id)
                    # Ensure teacher teaches the selected subject
                    if not teacher_obj.subjects.filter(id=subject_obj.id).exists():
                        messages.error(request, f"{teacher_obj.user.get_full_name()} does not teach {subject_obj.name}. Update the teacher's subjects first.")
                        return redirect('admin_classes')

                    assignment, created = TeacherClassAssignment.objects.get_or_create(
                        class_group=class_obj, subject=subject_obj,
                        defaults={'teacher': teacher_obj}
                    )
                    if not created:
                        # Reassign to a different teacher
                        if assignment.teacher_id == teacher_obj.id:
                            messages.info(request, f"{subject_obj.name} is already assigned to {teacher_obj.user.get_full_name()} for {class_obj.name}.")
                        else:
                            assignment.teacher = teacher_obj
                            assignment.save()
                            messages.success(request, f"Reassigned {subject_obj.name} to {teacher_obj.user.get_full_name()} for {class_obj.name}.")
                    else:
                        messages.success(request, f"Assigned {teacher_obj.user.get_full_name()} to {subject_obj.name} for {class_obj.name}.")
                except (Class.DoesNotExist, Subject.DoesNotExist, Teacher.DoesNotExist):
                    messages.error(request, "Invalid class, subject, or teacher selection.")
            else:
                messages.error(request, "Please select class and subject.")
        elif 'assign_class_teacher' in request.POST:
            class_id = request.POST.get('class_id')
            teacher_id = request.POST.get('teacher_id')
            if class_id and teacher_id:
                try:
                    class_obj = Class.objects.get(id=class_id)
                    teacher_obj = Teacher.objects.get(id=teacher_id)
                    class_obj.class_teacher = teacher_obj
                    class_obj.save()
                    messages.success(request, f"Changed class teacher for {class_obj.name} to {teacher_obj.user.get_full_name()}.")
                except (Class.DoesNotExist, Teacher.DoesNotExist):
                    messages.error(request, "Invalid class or teacher selection.")
            else:
                messages.error(request, "Please select both class and teacher.")
        elif 'add_class' in request.POST:
            add_class_form = AddClassForm(request.POST)
            if add_class_form.is_valid():
                add_class_form.save()
                messages.success(request, "Class added successfully.")
                add_class_form = AddClassForm()  # reset form
            else:
                messages.error(request, "Please correct the errors below.")
    # Prepare data for GET
    classes = Class.objects.all()
    # Only exclude component (child) subjects; allow parents and atoms
    all_subjects = Subject.objects.filter(part_of__isnull=True).order_by('name').distinct()
    all_teachers = Teacher.objects.select_related('user').all()
    class_list = [];
    for c in classes:
        # For this class, get all subjects (could be all, or you may want to filter)
        # Build a mapping of teachers to their assigned subjects for this class
        assignments = TeacherClassAssignment.objects.filter(class_group=c).select_related('teacher__user', 'subject')
        teacher_subject_map = {}
        for assign in assignments:
            teacher = assign.teacher
            if teacher not in teacher_subject_map:
                teacher_subject_map[teacher] = []
            teacher_subject_map[teacher].append(assign.subject)
        teachers_and_subjects = []
        for teacher, subjects in teacher_subject_map.items():
            teachers_and_subjects.append({
                'teacher': teacher,
                'subjects': subjects
            })
        # For backward compatibility in template, assign to 'subjects_and_teachers'
        subjects_and_teachers = teachers_and_subjects

        class_students = Student.objects.filter(class_group=c, graduated=False).select_related('user')
        total_students = class_students.count()
        boys_count = class_students.filter(gender__iexact='male').count()
        girls_count = class_students.filter(gender__iexact='female').count()
        # Compute available subjects for new assignment: non-child and not already assigned
        assigned_subject_ids = set(a.subject_id for a in assignments)
        available_subjects = [s for s in all_subjects if s.id not in assigned_subject_ids]

        # Compute subject performance for this class
        from django.db.models import Avg
        subject_performance = []
        for subj in all_subjects:
            avg_score = Grade.objects.filter(student__class_group=c, subject=subj).aggregate(avg=Avg('score'))['avg']
            subject_performance.append({'subject': subj.name, 'avg_score': avg_score or 0})
        class_list.append({
            'id': c.id,
            'name': c.name,
            'level': c.level,
            'class_teacher': c.class_teacher.user.get_full_name() if c.class_teacher else None,
            'class_teacher_username': c.class_teacher.user.username if c.class_teacher else None,
            'subject_performance': subject_performance,
            'subjects_and_teachers': subjects_and_teachers,
            'available_subjects': available_subjects,
            'total_students': total_students,
            'boys_count': boys_count,
            'girls_count': girls_count,
        })
    # Optionally, provide all graduated students for display
    graduated_students = Student.objects.filter(graduated=True).select_related('user')
    # Calculate graduation year info
    import datetime
    current_year = datetime.date.today().year
    graduated_info = []
    for student in graduated_students:
        # Try to find the last academic year in which the student was active (based on their grades)
        last_grade = student.grade_set.select_related('exam__term__academic_year').order_by('-exam__term__academic_year__year').first()
        grad_year = None
        if last_grade and last_grade.exam and last_grade.exam.term and last_grade.exam.term.academic_year:
            # Extract the year as int (handles '2024' or '2024/2025')
            import re
            match = re.match(r'(\d{4})', str(last_grade.exam.term.academic_year.year))
            if match:
                grad_year = int(match.group(1))
        if grad_year:
            years_ago = current_year - grad_year
            if years_ago == 0:
                status = 'graduated this year'
            elif years_ago == 1:
                status = 'graduated last year'
            else:
                status = f'graduated {years_ago} years ago'
        else:
            status = 'graduation year unknown'
        graduated_info.append({'student': student, 'status': status})
    context = {
        'class_list': class_list,
        'graduated_students': graduated_students,
        'graduated_info': graduated_info,
        'add_class_form': add_class_form,
        'all_subjects': all_subjects,
        'all_teachers': all_teachers,
    }
    return render(request, 'dashboards/admin_classes.html', context)

@login_required(login_url='login')
def admin_graduated_students(request):
    """Admin view showing all graduated students with simple search and filters."""
    if request.user.role != 'admin':
        return redirect('login')
    from .models import Student
    from django.db.models import Q
    # Filters
    q = (request.GET.get('q') or '').strip()
    gender = (request.GET.get('gender') or '').strip()
    # Base queryset: graduated students only
    students = Student.objects.filter(graduated=True).select_related('user', 'class_group')
    if q:
        students = students.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(user__username__icontains=q) |
            Q(admission_no__icontains=q)
        )
    if gender:
        students = students.filter(gender__iexact=gender)
    students = students.order_by('user__last_name', 'user__first_name')
    context = {
        'students': students,
        'q': q,
        'gender': gender,
    }
    return render(request, 'dashboards/admin_graduated_students.html', context)

@login_required(login_url='login')
def class_profile(request, class_id):
    from .models import Class, Student, TeacherClassAssignment, Subject, FeeAssignment, FeePayment
    class_obj = Class.objects.select_related('class_teacher').get(id=class_id)
    students = Student.objects.filter(class_group=class_obj).select_related('user')
    assignments = TeacherClassAssignment.objects.filter(class_group=class_obj).select_related('teacher__user', 'subject')
    subjects_and_teachers = []
    for subj in Subject.objects.all():
        assignment = assignments.filter(subject=subj).first()
        subjects_and_teachers.append({
            'subject': subj,
            'teacher': assignment.teacher if assignment else None,
        })
    # --- Level Ranking Table for This Class's Level ---
    from django.db.models import Avg, Sum
    from django.utils import timezone
    from .models import Exam
    level_students = Student.objects.filter(class_group__level=class_obj.level)
    today = timezone.now().date()
    current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()
    level_scores = []
    if current_term:
        for s in level_students:
            exams = Exam.objects.filter(term=current_term, level=class_obj.level)
            grades = Grade.objects.filter(student=s, exam__in=exams)
            avg = grades.aggregate(avg=Avg('score'))['avg'] if grades.exists() else None
            if avg is not None:
                name = s.user.get_full_name() or s.user.username
                level_scores.append({'name': name, 'avg': avg})
        level_scores.sort(key=lambda x: x['avg'], reverse=True)

    # --- Finance Data ---
    fee_assignments = FeeAssignment.objects.filter(class_group=class_obj)
    total_billed = fee_assignments.aggregate(total=Sum('amount'))['total'] or 0
    payments = FeePayment.objects.filter(fee_assignment__in=fee_assignments)
    total_paid = payments.aggregate(total=Sum('amount_paid'))['total'] or 0
    balance = total_billed - total_paid
    class_finances = {
        'total_billed': total_billed,
        'total_paid': total_paid,
        'balance': balance,
        'fee_assignments': fee_assignments,
        'payments': payments,
    }

    # --- Students With Outstanding Balances ---
    student_fees = fee_assignments.aggregate(total=Sum('amount'))['total'] or 0
    students_with_balances = []
    for student in students:
        student_paid = FeePayment.objects.filter(fee_assignment__in=fee_assignments, student=student).aggregate(total=Sum('amount_paid'))['total'] or 0
        student_balance = student_fees - student_paid
        if student_balance > 0:
            students_with_balances.append({
                'student': student,
                'balance': student_balance,
            })

    context = {
        'class_obj': class_obj,
        'students': students,
        'subjects_and_teachers': subjects_and_teachers,
        'level_scores_table': level_scores,
        'class_finances': class_finances,
        'students_with_balances': students_with_balances,
        'current_term': current_term,
    }

    return render(request, 'dashboards/class_profile.html', context)

@login_required(login_url='login')
def edit_class(request, class_id):
    from .models import Class, Teacher
    # ... (rest of the code remains the same)
    from django.contrib import messages
    class_obj = Class.objects.get(id=class_id)
    if request.method == 'POST':
        level = request.POST.get('level')
        stream = request.POST.get('stream')
        class_teacher_id = request.POST.get('class_teacher')
        if level and stream:
            class_obj.name = f"Grade {level} {stream}"
            class_obj.level = level
            if class_teacher_id:
                # Prevent choosing a teacher already assigned as class teacher for another class
                if Class.objects.exclude(id=class_id).filter(class_teacher_id=class_teacher_id).exists():
                    messages.error(request, 'Selected teacher is already a class teacher of another class.')
                    # Re-render with same filtered teacher list
                    assigned_teacher_ids = (
                        Class.objects
                        .exclude(class_teacher=None)
                        .exclude(id=class_id)
                        .values_list('class_teacher', flat=True)
                    )
                    teachers = Teacher.objects.exclude(id__in=assigned_teacher_ids)
                    context = {'class_obj': class_obj, 'teachers': teachers}
                    return render(request, 'dashboards/edit_class.html', context)
                try:
                    class_obj.class_teacher = Teacher.objects.get(id=class_teacher_id)
                except Teacher.DoesNotExist:
                    class_obj.class_teacher = None
            else:
                class_obj.class_teacher = None
            class_obj.save()
            messages.success(request, 'Class updated successfully!')
            return redirect('admin_classes')
    # Exclude teachers who are already class teachers for other classes
    assigned_teacher_ids = (
        Class.objects
        .exclude(class_teacher=None)
        .exclude(id=class_id)
        .values_list('class_teacher', flat=True)
    )
    teachers = Teacher.objects.exclude(id__in=assigned_teacher_ids)
    context = {'class_obj': class_obj, 'teachers': teachers}
    return render(request, 'dashboards/edit_class.html', context)

@login_required(login_url='login')
def delete_class(request, class_id):
    from .models import Class, Student, TeacherClassAssignment
    from django.contrib import messages
    class_obj = Class.objects.get(id=class_id)
    if Student.objects.filter(class_group=class_obj).exists() or TeacherClassAssignment.objects.filter(class_group=class_obj).exists():
        messages.error(request, 'Cannot delete class: class has students or teacher assignments.')
        return redirect('admin_classes')
    if request.method == 'POST':
        class_obj.delete()
        messages.success(request, 'Class deleted successfully.')
        return redirect('admin_classes')
    context = {'class_obj': class_obj}
    return render(request, 'dashboards/delete_class.html', context)



# Timetable View (placeholder for NoReverseMatch)
def timetable_view(request):
    teacher = None
    if request.user.is_authenticated and hasattr(request.user, 'role') and request.user.role == 'teacher':
        try:
            teacher = Teacher.objects.get(user=request.user)
        except Teacher.DoesNotExist:
            teacher = None

    # This view now primarily handles class-based timetables.
    # It's kept simple, but we pass the teacher for the sidebar.
    classes = Class.objects.all().order_by('level', 'name')
    # Exclude component (child) subjects from timetable allocations
    subjects = Subject.objects.filter(part_of__isnull=True).order_by('name')
    teachers = Teacher.objects.select_related('user').all().order_by('user__last_name')
    periods = PeriodSlot.objects.filter(is_class_slot=True).order_by('start_time')
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

    selected_class_id = request.GET.get('class_id')
    selected_class = None
    timetable_grid = {}

    if selected_class_id:
        selected_class = get_object_or_404(Class, id=selected_class_id)
        # Build the grid for the selected class
        entries = DefaultTimetable.objects.filter(class_group=selected_class).select_related('period', 'subject', 'teacher__user')
        for entry in entries:
            if entry.period.id not in timetable_grid:
                timetable_grid[entry.period.id] = {}
            timetable_grid[entry.period.id][entry.day] = entry

    context = {
        'teacher': teacher,
        'classes': classes,
        'selected_class': selected_class,
        'subjects': subjects,
        'teachers': teachers,
        'periods': periods,
        'days': days,
        'timetable_grid': timetable_grid,
    }
    return render(request, 'timetable/timetable.html', context)

# Admin Academic Years & Terms View
from .models import AcademicYear, Term
from django.views.decorators.csrf import csrf_protect

@login_required(login_url='login')
def admin_class_result_slip(request, class_id):
    if request.user.role != 'admin':
        return redirect('login')
    from .models import Class, Student, Subject, Exam, Grade, Term
    class_obj = get_object_or_404(Class, id=class_id)
    students = Student.objects.filter(class_group=class_obj).select_related('user')
    # Use top-level subjects only (exclude component/child subjects)
    subjects = Subject.objects.filter(part_of__isnull=True).order_by('name')
    terms = Term.objects.order_by('-start_date')
    exams = Exam.objects.none()
    selected_term = None
    selected_exam = None
    grades = {}
    averages = {}
    totals = {}
    ranks = {}
    points_sums = {}
    if terms:
        # Get selected term from GET or default to latest
        term_id = request.GET.get('term')
        if term_id:
            selected_term = terms.filter(id=term_id).first()
        if not selected_term:
            selected_term = terms.first()
        # Only show done exams for result slip
        from django.utils import timezone
        today = timezone.now().date()
        exams = Exam.objects.filter(term=selected_term, end_date__lte=today)
        # Get selected exam from GET or default to latest in term
        exam_id = request.GET.get('exam')
        if exam_id:
            selected_exam = exams.filter(id=exam_id).first()
        else:
            # Prefer latest done exam
            selected_exam = exams.order_by('-start_date').first()
        if selected_exam:
            # Build grades dict: grades[student.id][subject.id] = {'score': ..., 'grade_letter': ...}
            grades = {s.id: {} for s in students}
            grade_qs = Grade.objects.filter(student__in=students, exam=selected_exam)
            from .models import SubjectGradingScheme, SubjectComponent
            grading_schemes = {s.id: getattr(s, 'grading_scheme', None) for s in Subject.objects.all()}
            for g in grade_qs:
                scheme = grading_schemes.get(g.subject.id)
                grade_letter = None
                if scheme:
                    grade_letter = scheme.get_grade_letter(g.score)
                grades[g.student.id][g.subject.id] = {'score': g.score, 'grade_letter': grade_letter, 'remarks': getattr(g, 'remarks', '')}
            # Compute composite subject scores (e.g., English = Language + Composition)
            parent_components = {}
            for sc in SubjectComponent.objects.select_related('parent', 'child').all():
                parent_components.setdefault(sc.parent_id, []).append((sc.child_id, sc.weight))
            for s in students:
                for parent_id, comps in parent_components.items():
                    total = 0.0
                    has_any = False
                    for child_id, weight in comps:
                        child_entry = grades[s.id].get(child_id)
                        if child_entry and child_entry.get('score') is not None:
                            total += float(child_entry['score']) * float(weight)
                            has_any = True
                    if has_any:
                        scheme = grading_schemes.get(parent_id)
                        grade_letter = scheme.get_grade_letter(total) if scheme else None
                        grades[s.id][parent_id] = {'score': total, 'grade_letter': grade_letter, 'remarks': ''}
            # Compute totals, averages (optional), points, and ranks based on top-level subjects only
            top_ids = set(subj.id for subj in subjects)
            letter_points = {'A': 4, 'B': 3, 'C': 2, 'D': 1}
            for s in students:
                subj_scores = [v['score'] for sid, v in grades[s.id].items() if (sid in top_ids) and v.get('score') is not None]
                totals[s.id] = sum(subj_scores) if subj_scores else 0
                averages[s.id] = (totals[s.id] / len(subj_scores)) if subj_scores else 0
                # Compute total points from grade letters for top-level subjects
                letters = [v.get('grade_letter') for sid, v in grades[s.id].items() if sid in top_ids and v.get('grade_letter')]
                total_pts = 0
                for lt in letters:
                    key = str(lt).upper()[:1]
                    total_pts += letter_points.get(key, 0)
                points_sums[s.id] = total_pts
            # Rank students by totals (not averages)
            sorted_students = sorted(students, key=lambda s: totals.get(s.id, 0), reverse=True)
            for idx, s in enumerate(sorted_students, 1):
                ranks[s.id] = idx
            # Sort students by rank before passing to template
            students = sorted(students, key=lambda s: ranks.get(s.id, 9999))
    # --- Class-level comparison ---
    comparison_classes = []
    class_comparison_averages = {}
    if selected_exam and class_obj.level:
        peer_classes = Class.objects.filter(level=class_obj.level)
        for peer_class in peer_classes:
            peer_students = Student.objects.filter(class_group=peer_class)
            peer_grades = Grade.objects.filter(student__in=peer_students, exam=selected_exam)
            peer_scores = [g.score for g in peer_grades if g.score is not None]
            avg_score = sum(peer_scores) / len(peer_scores) if peer_scores else 0
            class_comparison_averages[peer_class.name] = avg_score
            comparison_classes.append(peer_class.name)
    # --- Overall student ranking across all classes at this level ---
    overall_students_ranked = []
    if selected_exam and class_obj.level:
        all_level_classes = Class.objects.filter(level=class_obj.level)
        all_level_students = Student.objects.filter(class_group__in=all_level_classes).select_related('user', 'class_group')
        student_averages = {}
        for s in all_level_students:
            grades_qs = Grade.objects.filter(student=s, exam=selected_exam)
            scores = [g.score for g in grades_qs if g.score is not None]
            avg = sum(scores) / len(scores) if scores else 0
            student_averages[s.id] = avg
        ranked_students = sorted(all_level_students, key=lambda s: student_averages.get(s.id, 0), reverse=True)
        for idx, s in enumerate(ranked_students, 1):
            overall_students_ranked.append({
                'rank': idx,
                'student': s,
                'class_name': s.class_group.name if s.class_group else '',
                'average': student_averages.get(s.id, 0)
            })
    class_performance = [
        {'name': cname, 'average': class_comparison_averages.get(cname, 0)}
        for cname in comparison_classes
    ]
    context = {
        'class_obj': class_obj,
        'students': students,
        'subjects': subjects,
        'terms': terms,
        'exams': exams,
        'selected_term': selected_term,
        'selected_exam': selected_exam,
        'grades': grades,
        'averages': averages,
        'totals': totals,
        'ranks': ranks,
        'points_sums': points_sums if 'points_sums' in locals() else {},
        'comparison_classes': comparison_classes,
        'class_comparison_averages': class_comparison_averages,
        'overall_students_ranked': overall_students_ranked,
        'class_performance': class_performance,
        'class_teacher_name': class_obj.class_teacher.user.get_full_name() if getattr(class_obj.class_teacher, 'user', None) else None,
    }
    return render(request, 'dashboards/admin_class_result_slip.html', context)

@login_required(login_url='login')
@cache_page(30)
def admin_block_result_slip(request):
    if request.user.role not in ('admin', 'teacher'):
        return redirect('login')
    # Aggregated result slips for all classes in a given level
    from .models import Class, Student, Subject, Exam, Grade, Term, FeeAssignment, FeePayment, FeeCategory, SubjectComponent, SubjectGradingScheme, ResultBlock
    from django.db.models import Sum
    from django.core.cache import cache
    level = request.GET.get('level')
    classes = Class.objects.none()
    if level is not None and level != "":
        classes = Class.objects.filter(level=level).order_by('name')
    # Use top-level subjects only (exclude component/child subjects), but restrict to those taught
    subjects = Subject.objects.filter(part_of__isnull=True)
    if classes.exists():
        subjects = subjects.filter(classes__in=classes).distinct()
    subjects = subjects.order_by('name')
    terms = Term.objects.order_by('-start_date')
    selected_term = None
    selected_exam = None
    grades = {}
    averages = {}
    totals = {}
    ranks = {}
    points_dict = {}
    students_by_class = {}
    all_students_ranked = []
    student_balances = {}
    balance_min_param = request.GET.get('balance_min')
    try:
        balance_min = float(balance_min_param) if balance_min_param not in (None, "") else None
    except ValueError:
        balance_min = None
    only_blocked = str(request.GET.get('only_blocked', '')).lower() in ('1','true','yes','on')
    # Persist/clear runtime block bar in cache (admins only)
    is_admin = (request.user.role == 'admin')
    if is_admin:
        if balance_min is not None:
            cache.set('results_bar_enabled', True, timeout=None)
            cache.set('results_bar_balance_min', balance_min, timeout=None)
        else:
            cache.delete('results_bar_enabled')
            cache.delete('results_bar_balance_min')
    if terms:
        # Term selection
        term_id = request.GET.get('term')
        if term_id:
            selected_term = terms.filter(id=term_id).first()
        if not selected_term:
            selected_term = terms.first()
        exams = Exam.objects.filter(term=selected_term)
        # Exam selection
        exam_id = request.GET.get('exam')
        if exam_id:
            selected_exam = exams.filter(id=exam_id).first()
        if not selected_exam:
            selected_exam = exams.order_by('-start_date').first()
        if selected_exam and classes.exists():
            # Collect all students in the level
            all_students = (
                Student.objects.filter(class_group__in=classes)
                .select_related('user', 'class_group')
            )
            # Build grades dict: grades[student.id][subject.id] = {score, grade_letter}
            grades = {s.id: {} for s in all_students}
            grade_qs = Grade.objects.filter(student__in=all_students, exam=selected_exam)
            for g in grade_qs:
                # Many schemas store grade letter on Grade model already; if not present, leave None
                grade_letter = getattr(g, 'grade_letter', None)
                grades[g.student_id][g.subject_id] = {'score': g.score, 'grade_letter': grade_letter}
            # Aggregate composite (parent) subjects from their child components if direct grade missing
            # Preload components per parent
            parent_components = {
                p.id: list(SubjectComponent.objects.filter(parent=p).select_related('child'))
                for p in subjects
            }
            # Preload grading schemes for parents
            grading_by_parent = {
                sg.subject_id: sg for sg in SubjectGradingScheme.objects.filter(subject__in=subjects)
            }
            for s in all_students:
                for parent in subjects:
                    # Skip if we already have a direct grade for the parent
                    if grades.get(s.id, {}).get(parent.id):
                        continue
                    comps = parent_components.get(parent.id, [])
                    if not comps:
                        continue
                    total = 0.0
                    have_any = False
                    for sc in comps:
                        child_val = grades.get(s.id, {}).get(sc.child_id)
                        if child_val and child_val.get('score') is not None:
                            have_any = True
                            w = sc.weight if sc.weight is not None else 1.0
                            total += float(child_val['score']) * float(w)
                    if have_any:
                        # Derive grade letter from parent's grading scheme if available
                        letter = None
                        sg = grading_by_parent.get(parent.id)
                        if sg:
                            try:
                                letter = sg.get_grade_letter(total)
                            except Exception:
                                letter = None
                        grades[s.id][parent.id] = {'score': total, 'grade_letter': letter}
            # Compute totals and averages per student using top-level subjects only
            for s in all_students:
                top_ids = set(subj.id for subj in subjects)
                subj_scores = [val['score'] for sid, val in grades[s.id].items() if sid in top_ids and val.get('score') is not None]
                totals[s.id] = sum(subj_scores) if subj_scores else 0
                averages[s.id] = (totals[s.id] / len(subj_scores)) if subj_scores else 0
            # Compute total points per student from grade letters (A=4, B=3, C=2, D=1)
            letter_points = {'A': 4, 'B': 3, 'C': 2, 'D': 1}
            for s in all_students:
                top_ids = set(subj.id for subj in subjects)
                letters = [val.get('grade_letter') for sid, val in grades[s.id].items() if sid in top_ids and val.get('grade_letter')]
                total_pts = 0
                for lt in letters:
                    key = str(lt).upper()[:1]
                    total_pts += letter_points.get(key, 0)
                points_dict[s.id] = total_pts
            # Compute balances similar to admin_fees view, but in bulk to avoid per-student queries
            all_fee_categories = list(FeeCategory.objects.all())
            if selected_term:
                previous_terms = Term.objects.filter(start_date__lt=selected_term.start_date).order_by('start_date')
                last_term = previous_terms.last() if previous_terms.exists() else None
            else:
                previous_terms = Term.objects.all().order_by('start_date')
                last_term = previous_terms.last() if previous_terms.exists() else None

            # Pre-aggregate amounts to cut N+1 queries
            student_ids = list(all_students.values_list('id', flat=True))
            class_ids = list(classes.values_list('id', flat=True))

            # All-time paid per student
            paid_by_student = {
                row['student_id']: float(row['total'] or 0)
                for row in (
                    FeePayment.objects
                    .filter(student_id__in=student_ids)
                    .values('student_id')
                    .annotate(total=Sum('amount_paid'))
                )
            }

            # Current term billed per class
            billed_current_by_class = {
                row['class_group_id']: float(row['total'] or 0)
                for row in (
                    FeeAssignment.objects
                    .filter(class_group_id__in=class_ids, term=selected_term) if selected_term else FeeAssignment.objects.filter(class_group_id__in=class_ids)
                ).values('class_group_id').annotate(total=Sum('amount'))
            }

            # Previous terms billed per class
            billed_prev_by_class = {}
            paid_prev_by_student = {}
            if previous_terms.exists():
                billed_prev_by_class = {
                    row['class_group_id']: float(row['total'] or 0)
                    for row in (
                        FeeAssignment.objects
                        .filter(class_group_id__in=class_ids, term__in=previous_terms)
                        .values('class_group_id')
                        .annotate(total=Sum('amount'))
                    )
                }
                paid_prev_by_student = {
                    row['student_id']: float(row['total'] or 0)
                    for row in (
                        FeePayment.objects
                        .filter(student_id__in=student_ids, fee_assignment__term__in=previous_terms)
                        .values('student_id')
                        .annotate(total=Sum('amount_paid'))
                    )
                }

            # Optional last term extra outstanding component
            last_billed_by_class = {}
            last_paid_by_student = {}
            if last_term and not previous_terms.filter(pk=last_term.pk).exists():
                last_billed_by_class = {
                    row['class_group_id']: float(row['total'] or 0)
                    for row in (
                        FeeAssignment.objects
                        .filter(class_group_id__in=class_ids, term=last_term)
                        .values('class_group_id')
                        .annotate(total=Sum('amount'))
                    )
                }
                last_paid_by_student = {
                    row['student_id']: float(row['total'] or 0)
                    for row in (
                        FeePayment.objects
                        .filter(student_id__in=student_ids, fee_assignment__term=last_term)
                        .values('student_id')
                        .annotate(total=Sum('amount_paid'))
                    )
                }

            # Final per-student balance computation using dictionaries
            for s in all_students:
                total_billed = billed_current_by_class.get(getattr(s.class_group, 'id', None), 0.0)
                paid_total = paid_by_student.get(s.id, 0.0)
                outstanding = 0.0
                if previous_terms.exists():
                    prev_billed = billed_prev_by_class.get(getattr(s.class_group, 'id', None), 0.0)
                    prev_paid = paid_prev_by_student.get(s.id, 0.0)
                    outstanding += (prev_billed - prev_paid)
                if last_term and not previous_terms.filter(pk=last_term.pk).exists():
                    last_billed = last_billed_by_class.get(getattr(s.class_group, 'id', None), 0.0)
                    last_paid = last_paid_by_student.get(s.id, 0.0)
                    outstanding += (last_billed - last_paid)
                student_balances[s.id] = (total_billed + outstanding - paid_total)
            # Global ranks across the level
            ranked = sorted(all_students, key=lambda s: averages.get(s.id, 0), reverse=True)
            # Threshold source (use request param first, else active cached bar)
            from django.core.cache import cache
            threshold = balance_min
            if threshold is None:
                if cache.get('results_bar_enabled'):
                    try:
                        threshold = float(cache.get('results_bar_balance_min'))
                    except (TypeError, ValueError):
                        threshold = None
            # KPIs
            balances_list = [float(student_balances.get(s.id, 0) or 0) for s in all_students]
            blocked_count = 0
            avg_balance = None
            median_balance = None
            try:
                if balances_list:
                    blocked_count = sum(1 for b in balances_list if (threshold is not None and b >= threshold)) if threshold is not None else 0
                    import statistics
                    avg_balance = statistics.fmean(balances_list)
                    median_balance = statistics.median(balances_list)
            except Exception:
                pass
            # Persist blocks when admin applies a threshold for a specific exam
            if is_admin and selected_exam and threshold is not None:
                try:
                    # Build set of student IDs to be blocked
                    to_block_ids = {s.id for s in all_students if float(student_balances.get(s.id, 0) or 0) >= float(threshold)}
                    # Upsert blocks
                    existing = {rb.student_id: rb for rb in ResultBlock.objects.filter(exam=selected_exam, student_id__in=list(all_students.values_list('id', flat=True)))}
                    from django.utils import timezone
                    # Activate or create blocks
                    for sid in to_block_ids:
                        rb = existing.get(sid)
                        if rb:
                            if not rb.active:
                                rb.active = True
                                rb.reason = rb.reason or 'Blocked by fee threshold'
                                rb.balance_threshold = threshold
                                rb.cleared_at = None
                                rb.save(update_fields=['active','reason','balance_threshold','cleared_at'])
                        else:
                            ResultBlock.objects.create(student_id=sid, exam=selected_exam, active=True, reason='Blocked by fee threshold', balance_threshold=threshold, created_by=request.user)
                    # Deactivate blocks for students now below threshold
                    for sid, rb in existing.items():
                        if sid not in to_block_ids and rb.active:
                            rb.active = False
                            rb.cleared_at = timezone.now()
                            rb.save(update_fields=['active','cleared_at'])
                except Exception:
                    # Non-fatal: if persistence fails, UI still works with cached threshold
                    pass
            # Optional: filter by threshold (explicit balance_min) and/or only_blocked
            if balance_min is not None:
                ranked = [s for s in ranked if student_balances.get(s.id, 0) >= balance_min]
            elif only_blocked and threshold is not None:
                ranked = [s for s in ranked if student_balances.get(s.id, 0) >= threshold]
            for idx, s in enumerate(ranked, 1):
                ranks[s.id] = idx
            # Expose a flat ranked list across the level
            all_students_ranked = ranked
            # Split students by class and sort by rank
            for c in classes:
                cls_students = [s for s in ranked if getattr(s.class_group, 'id', None) == c.id]
                students_by_class[c.id] = cls_students
    context = {
        'level': level,
        'classes': list(classes) if classes is not None else [],
        'students_by_class': students_by_class,
        'all_students_ranked': all_students_ranked,
        'subjects': subjects,
        'selected_term': selected_term,
        'selected_exam': selected_exam,
        'grades': grades,
        'averages': averages,
        'totals': totals,
        'ranks': ranks,
        # points_sums optional; template tolerates missing entries
        'points_sums': {},
        'points_dict': points_dict,
        'student_balances': student_balances,
        'balance_min': balance_min,
        'only_blocked': only_blocked,
        'blocked_count': locals().get('blocked_count', 0),
        'avg_balance': locals().get('avg_balance', None),
        'median_balance': locals().get('median_balance', None),
        'effective_threshold': locals().get('threshold', None),
    }
    return render(request, 'dashboards/admin_block_result_slip.html', context)
@login_required(login_url='login')
def teacher_class_result_slip(request, class_id):
    if request.user.role != 'teacher':
        return redirect('login')
    from .models import Class, Student, Subject, Exam, Grade, Term
    class_obj = get_object_or_404(Class, id=class_id)
    students = Student.objects.filter(class_group=class_obj).select_related('user')
    subjects = Subject.objects.all()
    terms = Term.objects.order_by('-start_date')
    exams = Exam.objects.none()
    selected_term = None
    selected_exam = None
    grades = {}
    averages = {}
    ranks = {}
    if terms:
        # Get selected term from GET or default to latest
        term_id = request.GET.get('term')
        if term_id:
            selected_term = terms.filter(id=term_id).first()
        if not selected_term:
            selected_term = terms.first()
        # Only show done exams for result slip
        from django.utils import timezone
        today = timezone.now().date()
        exams = Exam.objects.filter(term=selected_term, end_date__lte=today)
        # Get selected exam from GET or default to latest in term
        exam_id = request.GET.get('exam')
        if exam_id:
            selected_exam = exams.filter(id=exam_id).first()
        if not selected_exam:
            selected_exam = exams.order_by('-start_date').first()
        if selected_exam:
            # Build grades dict: grades[student.id][subject.id] = Grade instance
            all_level_classes = Class.objects.filter(level=class_obj.level)
            all_level_students = Student.objects.filter(class_group__in=all_level_classes).select_related('user', 'class_group')
            grades = {s.id: {} for s in all_level_students}
            grade_qs = Grade.objects.filter(student__in=all_level_students, exam=selected_exam)
            for g in grade_qs:
                grades[g.student.id][g.subject.id] = g
            # Compute averages and ranks
            for s in students:
                subj_scores = [g.score for g in grades[s.id].values() if hasattr(g, 'score') and g.score is not None]
                averages[s.id] = sum(subj_scores) / len(subj_scores) if subj_scores else 0
            sorted_students = sorted(students, key=lambda s: averages[s.id], reverse=True)
            for idx, s in enumerate(sorted_students, 1):
                ranks[s.id] = idx
            # Sort students by rank before passing to template
            students = sorted(students, key=lambda s: ranks.get(s.id, 9999))
    # --- Class-level comparison ---
    comparison_classes = []
    class_comparison_averages = {}
    if selected_exam and class_obj.level:
        peer_classes = Class.objects.filter(level=class_obj.level)
        for peer_class in peer_classes:
            peer_students = Student.objects.filter(class_group=peer_class)
            peer_grades = Grade.objects.filter(student__in=peer_students, exam=selected_exam)
            peer_scores = [g.score for g in peer_grades if g.score is not None]
            avg_score = sum(peer_scores) / len(peer_scores) if peer_scores else 0
            class_comparison_averages[peer_class.name] = avg_score
            comparison_classes.append(peer_class.name)
    # --- Overall student ranking across all classes at this level ---
    overall_students_ranked = []
    if selected_exam and class_obj.level:
        all_level_classes = Class.objects.filter(level=class_obj.level)
        all_level_students = Student.objects.filter(class_group__in=all_level_classes).select_related('user', 'class_group')
        student_averages = {}
        for s in all_level_students:
            grades_qs = Grade.objects.filter(student=s, exam=selected_exam)
            scores = [g.score for g in grades_qs if g.score is not None]
            avg = sum(scores) / len(scores) if scores else 0
            student_averages[s.id] = avg
        ranked_students = sorted(all_level_students, key=lambda s: student_averages.get(s.id, 0), reverse=True)
        for idx, s in enumerate(ranked_students, 1):
            overall_students_ranked.append({
                'rank': idx,
                'student': s,
                'class_name': s.class_group.name if s.class_group else '',
                'average': student_averages.get(s.id, 0)
            })
    class_performance = [
        {'name': cname, 'average': class_comparison_averages.get(cname, 0)}
        for cname in comparison_classes
    ]
    context = {
        'class_obj': class_obj,
        'students': students,
        'subjects': subjects,
        'terms': terms,
        'exams': exams,
        'selected_term': selected_term,
        'selected_exam': selected_exam,
        'grades': grades,
        'averages': averages,
        'ranks': ranks,
        'comparison_classes': comparison_classes,
        'class_comparison_averages': class_comparison_averages,
        'overall_students_ranked': overall_students_ranked,
        'class_performance': class_performance,
    }
    return render(request, 'dashboards/admin_class_result_slip.html', context)

@login_required(login_url='login')
def overall_student_results(request, class_id):
    if request.user.role != 'admin':
        return redirect('login')
    from .models import Class, Student, Subject, Exam, Grade
    class_obj = get_object_or_404(Class, id=class_id)
    level = class_obj.level
    subjects = Subject.objects.all()
    # Use same term/exam selection logic as result slip
    terms = Term.objects.order_by('-start_date')
    selected_term = None
    selected_exam = None
    if terms:
        term_id = request.GET.get('term')
        if term_id:
            selected_term = terms.filter(id=term_id).first()
        if not selected_term:
            selected_term = terms.first()
        # Only show done exams for overall results
        from django.utils import timezone
        today = timezone.now().date()
        exams = Exam.objects.filter(term=selected_term, end_date__lte=today)
        exam_id = request.GET.get('exam')
        if exam_id:
            selected_exam = exams.filter(id=exam_id).first()
        if not selected_exam:
            selected_exam = exams.order_by('-start_date').first()
    overall_students_ranked = []
    if selected_exam and level:
        all_level_classes = Class.objects.filter(level=level)
        all_level_students = Student.objects.filter(class_group__in=all_level_classes).select_related('user', 'class_group')
        student_averages = {}
        student_grades = {}
        for s in all_level_students:
            grades_qs = Grade.objects.filter(student=s, exam=selected_exam)
            student_grades[s.id] = {g.subject.id: g for g in grades_qs}
            scores = [g.score for g in grades_qs if g.score is not None]
            avg = sum(scores) / len(scores) if scores else 0
            student_averages[s.id] = avg
        ranked_students = sorted(all_level_students, key=lambda s: student_averages.get(s.id, 0), reverse=True)
        for idx, s in enumerate(ranked_students, 1):
            overall_students_ranked.append({
                'rank': idx,
                'student': s,
                'class_name': s.class_group.name if s.class_group else '',
                'grades': student_grades.get(s.id, {}),
                'average': student_averages.get(s.id, 0)
            })
    context = {
        'level': level,
        'subjects': subjects,
        'selected_exam': selected_exam,
        'overall_students_ranked': overall_students_ranked,
        'back_url': request.META.get('HTTP_REFERER', '/')
    }
    return render(request, 'dashboards/overall_student_results.html', context)

@login_required(login_url='login')
def admin_academic_years(request):
    from django.contrib import messages
    if request.user.role != 'admin':
        return redirect('login')

    edit_term_id = None
    edit_term_form = None

    if request.method == 'POST':
        if 'show_edit_term' in request.POST:
            edit_term_id = request.POST.get('term_id')
            if edit_term_id:
                try:
                    term = Term.objects.get(id=edit_term_id)
                    edit_term_form = EditTermDatesForm(instance=term)
                except Term.DoesNotExist:
                    messages.error(request, "Term not found.")
        elif 'add_year' in request.POST:
            year = request.POST.get('year')
            if year:
                # Check if the year already exists (case insensitive)
                if AcademicYear.objects.filter(year__iexact=year).exists():
                    messages.error(request, f"Academic Year '{year}' already exists. If you want to add terms, use the form below.")
                    return redirect('admin_academic_years')
                academic_year_obj = AcademicYear.objects.create(year=year)
                # Auto-create terms if none exist for this year
                import re
                year_match = re.match(r'(\d{4})', year)
                if year_match:
                    base_year = int(year_match.group(1))
                    # Define terms and holidays
                    import datetime
                    terms = [
                        ('Term 1', datetime.date(base_year, 1, 1), datetime.date(base_year, 3, 31)),
                        ('Term 2', datetime.date(base_year, 5, 1), datetime.date(base_year, 7, 31)),
                        ('Term 3', datetime.date(base_year, 9, 1), datetime.date(base_year, 11, 30)),
                    ]
                    for name, start, end in terms:
                        Term.objects.create(name=name, academic_year=academic_year_obj, start_date=start, end_date=end)
                messages.success(request, f'Academic Year {year} added.')
                return redirect('admin_academic_years')
        elif 'add_term' in request.POST:
            year_id = request.POST.get('year_id')
            term_name = request.POST.get('term_name')
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            if year_id and term_name and start_date and end_date:
                academic_year = AcademicYear.objects.get(id=year_id)
                Term.objects.get_or_create(
                    name=term_name,
                    academic_year=academic_year,
                    defaults={
                        'start_date': start_date,
                        'end_date': end_date
                    }
                )
                messages.success(request, f'Term {term_name} added to {academic_year.year}.')
                return redirect('admin_academic_years')
        elif 'edit_term_dates' in request.POST:
            edit_term_id = request.POST.get('term_id')
            term = Term.objects.get(id=edit_term_id)
            edit_term_form = EditTermDatesForm(request.POST, instance=term)
            if edit_term_form.is_valid():
                edit_term_form.save()
                messages.success(request, f"Updated dates for {term.name}.")
                return redirect('admin_academic_years')
            else:
                messages.error(request, "Failed to update term dates. Please check the form.")
        elif 'run_promotion' in request.POST:
            # Trigger the promote_students management command
            import subprocess
            import sys
            result = subprocess.run([sys.executable, 'manage.py', 'promote_students'], capture_output=True, text=True)
            if result.returncode == 0:
                messages.success(request, 'Promotion and graduation process completed successfully.')
                messages.warning(request, 'Promotion completed. Please review and update class names if needed.')
            else:
                messages.error(request, 'Promotion failed: ' + result.stderr)
            return redirect('admin_academic_years')

    import datetime
    today = datetime.date.today()
    # Only show years with at least one term whose end_date is today or in the future
    all_years = AcademicYear.objects.prefetch_related('terms').all().order_by('-year')
    academic_years = []
    for year in all_years:
        terms = year.terms.all()
        if any(term.end_date and term.end_date >= today for term in terms):
            academic_years.append(year)

    # Find the current term based on today's date
    current_term = None
    current_year = None
    graduation_ready = False
    graduation_year = None
    for year in academic_years:
        terms = sorted(year.terms.all(), key=lambda t: t.start_date or datetime.date.min)
        for idx, term in enumerate(terms):
            if term.start_date and term.end_date:
                if term.start_date <= today <= term.end_date:
                    current_term = term
                    current_year = year
                    break
        # Graduation: after last term's end date
        if terms and all(t.end_date for t in terms):
            last_term = terms[-1]
            if last_term.end_date < today:
                graduation_ready = True
                graduation_year = year
    context = {
        'academic_years': academic_years,
        'current_term': current_term,
        'current_year': current_year,
        'graduation_ready': graduation_ready,
        'graduation_year': graduation_year,
        'edit_term_id': edit_term_id,
        'edit_term_form': edit_term_form,
    }
    return render(request, 'dashboards/admin_academic_years.html', context)

import json
from .forms import AddSubjectForm
from django.db.models import Avg

from django.views.generic import ListView
from django.db.models import Q
from .models import FeePayment, Term, Class, FeeCategory, Subject, Grade, Student, FeeAssignment
from .forms import FeePaymentForm

class AdminAnalyticsView(ListView):
    model = FeePayment
    template_name = 'dashboards/admin_analytics.html'
    context_object_name = 'fee_payments'
    paginate_by = 25

    def get_queryset(self):
        queryset = FeePayment.objects.select_related('student__user', 'fee_assignment__fee_category', 'fee_assignment__class_group', 'fee_assignment__term')
        term = self.request.GET.get('term')
        class_group = self.request.GET.get('class_group')
        fee_category = self.request.GET.get('fee_category')
        search = self.request.GET.get('search')
        if term:
            queryset = queryset.filter(fee_assignment__term__id=term)
        if class_group:
            queryset = queryset.filter(fee_assignment__class_group__id=class_group)
        if fee_category:
            queryset = queryset.filter(fee_assignment__fee_category__id=fee_category)
        if search:
            queryset = queryset.filter(
                Q(student__user__username__icontains=search) |
                Q(student__admission_no__icontains=search) |
                Q(fee_assignment__fee_category__name__icontains=search) |
                Q(fee_assignment__class_group__name__icontains=search) |
                Q(amount_paid__icontains=search) |
                Q(reference__icontains=search)
            )
        return queryset.order_by('-payment_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['terms'] = Term.objects.all()
        context['class_groups'] = Class.objects.all()
        context['fee_categories'] = FeeCategory.objects.all()
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_term'] = self.request.GET.get('term', '')
        context['selected_class'] = self.request.GET.get('class_group', '')
        context['selected_category'] = self.request.GET.get('fee_category', '')
        context['add_form'] = FeePaymentForm()

        # --- Analytics calculations ---
        today = timezone.now().date()
        current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()
        context['current_term'] = current_term

        # Average performance per subject (across all grades)
        from django.db.models import Avg
        avg_performance = []
        subject_labels = []
        avg_scores = []
        subjects = Subject.objects.all()
        for subject in subjects:
            avg = Grade.objects.filter(subject=subject).aggregate(avg_score=Avg('score'))['avg_score']
            if avg is not None:
                avg_performance.append({'subject': subject.name, 'avg_score': avg})
                subject_labels.append(subject.name)
                avg_scores.append(round(avg, 2))
        context['avg_performance'] = avg_performance
        context['subject_labels'] = json.dumps(subject_labels)
        context['avg_scores'] = json.dumps(avg_scores)

        # Fee analytics
        fee_assignments = FeeAssignment.objects.all()
        total_assigned = 0
        for assignment in fee_assignments:
            num_students = Student.objects.filter(class_group=assignment.class_group).count()
            total_assigned += float(assignment.amount) * num_students
        total_paid = FeePayment.objects.aggregate(total=Sum('amount_paid'))['total'] or 0
        payment_percentage = (float(total_paid) / float(total_assigned) * 100) if total_assigned else 0
        context['total_assigned'] = total_assigned
        context['total_paid'] = total_paid
        context['payment_percentage'] = round(payment_percentage, 2)
        # Pie chart: Paid vs Unpaid
        paid_value = float(total_paid)
        unpaid_value = float(total_assigned) - float(total_paid)
        context['pie_labels'] = json.dumps(["Paid", "Unpaid"])
        context['pie_data'] = json.dumps([paid_value, unpaid_value if unpaid_value > 0 else 0])
        # Bar chart: Payments over time (by month)
        from collections import defaultdict
        payments = FeePayment.objects.filter(fee_assignment__in=fee_assignments)
        monthly_totals = defaultdict(float)
        for payment in payments:
            if payment.payment_date:
                month = payment.payment_date.strftime('%Y-%m')
                monthly_totals[month] += float(payment.amount_paid)
        months_sorted = sorted(monthly_totals.keys())
        context['bar_labels'] = json.dumps(months_sorted)
        context['bar_data'] = json.dumps([monthly_totals[m] for m in months_sorted])

        # Payments by method (Cash/Bank/Mpesa Paybill)
        from collections import defaultdict as _dd
        method_totals = _dd(float)
        method_qs = FeePayment.objects.all()
        if current_term:
            method_qs = method_qs.filter(fee_assignment__term=current_term)
        for p in method_qs:
            label = p.payment_method or 'Unknown'
            method_totals[label] += float(p.amount_paid)
        method_labels = list(method_totals.keys()) or ['No Data']
        method_data = [method_totals[k] for k in method_labels] or [0]
        context['method_labels'] = json.dumps(method_labels)
        context['method_data'] = json.dumps(method_data)

        # Outstanding by class (Assigned - Paid) for current term
        class_outstanding = []
        classes = Class.objects.all()
        for c in classes:
            assignments = FeeAssignment.objects.filter(class_group=c)
            if current_term:
                assignments = assignments.filter(term=current_term)
            assigned_total = 0.0
            for fa in assignments:
                num_students = Student.objects.filter(class_group=c).count()
                assigned_total += float(getattr(fa, 'amount', 0)) * num_students
            paid_total = FeePayment.objects.filter(fee_assignment__in=assignments).aggregate(total=Sum('amount_paid'))['total'] or 0.0
            outstanding = float(assigned_total) - float(paid_total)
            if outstanding > 0:
                class_outstanding.append((c.name, outstanding))
        # Top 10 by outstanding
        class_outstanding.sort(key=lambda t: t[1], reverse=True)
        top_labels = [t[0] for t in class_outstanding[:10]] or ['No Data']
        top_values = [t[1] for t in class_outstanding[:10]] or [0]
        context['class_out_labels'] = json.dumps(top_labels)
        context['class_out_data'] = json.dumps(top_values)

        # Additional visuals: Payments by Category (current term)
        from collections import defaultdict as __dd
        category_totals = __dd(float)
        category_qs = FeePayment.objects.select_related('fee_assignment__fee_category')
        if current_term:
            category_qs = category_qs.filter(fee_assignment__term=current_term)
        for p in category_qs:
            cat = getattr(getattr(p.fee_assignment, 'fee_category', None), 'name', 'Unknown')
            category_totals[cat] += float(p.amount_paid)
        category_labels = list(category_totals.keys()) or ['No Data']
        category_data = [category_totals[k] for k in category_labels] or [0]
        context['category_labels'] = json.dumps(category_labels)
        context['category_data'] = json.dumps(category_data)

        # Additional visuals: Daily Payments (last 14 days)
        from datetime import timedelta
        start_14 = today - timedelta(days=13)
        daily_qs = FeePayment.objects.filter(payment_date__date__gte=start_14, payment_date__date__lte=today)
        daily_map = __dd(float)
        for p in daily_qs:
            key = p.payment_date.date().isoformat() if p.payment_date else None
            if key:
                daily_map[key] += float(p.amount_paid)
        daily_labels = [(start_14 + timedelta(days=i)).isoformat() for i in range(14)]
        daily_data = [daily_map[d] for d in daily_labels]
        context['daily_labels'] = json.dumps(daily_labels)
        context['daily_data'] = json.dumps(daily_data)

        # Additional visuals: Cumulative Collections in current term
        cum_labels, cum_paid = [], []
        if current_term:
            cur = current_term.start_date
            cum_total = 0.0
            # Pre-aggregate per-date totals for efficiency
            term_qs = FeePayment.objects.filter(
                fee_assignment__term=current_term,
                payment_date__date__gte=current_term.start_date,
                payment_date__date__lte=today,
            )
            per_day = __dd(float)
            for p in term_qs:
                k = p.payment_date.date().isoformat() if p.payment_date else None
                if k:
                    per_day[k] += float(p.amount_paid)
            while cur <= today:
                key = cur.isoformat()
                cum_total += per_day[key]
                cum_labels.append(key)
                cum_paid.append(cum_total)
                cur += timedelta(days=1)
        context['cum_labels'] = json.dumps(cum_labels)
        context['cum_paid'] = json.dumps(cum_paid)

        return context

from django.http import JsonResponse
from django.views.decorators.http import require_GET

@login_required(login_url='login')
@require_GET
def admin_analytics_data(request):
    """Return analytics datasets as JSON for dynamic filtering and comparisons.
    Params (GET):
      term: term id (optional)
      class_group: class id (optional)
      fee_category: category id (optional)
      compare_prev: '1' to include previous-term comparison series
    """
    from collections import defaultdict
    from datetime import timedelta
    # Filters
    term_id = request.GET.get('term')
    class_id = request.GET.get('class_group')
    category_id = request.GET.get('fee_category')
    compare_prev = request.GET.get('compare_prev') == '1'

    today = timezone.now().date()
    current_term = None
    if term_id:
        current_term = Term.objects.filter(id=term_id).first()
    if not current_term:
        current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()

    fa_qs = FeeAssignment.objects.all()
    if current_term:
        fa_qs = fa_qs.filter(term=current_term)
    if class_id:
        fa_qs = fa_qs.filter(class_group_id=class_id)
    if category_id:
        fa_qs = fa_qs.filter(fee_category_id=category_id)

    # Totals assigned
    total_assigned = 0.0
    for fa in fa_qs:
        num_students = Student.objects.filter(class_group=fa.class_group).count()
        total_assigned += float(getattr(fa, 'amount', 0)) * num_students

    # Payments filtered
    pay_qs = FeePayment.objects.select_related('student', 'fee_assignment', 'fee_assignment__term', 'fee_assignment__class_group', 'fee_assignment__fee_category')
    if current_term:
        pay_qs = pay_qs.filter(fee_assignment__term=current_term)
    if class_id:
        pay_qs = pay_qs.filter(fee_assignment__class_group_id=class_id)
    if category_id:
        pay_qs = pay_qs.filter(fee_assignment__fee_category_id=category_id)

    total_paid = float(pay_qs.aggregate(total=Sum('amount_paid'))['total'] or 0.0)
    unpaid_value = max(total_assigned - total_paid, 0.0)

    # Pie: Paid vs Unpaid
    pie = {
        'labels': ["Paid", "Unpaid"],
        'data': [total_paid, unpaid_value],
    }

    # Payments over time (monthly)
    monthly_totals = defaultdict(float)
    for p in pay_qs:
        if p.payment_date:
            m = p.payment_date.strftime('%Y-%m')
            monthly_totals[m] += float(p.amount_paid)
    months_sorted = sorted(monthly_totals.keys())
    bar = {
        'labels': months_sorted,
        'data': [monthly_totals[m] for m in months_sorted],
    }

    # Payments by method
    method_totals = defaultdict(float)
    for p in pay_qs:
        method_totals[p.payment_method or 'Unknown'] += float(p.amount_paid)
    method = {
        'labels': list(method_totals.keys()) or ['No Data'],
        'data': [method_totals[k] for k in (list(method_totals.keys()) or ['No Data'])] or [0],
    }

    # Outstanding by class (Top 10)
    class_outstanding = []
    classes = Class.objects.all()
    if class_id:
        classes = classes.filter(id=class_id)
    for c in classes:
        fa_cls = FeeAssignment.objects.filter(class_group=c)
        if current_term:
            fa_cls = fa_cls.filter(term=current_term)
        if category_id:
            fa_cls = fa_cls.filter(fee_category_id=category_id)
        assigned_total = 0.0
        num_students = Student.objects.filter(class_group=c).count()
        for fa in fa_cls:
            assigned_total += float(getattr(fa, 'amount', 0)) * num_students
        paid_total = FeePayment.objects.filter(fee_assignment__in=fa_cls).aggregate(total=Sum('amount_paid'))['total'] or 0.0
        outstanding = float(assigned_total) - float(paid_total)
        if outstanding > 0:
            class_outstanding.append((c.name, outstanding))
    class_outstanding.sort(key=lambda t: t[1], reverse=True)
    class_out = {
        'labels': [t[0] for t in class_outstanding[:10]] or ['No Data'],
        'data': [t[1] for t in class_outstanding[:10]] or [0],
    }

    # Payments by category (current filter scope)
    cat_totals = defaultdict(float)
    for p in pay_qs:
        cat = getattr(getattr(p.fee_assignment, 'fee_category', None), 'name', 'Unknown')
        cat_totals[cat] += float(p.amount_paid)
    category = {
        'labels': list(cat_totals.keys()) or ['No Data'],
        'data': [cat_totals[k] for k in (list(cat_totals.keys()) or ['No Data'])] or [0],
    }

    # Daily payments last 14 days
    start_14 = today - timedelta(days=13)
    daily_map = defaultdict(float)
    for p in pay_qs.filter(payment_date__date__gte=start_14, payment_date__date__lte=today):
        d = p.payment_date.date().isoformat() if p.payment_date else None
        if d:
            daily_map[d] += float(p.amount_paid)
    daily_labels = [(start_14 + timedelta(days=i)).isoformat() for i in range(14)]
    daily = {
        'labels': daily_labels,
        'data': [daily_map[d] for d in daily_labels],
    }

    # Cumulative for current term
    cum_labels, cum_paid = [], []
    if current_term:
        cur = current_term.start_date
        per_day = defaultdict(float)
        for p in pay_qs.filter(payment_date__date__gte=current_term.start_date, payment_date__date__lte=today):
            k = p.payment_date.date().isoformat() if p.payment_date else None
            if k:
                per_day[k] += float(p.amount_paid)
        total = 0.0
        while cur <= today:
            key = cur.isoformat()
            total += per_day[key]
            cum_labels.append(key)
            cum_paid.append(total)
            cur += timedelta(days=1)

    # Previous term comparison
    cum_prev = {'labels': [], 'data': []}
    bar_prev = {'labels': [], 'data': []}
    if compare_prev and current_term:
        prev_term = Term.objects.filter(end_date__lt=current_term.start_date).order_by('-end_date').first()
        if prev_term:
            pay_prev = FeePayment.objects.filter(fee_assignment__term=prev_term)
            if class_id:
                pay_prev = pay_prev.filter(fee_assignment__class_group_id=class_id)
            if category_id:
                pay_prev = pay_prev.filter(fee_assignment__fee_category_id=category_id)
            # monthly
            m_prev = defaultdict(float)
            for p in pay_prev:
                if p.payment_date:
                    m = p.payment_date.strftime('%Y-%m')
                    m_prev[m] += float(p.amount_paid)
            months_prev = sorted(m_prev.keys())
            bar_prev = {'labels': months_prev, 'data': [m_prev[m] for m in months_prev]}
            # cumulative over prev term
            per_day_prev = defaultdict(float)
            cur = prev_term.start_date
            end = prev_term.end_date
            for p in pay_prev.filter(payment_date__date__gte=prev_term.start_date, payment_date__date__lte=prev_term.end_date):
                k = p.payment_date.date().isoformat() if p.payment_date else None
                if k:
                    per_day_prev[k] += float(p.amount_paid)
            labels_prev, data_prev = [], []
            running = 0.0
            while cur <= end:
                k = cur.isoformat()
                running += per_day_prev[k]
                labels_prev.append(k)
                data_prev.append(running)
                cur += timedelta(days=1)
            cum_prev = {'labels': labels_prev, 'data': data_prev}

    # Top debtors in selected class (top 10)
    top_debtors = []
    if class_id:
        cls = Class.objects.filter(id=class_id).first()
        if cls:
            students = Student.objects.filter(class_group=cls)
            # Assigned for each student = sum of assignments for the class (respecting filters)
            assign_qs = FeeAssignment.objects.filter(class_group=cls)
            if current_term:
                assign_qs = assign_qs.filter(term=current_term)
            if category_id:
                assign_qs = assign_qs.filter(fee_category_id=category_id)
            assigned_amount = sum(float(getattr(a, 'amount', 0)) for a in assign_qs)
            for s in students:
                paid = FeePayment.objects.filter(student=s, fee_assignment__in=assign_qs).aggregate(total=Sum('amount_paid'))['total'] or 0.0
                balance = max(assigned_amount - float(paid), 0.0)
                if balance > 0:
                    top_debtors.append({
                        'student': getattr(s.user, 'get_full_name', lambda: s.user.username)() if s.user else s.admission_no,
                        'admission_no': s.admission_no,
                        'balance': balance,
                    })
            top_debtors.sort(key=lambda x: x['balance'], reverse=True)
            top_debtors = top_debtors[:10]

    # --- Academic Visualizations ---
    # Helpers for grading bands and pass threshold
    def band_from_score(score: float):
        # Fallback bands if no SubjectGradingScheme is present
        if score is None:
            return 'E'
        if score >= 80:
            return 'A'
        if score >= 70:
            return 'B'
        if score >= 60:
            return 'C'
        if score >= 50:
            return 'D'
        return 'E'

    PASS_THRESHOLD = 50.0

    # Resolve current_term again for grades context
    # subject trend over last 3 terms (top 5 subjects by current-term average)
    subject_trend = {'labels': [], 'datasets': []}
    class_avg = {'labels': [], 'data': []}
    grade_bands = {'labels': ['A', 'B', 'C', 'D', 'E'], 'data': [0, 0, 0, 0, 0]}
    pass_rate = {'labels': ['Pass', 'Fail'], 'data': [0, 0]}

    # Determine last 3 terms chronologically
    last_terms = list(Term.objects.order_by('-start_date')[:3])[::-1]
    term_label_map = {t.id: f"{t.name} {t.academic_year.year}" for t in last_terms}
    subject_trend['labels'] = [term_label_map[t.id] for t in last_terms]

    # Compute per-subject avg for current term to pick top subjects
    if current_term:
        from collections import defaultdict as _dd
        cur_term_scores = _dd(list)
        q = Grade.objects.filter(exam__term=current_term)
        if class_id:
            q = q.filter(student__class_group_id=class_id)
        for g in q.select_related('subject'):
            cur_term_scores[g.subject_id].append(g.score)
        subj_avgs = []
        for sid, scores in cur_term_scores.items():
            if scores:
                subj_avgs.append((sid, sum(scores) / len(scores)))
        subj_avgs.sort(key=lambda x: x[1], reverse=True)
        top_subject_ids = [sid for sid, _ in subj_avgs[:5]]
        # Build datasets across last terms
        for sid in top_subject_ids:
            data_points = []
            for t in last_terms:
                q = Grade.objects.filter(exam__term=t, subject_id=sid)
                if class_id:
                    q = q.filter(student__class_group_id=class_id)
                vals = list(q.values_list('score', flat=True))
                data_points.append(round(sum(vals) / len(vals), 2) if vals else 0)
            subj_name = Subject.objects.filter(id=sid).values_list('name', flat=True).first() or f"Subject {sid}"
            subject_trend['datasets'].append({'label': subj_name, 'data': data_points})

    # Class averages comparison within same level for selected class and term
    if current_term and class_id:
        cls = Class.objects.filter(id=class_id).first()
        if cls:
            level_classes = Class.objects.filter(level=cls.level)
            for c in level_classes:
                q = Grade.objects.filter(exam__term=current_term, student__class_group=c)
                vals = list(q.values_list('score', flat=True))
                avg = round(sum(vals) / len(vals), 2) if vals else 0
                class_avg['labels'].append(c.name)
                class_avg['data'].append(avg)

    # Grade bands and pass rate for selected class & term
    if current_term and class_id:
        q = Grade.objects.filter(exam__term=current_term, student__class_group_id=class_id)
        counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'E': 0}
        passed = 0
        failed = 0
        for g in q:
            b = g.grade_letter or None
            if not b:
                b = band_from_score(g.score)
            counts[b] = counts.get(b, 0) + 1
            if (g.score or 0) >= PASS_THRESHOLD:
                passed += 1
            else:
                failed += 1
        grade_bands['data'] = [counts.get('A', 0), counts.get('B', 0), counts.get('C', 0), counts.get('D', 0), counts.get('E', 0)]
        pass_rate['data'] = [passed, failed]

    return JsonResponse({
        'pie': pie,
        'bar': bar,
        'bar_prev': bar_prev,
        'method': method,
        'class_out': class_out,
        'category': category,
        'daily': daily,
        'cum': {'labels': cum_labels, 'data': cum_paid},
        'cum_prev': cum_prev,
        # Academic datasets
        'subject_trend': subject_trend,
        'class_avg': class_avg,
        'grade_bands': grade_bands,
        'pass_rate': pass_rate,
        'summary': {
            'total_assigned': total_assigned,
            'total_paid': total_paid,
            'payment_percentage': round((total_paid / total_assigned * 100) if total_assigned else 0, 2),
        },
        'top_debtors': top_debtors,
    })

@login_required(login_url='login')
def admin_analytics(request):
    if request.user.role not in ('admin', 'clerk'):
        return redirect('login')
    return AdminAnalyticsView.as_view()(request)

# Finance-specific entry point to analytics (same data, finance-friendly route)
@login_required(login_url='login')
def finance_analytics(request):
    if request.user.role not in ('admin', 'clerk'):
        return redirect('login')
    return AdminAnalyticsView.as_view()(request)

# ... rest of the code remains the same ...
    fee_assignments = FeeAssignment.objects.all()

    # Calculate total assigned as the sum of (amount * number of students in class) for all assignments
    total_assigned = 0
    for assignment in fee_assignments:
        num_students = Student.objects.filter(class_group=assignment.class_group).count()
        total_assigned += float(assignment.amount) * num_students
    total_paid = FeePayment.objects.aggregate(total=Sum('amount_paid'))['total'] or 0
    payment_percentage = (float(total_paid) / float(total_assigned) * 100) if total_assigned else 0

    # Pie chart: Paid vs Unpaid
    paid_value = float(total_paid)
    unpaid_value = float(total_assigned) - float(total_paid)
    pie_labels = json.dumps(["Paid", "Unpaid"])
    pie_data = json.dumps([paid_value, unpaid_value if unpaid_value > 0 else 0])

    # Bar chart: Payments over time (by month)
    from collections import defaultdict
    import calendar
    payments = FeePayment.objects.filter(fee_assignment__in=fee_assignments)
    monthly_totals = defaultdict(float)
    for payment in payments:
        month = payment.payment_date.strftime('%Y-%m')
        monthly_totals[month] += float(payment.amount_paid)
    months_sorted = sorted(monthly_totals.keys())
    bar_labels = json.dumps(months_sorted)
    bar_data = json.dumps([monthly_totals[m] for m in months_sorted])

    context = {
        'avg_performance': avg_performance,
        'subject_labels': subject_labels,
        'avg_scores': avg_scores,
        'total_assigned': total_assigned,
        'total_paid': total_paid,
        'payment_percentage': round(payment_percentage, 2),
        'pie_labels': pie_labels,
        'pie_data': pie_data,
        'bar_labels': bar_labels,
        'bar_data': bar_data,
        'current_term': current_term,
    }
    return render(request, 'dashboards/admin_analytics.html', context)


@login_required(login_url='login')
def admin_subjects(request):
    if request.user.role != 'admin':
        return redirect('login')
    from .models import Subject, Teacher, Department
    subjects = Subject.objects.all()
    # Build a department list from subjects (placeholder, since Subject has no department FK)
    departments = set()
    for subject in subjects:
        # If you later add a department field to Subject, update here
        dept = getattr(subject, 'department', None)
        if dept:
            departments.add(dept)
        # Attach teachers to each subject (assuming a many-to-many or related name)
        if hasattr(subject, 'teachers'):
            subject.teachers = subject.teachers.all()
        elif hasattr(subject, 'teacher_set'):
            subject.teachers = subject.teacher_set.all()
        else:
            subject.teachers = []
    add_subject_form = AddSubjectForm()
    if request.method == 'POST' and 'add_subject' in request.POST:
        add_subject_form = AddSubjectForm(request.POST)
        if add_subject_form.is_valid():
            add_subject_form.save()
            messages.success(request, 'Subject added successfully!')
            add_subject_form = AddSubjectForm()
        else:
            messages.error(request, 'Please correct the errors below.')
    context = {
        'subjects': subjects,
        'departments': list(departments),
        'add_subject_form': add_subject_form,
    }
    return render(request, 'dashboards/admin_subjects.html', context)


@login_required(login_url='login')
def admin_subject_components(request):
    # Admin-only
    if getattr(request.user, 'role', None) != 'admin':
        return redirect('login')

    from .models import Subject, SubjectComponent
    from .forms import SubjectComponentFormSet

    # Accept both 'subject_id' and legacy/shortcut 'parent' query params
    subject_id = request.GET.get('subject_id') or request.GET.get('parent') or request.POST.get('subject_id')
    parent = None

    if subject_id:
        parent = get_object_or_404(Subject, pk=subject_id)

    # Build allowed parent subjects: exclude any subject already used anywhere (as parent or child),
    # except the currently selected parent so it remains visible/editable in the dropdown.
    used_parent_ids = set(SubjectComponent.objects.values_list('parent_id', flat=True))
    used_child_ids = set(SubjectComponent.objects.values_list('child_id', flat=True))
    used_ids = used_parent_ids.union(used_child_ids)
    if parent:
        used_ids.discard(parent.pk)
    subject_qs = Subject.objects.order_by('name').exclude(id__in=used_ids)

    if request.method == 'POST' and parent:
        formset = SubjectComponentFormSet(request.POST, instance=parent, prefix='components')
        if formset.is_valid():
            # Prevent selecting the parent as its own child
            for form in formset:
                if not hasattr(form, 'cleaned_data'):
                    continue
                if form.cleaned_data.get('DELETE'):
                    continue
                child = form.cleaned_data.get('child')
                if child and parent and child.pk == parent.pk:
                    form.add_error('child', 'A subject cannot be a child of itself.')
            if not any(f.errors for f in formset.forms):
                formset.save()
                messages.success(request, 'Subject components updated successfully.')
                # Redirect to persist selection and avoid resubmission
                return redirect(f"{reverse('admin_subject_components')}?subject_id={parent.pk}")
        else:
            messages.error(request, 'Please fix the errors in the form.')
    else:
        formset = SubjectComponentFormSet(instance=parent, prefix='components') if parent else None

    context = {
        'subjects': subject_qs,
        'selected_subject': parent,
        'formset': formset,
    }
    return render(request, 'dashboards/admin_subject_components.html', context)



def custom_logout_view(request):
    # Clear all messages
    list(messages.get_messages(request))
    logout(request)
    messages.success(request, 'Logged out successfully')
    return redirect('login')

@login_required(login_url='login')
def teacher_exams(request, teacher_id):
    from .models import TeacherClassAssignment, Student, Subject, Teacher
    teacher = Teacher.objects.get(id=teacher_id)
    # Access control: only the teacher themselves or an admin can view
    is_admin = getattr(request.user, 'role', None) == 'admin'
    if not is_admin and request.user != teacher.user:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('You are not allowed to view these exams.')
    assignments = TeacherClassAssignment.objects.filter(teacher=teacher).select_related('subject', 'class_group')
    subject_students = {}
    subject_grades = {}
    from .models import Exam, Grade
    for assignment in assignments:
        students = Student.objects.filter(class_group=assignment.class_group)
        subject_students[(assignment.subject, assignment.class_group)] = students

        # Get latest exam for this class level
        latest_exam = Exam.objects.filter(level=assignment.class_group.level).order_by('-date').first()
        grades_map = {}
        if latest_exam:
            grades = Grade.objects.filter(exam=latest_exam, subject=assignment.subject, student__in=students)
            for grade in grades:
                grades_map[grade.student_id] = grade

        # Use nested dictionary instead of tuple key
        if assignment.subject.id not in subject_grades:
            subject_grades[assignment.subject.id] = {}
        subject_grades[assignment.subject.id][assignment.class_group.id] = {'exam': latest_exam, 'grades': grades_map}
    # Get subject_id and class_id from query params to show specific table
    show_subject_id = request.GET.get('subject_id')
    show_class_id = request.GET.get('class_id')

    context = {
        'teacher': teacher,
        'subject_students': subject_students,
        'subject_grades': subject_grades,
        'show_subject_id': show_subject_id,
        'show_class_id': show_class_id,
    }
    return render(request, 'dashboards/teacher_exams.html', context)

@login_required(login_url='login')
def teacher_exam_entry(request, teacher_id, class_id, subject_id, exam_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.user != teacher.user:
        return redirect('login')

    if request.method == 'POST':
        scores = request.POST.getlist('scores[]')
        student_ids = request.POST.getlist('student_ids[]')
        
        exam = get_object_or_404(Exam, id=exam_id)
        subject = get_object_or_404(Subject, id=subject_id)
        class_group = get_object_or_404(Class, id=class_id)
        # Security: Only allow if this teacher is assigned to this class+subject (or is admin)
        is_admin = getattr(request.user, 'role', None) == 'admin'
        if not is_admin:
            from .models import TeacherClassAssignment
            has_assignment = TeacherClassAssignment.objects.filter(
                teacher=teacher, class_group=class_group, subject=subject
            ).exists()
            if not has_assignment:
                from django.http import HttpResponseForbidden
                return HttpResponseForbidden('You are not assigned to this class/subject.')
        
        # Validate scores
        for score in scores:
            try:
                score = float(score)
                if score < 0 or score > 100:
                    messages.error(request, 'Scores must be between 0 and 100')
                    return redirect('teacher_exam_entry', teacher_id=teacher_id, 
                                  class_id=class_id, subject_id=subject_id, exam_id=exam_id)
            except ValueError:
                messages.error(request, 'Invalid score format')
                return redirect('teacher_exam_entry', teacher_id=teacher_id, 
                              class_id=class_id, subject_id=subject_id, exam_id=exam_id)

        # Save scores
        try:
            for student_id, score in zip(student_ids, scores):
                student = Student.objects.get(id=student_id)
                Grade.objects.update_or_create(
                    student=student,
                    exam=exam,
                    subject=subject,
                    defaults={'score': score}
                )
            messages.success(request, 'Scores saved successfully')
        except Exception as e:
            messages.error(request, f'Error saving scores: {str(e)}')
            return redirect('teacher_exam_entry', teacher_id=teacher_id, 
                          class_id=class_id, subject_id=subject_id, exam_id=exam_id)

        return redirect('teacher_exam_entry', teacher_id=teacher_id, 
                       class_id=class_id, subject_id=subject_id, exam_id=exam_id)

    # Get teacher's subjects and classes
    subjects = teacher.subjects.all()
    classes = Class.objects.filter(teachers=teacher)
    
    # Get exam details
    exam = get_object_or_404(Exam, id=exam_id)
    subject = get_object_or_404(Subject, id=subject_id)
    class_group = get_object_or_404(Class, id=class_id)

    # Get students for this class and subject
    students = Student.objects.filter(
        class_group_id=class_id
    ).order_by('user__last_name', 'user__first_name')

    # Get existing grades for this exam
    grade_map = {}
    existing_grades = Grade.objects.filter(
        exam=exam,
        subject=subject,
        student__in=students
    )
    for grade in existing_grades:
        grade_map[grade.student_id] = grade.score

    context = {
        'teacher': teacher,
        'subjects': subjects,
        'classes': classes,
        'exam': exam,
        'subject': subject,
        'class_group': class_group,
        'students': students,
        'grade_map': grade_map,
    }
    return render(request, 'dashboards/teacher_exam_entry.html', context)

@login_required(login_url='login')
def upload_marksheet(request):
    if request.method == 'POST':
        file = request.FILES['file']
        import pandas as pd
        import datetime

        df = pd.read_excel(file) if file.name.endswith('.xlsx') else pd.read_csv(file)
        df.columns = [col.strip() for col in df.columns]

        # Flexible required columns check
        required_cols_student = ['admission_no', 'student_name']
        required_cols_subject = ['subject_code', 'subject_name']
        other_required_cols = ['exam_name', 'term', 'academic_year', 'score']

        has_student_id = any(col in df.columns for col in required_cols_student)
        has_subject_id = any(col in df.columns for col in required_cols_subject)
        has_others = all(col in df.columns for col in other_required_cols)

        if not (has_student_id and has_subject_id and has_others):
            msg = "File is missing required columns. It must contain ('admission_no' or 'student_name'), ('subject_code' or 'subject_name'), 'exam_name', 'term', 'academic_year', and 'score'."
            messages.error(request, msg)
            return redirect('upload_marksheet')

        errors = []
        try:
            for index, row in df.iterrows():
                try:
                    # Find Student
                    student = None
                    if 'admission_no' in df.columns and pd.notna(row.get('admission_no')):
                        student = Student.objects.get(admission_no=row['admission_no'])
                    elif 'student_name' in df.columns and pd.notna(row.get('student_name')):
                        full_name = str(row['student_name']).strip()
                        parts = full_name.split(' ', 1)
                        first_name = parts[0]
                        last_name = parts[1] if len(parts) > 1 else ''
                        
                        user_qs = User.objects.filter(
                            Q(username__iexact=full_name) |
                            (Q(first_name__iexact=first_name) & Q(last_name__iexact=last_name))
                        )
                        
                        if user_qs.count() == 1:
                            student = Student.objects.get(user=user_qs.first())
                        elif user_qs.count() > 1:
                            errors.append(f"Row {index+2}: Multiple students found for name '{full_name}'. Use admission number.")
                            continue
                        else:
                            errors.append(f"Row {index+2}: Student with name '{full_name}' not found.")
                            continue
                    else:
                        errors.append(f"Row {index+2}: Missing student identifier.")
                        continue

                    # Find Subject
                    subject = None
                    if 'subject_code' in df.columns and pd.notna(row.get('subject_code')):
                        subject = Subject.objects.get(code__iexact=str(row['subject_code']).strip())
                    elif 'subject_name' in df.columns and pd.notna(row.get('subject_name')):
                        subject_identifier = str(row['subject_name']).strip()
                        subject_qs = Subject.objects.filter(
                            Q(name__iexact=subject_identifier) | Q(code__iexact=subject_identifier)
                        )
                        if subject_qs.count() == 1:
                            subject = subject_qs.first()
                        else:
                            errors.append(f"Row {index+2}: Subject '{subject_identifier}' not found or is ambiguous.")
                            continue
                    else:
                        errors.append(f"Row {index+2}: Missing subject identifier.")
                        continue
                    
                    # Get or create other objects
                    academic_year, _ = AcademicYear.objects.get_or_create(year=row['academic_year'])
                    term, _ = Term.objects.get_or_create(name=row['term'], academic_year=academic_year)
                    exam, _ = Exam.objects.get_or_create(name=row['exam_name'], term=term, defaults={'date': datetime.date.today()})

                    # Update or create grade
                    Grade.objects.update_or_create(
                        student=student,
                        subject=subject,
                        exam=exam,
                        defaults={'score': row['score']}
                    )
                except Exception as e:
                    errors.append(f'Row {index+2}: An unexpected error occurred: {e}')

            if errors:
                for error in errors:
                    messages.error(request, error)
            else:
                messages.success(request, 'Marksheet uploaded successfully.')

        except Exception as e:
            messages.error(request, f'Error processing file: {e}')

        return redirect('upload_marksheet')

    teacher = request.user.teacher
    return render(request, 'dashboards/upload_marksheet.html', {'teacher': teacher})

from .forms import AddTeacherForm

@login_required(login_url='login')
def admin_teachers(request):
    from .forms import AddTeacherForm
    from .models import Teacher, Department, Subject
    from django.contrib import messages
    from django.db.models import Q
    # Admin-only access
    if not (getattr(request.user, 'role', None) == 'admin' or request.user.is_staff or request.user.is_superuser):
        return redirect('login')

    # Filtering logic
    teachers = Teacher.objects.all()
    search = request.GET.get('search', '').strip()
    department_id = request.GET.get('department', '')
    subject_id = request.GET.get('subject', '')
    gender = request.GET.get('gender', '')

    if search:
        teachers = teachers.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__username__icontains=search) |
            Q(user__email__icontains=search) |
            Q(phone__icontains=search) |
            Q(tsc_number__icontains=search) |
            Q(staff_id__icontains=search)
        )
    if department_id:
        teachers = teachers.filter(department_id=department_id)
    if subject_id:
        teachers = teachers.filter(subjects__id=subject_id)
    if gender:
        teachers = teachers.filter(gender=gender)
    teachers = teachers.distinct()

    # For filter dropdowns
    departments = Department.objects.all()
    subjects = Subject.objects.all()

    if request.method == 'POST':
        form = AddTeacherForm(request.POST)
        if form.is_valid():
            user = form.save_user()
            teacher = form.save(commit=False)
            teacher.user = user
            teacher.save()
            form.save_m2m()
            messages.success(request, 'Teacher added successfully!')
            return redirect('admin_teachers')
        else:
            add_teacher_form = form
    else:
        add_teacher_form = AddTeacherForm()
    context = {
        'teachers': teachers,
        'add_teacher_form': add_teacher_form,
        'departments': departments,
        'subjects': subjects,
        'search': search,
        'selected_department': department_id,
        'selected_subject': subject_id,
        'selected_gender': gender,
    }
    return render(request, 'dashboards/admin_teachers.html', context)




@login_required(login_url='login')
def admin_assign_responsibility(request):
    from .forms import TeacherResponsibilityForm
    from .models import TeacherResponsibility
    responsibilities = TeacherResponsibility.objects.select_related('teacher', 'teacher__user').order_by('-assigned_at')
    # Admin-only access
    if getattr(request.user, 'role', None) != 'admin':
        return redirect('login')
    if request.method == 'POST':
        form = TeacherResponsibilityForm(request.POST)
        if form.is_valid():
            resp = form.save(commit=False)
            resp.assigned_by = request.user
            resp.save()
            messages.success(request, 'Responsibility assigned successfully!')
            return redirect('admin_assign_responsibility')
    else:
        form = TeacherResponsibilityForm()
    return render(request, 'dashboards/admin_assign_responsibility.html', {
        'form': form,
        'responsibilities': responsibilities,
    })

@login_required(login_url='login')
def delete_teacher(request, teacher_id):
    from django.shortcuts import get_object_or_404, redirect, render
    from django.contrib import messages
    from .models import Teacher
    # Admin-only access
    if getattr(request.user, 'role', None) != 'admin':
        return redirect('login')
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        # Delete the user as well (cascades to Teacher)
        user = teacher.user
        teacher.delete()
        user.delete()
        messages.success(request, 'Teacher deleted successfully!')
        return redirect('admin_teachers')
    return render(request, 'dashboards/delete_teacher.html', {'teacher': teacher})

@login_required(login_url='login')
def edit_teacher(request, teacher_id):
    from django.shortcuts import get_object_or_404, redirect, render
    from django.contrib import messages
    from .models import Teacher
    from .forms import AddTeacherForm
    # Admin-only access
    if getattr(request.user, 'role', None) != 'admin':
        return redirect('login')
    teacher = get_object_or_404(Teacher, id=teacher_id)
    user = teacher.user
    if request.method == 'POST':
        form = AddTeacherForm(request.POST, instance=teacher)
        # Patch user fields for validation
        form.fields['first_name'].initial = user.first_name
        form.fields['last_name'].initial = user.last_name
        form.fields['username'].initial = user.username
        form.fields['email'].initial = user.email
        if form.is_valid():
            # Update user fields
            user.first_name = form.cleaned_data.get('first_name')
            user.last_name = form.cleaned_data.get('last_name')
            user.username = form.cleaned_data.get('username')
            user.email = form.cleaned_data.get('email')
            password = form.cleaned_data.get('password', None)
            if password:
                user.set_password(password)
            user.save()
            # Update teacher fields
            teacher = form.save(commit=False)
            teacher.user = user
            teacher.save()
            form.save_m2m()
            messages.success(request, 'Teacher updated successfully!')
            return redirect('admin_teachers')
    else:
        # Pre-populate form with user and teacher data
        initial = {
            'first_name': user.first_name,
            'last_name': user.last_name,
            'username': user.username,
            'email': user.email,
        }
        form = AddTeacherForm(instance=teacher, initial=initial)
    return render(request, 'dashboards/edit_teacher.html', {'form': form, 'teacher': teacher})

@login_required(login_url='login')
@user_passes_test(is_admin_or_clerk)
def admin_payment(request):
    from .models import Student, Term, FeeAssignment, FeePayment, MpesaTransaction
    from django.utils import timezone
    from django.contrib import messages
    from .mpesa_utils import initiate_stk_push
    today = timezone.now().date()
    current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()
    students = Student.objects.select_related('user', 'class_group').all().order_by('user__last_name', 'user__first_name')
    selected_student_id = request.GET.get('student_id')

    if request.method == 'POST':
        try:
            student_id = request.POST.get('student_id')
            amount_paid = request.POST.get('amount_paid')
            payment_method = request.POST.get('payment_method')
            reference = request.POST.get('reference')
            phone_number = request.POST.get('phone_number')
            student = Student.objects.get(id=student_id)
            fee_assignment = FeeAssignment.objects.filter(class_group=student.class_group, term=current_term).first()
            if not fee_assignment:
                fee_assignment = FeeAssignment.objects.first()
            # If payment method is Mpesa Paybill, initiate STK Push
            if payment_method == 'Mpesa Paybill':
                # Normalize phone number to 2547XXXXXXXX
                phone_number = phone_number.strip().replace(" ", "")
                if phone_number.startswith('+254'):
                    phone_number = phone_number[1:]
                elif phone_number.startswith('07'):
                    phone_number = '254' + phone_number[1:]
                elif phone_number.startswith('7') and len(phone_number) == 9:
                    phone_number = '254' + phone_number
                # else: assume already correct
                print("[DEBUG] Normalized phone number:", phone_number)
                print("[DEBUG] Payment method:", payment_method)
                account_ref = f"Account#{student.admission_no}"
                print("[DEBUG] About to call initiate_stk_push with:", phone_number, amount_paid, account_ref)
                stk_response = initiate_stk_push(
                    phone_number=phone_number,
                    amount=amount_paid,
                    account_ref=account_ref,
                    transaction_desc=f'Fee payment for {student.full_name}'
                )
                print("[DEBUG] STK Response:", stk_response)
                if stk_response.get('ResponseCode') == '0':
                    # Persist a pending MpesaTransaction for later verification
                    try:
                        MpesaTransaction.objects.create(
                            student=student,
                            fee_assignment=fee_assignment,
                            phone_number=phone_number,
                            amount=amount_paid,
                            account_reference=account_ref,
                            merchant_request_id=stk_response.get('MerchantRequestID'),
                            checkout_request_id=stk_response.get('CheckoutRequestID'),
                            status='pending'
                        )
                    except Exception as e:
                        print('[M-PESA][WARN] Could not persist MpesaTransaction:', e)
                    messages.success(request, 'STK Push sent! Complete payment on your phone.')
                else:
                    messages.error(request, f"STK Push failed: {stk_response.get('errorMessage', stk_response)}")
                    # Optionally, do not save payment if STK Push fails
                    return redirect(request.path + f'?student_id={student.id}')
            # Only create FeePayment immediately for non-Mpesa methods
            if payment_method != 'Mpesa Paybill':
                payment = FeePayment.objects.create(
                    student=student,
                    fee_assignment=fee_assignment,
                    amount_paid=amount_paid,
                    payment_method=payment_method,
                    reference=reference,
                    phone_number=phone_number
                )
            else:
                payment = None
            from django.db.models import Sum
            # Compute student's TOTAL balance (all categories/terms), not just this assignment
            total_billed = FeeAssignment.objects.filter(class_group=student.class_group).aggregate(total=Sum('amount'))['total'] or 0
            total_paid = FeePayment.objects.filter(student=student).aggregate(total=Sum('amount_paid'))['total'] or 0
            balance = (total_billed or 0) - (total_paid or 0)
            payment_date = payment.payment_date.strftime('%Y-%m-%d %H:%M') if payment else str(timezone.now())
            confirm_msg = f"Dear {student.full_name}, your payment of Ksh. {amount_paid} has been received on {payment_date}. Outstanding balance: Ksh. {balance:.2f}. Thank you."
            if student.user and student.user.email:
                from django.core.mail import send_mail
                send_mail(
                    subject='Fee Payment Confirmation',
                    message=confirm_msg,
                    from_email=None,
                    recipient_list=[student.user.email],
                    fail_silently=True
                )
            from .messaging_utils import send_sms
            # Send SMS to student's registered phone
            if student.phone:
                try:
                    send_sms(student.phone, confirm_msg)
                except Exception:
                    pass
            # Also send SMS to the payer's phone number if provided and different
            if phone_number and (not student.phone or str(phone_number).strip() != str(student.phone).strip()):
                try:
                    send_sms(phone_number, confirm_msg)
                except Exception:
                    pass
            # Notify student of successful payment
            from core.utils import create_notification
            if student.user:
                notif_msg = f"Your payment of Ksh. {amount_paid} has been received on {payment_date}. Outstanding balance: Ksh. {balance:.2f}."
                create_notification(student.user, notif_msg)
            if payment_method != 'Mpesa Paybill':
                messages.success(request, f'Payment of Ksh. {amount_paid} recorded for {student.full_name}. Confirmation sent.')
            else:
                messages.info(request, f'STK initiated for Ksh. {amount_paid}. Payment will reflect after confirmation.')
            return redirect(request.path + f'?student_id={student.id}')
        except Student.DoesNotExist:
            # Notify admin of error
            from core.utils import create_notification
            if request.user and hasattr(request.user, 'role') and request.user.role == 'admin':
                create_notification(request.user, 'Payment error: Selected student does not exist.')
            messages.error(request, 'Selected student does not exist.')
        except Exception as e:
            # Notify admin of error
            from core.utils import create_notification
            if request.user and hasattr(request.user, 'role') and request.user.role == 'admin':
                create_notification(request.user, f'Payment error: {e}')
            messages.error(request, f'Error recording payment: {e}')

    # Query all payments for admin table with efficient annotation and pagination
    from django.db.models import Case, When, Value, CharField, Q
    from django.core.paginator import Paginator
    # Optional filtering to keep result sets small and fast
    q = request.GET.get('q', '').strip()
    method = request.GET.get('method', '').strip()
    status_filter = request.GET.get('status', '').strip()
    start_date = request.GET.get('start', '').strip()
    end_date = request.GET.get('end', '').strip()

    all_payments_qs = (
        FeePayment.objects
        .select_related(
            'student__user',
            'fee_assignment__fee_category',
            'fee_assignment__class_group',
            'fee_assignment__term'
        )
        .annotate(
            computed_status=Case(
                When(Q(payment_method='Mpesa Paybill') & Q(reference__isnull=True), then=Value('Pending')),
                default=Value('Completed'),
                output_field=CharField(),
            )
        )
        .order_by('-payment_date')
    )
    # Apply filters
    if q:
        all_payments_qs = all_payments_qs.filter(
            Q(student__admission_no__icontains=q)
            | Q(student__user__first_name__icontains=q)
            | Q(student__user__last_name__icontains=q)
        )
    if method:
        all_payments_qs = all_payments_qs.filter(payment_method=method)
    if status_filter:
        # Support filtering by either the computed display status or the model field status
        all_payments_qs = all_payments_qs.filter(
            Q(computed_status__iexact=status_filter) | Q(status__iexact=status_filter)
        )
    if start_date:
        all_payments_qs = all_payments_qs.filter(payment_date__date__gte=start_date)
    if end_date:
        all_payments_qs = all_payments_qs.filter(payment_date__date__lte=end_date)
    page_number = request.GET.get('page', 1)
    paginator = Paginator(all_payments_qs, 25)  # 25 rows per page
    all_payments = paginator.get_page(page_number)
    context = {
        'students': students,
        'current_term': current_term,
        'selected_student_id': int(selected_student_id) if selected_student_id and selected_student_id.isdigit() else None,
        'all_payments': all_payments,
        'q': q,
        'method': method,
        'status_filter': status_filter,
        'start': start_date,
        'end': end_date,
    }
    return render(request, 'dashboards/admin_payment.html', context)

@login_required(login_url='login')
def admin_events(request):
    from .forms import EventForm
    from .models import Event
    from django.utils import timezone
    # Admin-only access
    if getattr(request.user, 'role', None) != 'admin':
        return redirect('login')

    # Handle marking event as done
    if request.method == 'POST' and 'mark_done_event_id' in request.POST:
        event_id = request.POST.get('mark_done_event_id')
        is_done = request.POST.get('is_done') == 'on'
        comment = request.POST.get('comment', '').strip()
        try:
            event = Event.objects.get(id=event_id)
            event.is_done = is_done
            event.comment = comment
            event.save()
            messages.success(request, 'Event status updated.')
        except Event.DoesNotExist:
            messages.error(request, 'Event not found.')
        return redirect(request.path + '?' + request.META.get('QUERY_STRING', ''))

    # Handle event deletion
    if 'delete' in request.GET:
        delete_id = request.GET['delete']
        try:
            event = Event.objects.get(id=delete_id)
            event.delete()
            messages.success(request, 'Event deleted successfully.')
            # Redirect to remove ?delete=... from URL
            return redirect('admin_events')
        except Event.DoesNotExist:
            messages.error(request, 'Event not found.')
            return redirect('admin_events')

    filter_type = request.GET.get('filter', 'all')
    now = timezone.now()
    events = Event.objects.all().order_by('-start')

    # Filtering logic
    if filter_type == 'upcoming':
        filtered_events = events.filter(start__gte=now, is_done=False)
    elif filter_type == 'done':
        filtered_events = events.filter(is_done=True)
    elif filter_type == 'undone':
        filtered_events = events.filter(is_done=False)
    else:
        filtered_events = events

    # Past events that are not marked as done
    past_events = events.filter(start__lt=now, is_done=False)

    # Handle event editing (POST)
    if request.method == 'POST':
        if 'edit_event_id' in request.POST:
            event_id = request.POST.get('edit_event_id')
            try:
                event = Event.objects.get(id=event_id)
            except Event.DoesNotExist:
                messages.error(request, 'Event not found.')
                return redirect('admin_events')
            form = EventForm(request.POST, instance=event)
            if form.is_valid():
                form.save()
                messages.success(request, 'Event updated successfully.')
                return redirect('admin_events')
            else:
                edit_event = event
        else:
            # Only create if not editing
            form = EventForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Event created successfully.')
                return redirect('admin_events')
            else:
                edit_event = None
    else:
        form = EventForm()
        edit_event = None
        if 'edit' in request.GET:
            try:
                edit_event = Event.objects.get(id=request.GET['edit'])
                form = EventForm(instance=edit_event)
            except Event.DoesNotExist:
                edit_event = None

    context = {
        'filtered_events': filtered_events,
        'past_events': past_events,
        'filter_type': filter_type,
        'form': form,
        'edit_event': edit_event,
    }
    return render(request, 'dashboards/admin_events.html', context)


@login_required(login_url='login')
def teacher_timetable(request, teacher_id):
    # Personalized timetable for a teacher across the week
    from .models import Teacher, DefaultTimetable, PeriodSlot

    # Access control: allow admins or the owner teacher to view
    is_admin = getattr(request.user, 'role', None) == 'admin'
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if not is_admin and getattr(request.user, 'id', None) != getattr(teacher.user, 'id', None):
        # Forbidden
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('You are not allowed to view this timetable.')

    # Match class timetable: only class slots, ordered by start
    periods = PeriodSlot.objects.filter(is_class_slot=True).order_by('start_time')
    days = [d[0] for d in DefaultTimetable.DAY_CHOICES]

    # Initialize grid: period.id -> day -> None/entry
    timetable_grid = {p.id: {day: None for day in days} for p in periods}

    # Fetch this teacher's assignments
    entries = (
        DefaultTimetable.objects
        .filter(teacher=teacher)
        .select_related('class_group', 'subject', 'period')
    )

    for entry in entries:
        if entry.period_id in timetable_grid and entry.day in timetable_grid[entry.period_id]:
            timetable_grid[entry.period_id][entry.day] = entry

    context = {
        'teacher': teacher,
        'days': days,
        'periods': periods,
        'timetable_grid': timetable_grid,
    }
    return render(request, 'dashboards/teacher_timetable.html', context)

@login_required(login_url='login')
def student_fees(request):
    from .models import FeeAssignment, FeePayment, Term, Student, MpesaTransaction
    from django.utils import timezone
    from django.db.models import Sum
    from django.contrib import messages
    from .mpesa_utils import initiate_stk_push

    user = request.user
    try:
        student = Student.objects.get(user=user)
    except Student.DoesNotExist:
        messages.error(request, 'Student profile not found.')
        return redirect('login')

    today = timezone.now().date()
    current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()
    fee_assignments = FeeAssignment.objects.filter(class_group=student.class_group, term=current_term)
    fee_payments = FeePayment.objects.filter(student=student, fee_assignment__in=fee_assignments)
    success = False

    if request.method == 'POST':
        amount_paid = request.POST.get('amount_paid')
        payment_method = request.POST.get('payment_method')
        reference = request.POST.get('reference')
        phone_number = request.POST.get('phone_number')
        fee_assignment = FeeAssignment.objects.filter(class_group=student.class_group, term=current_term).first()
        # Allow payment even if no FeeAssignment exists
        # If payment method is Mpesa Paybill, initiate STK Push
        try:
            if payment_method == 'Mpesa Paybill':
                phone_number = phone_number.strip().replace(" ", "")
                if phone_number.startswith('+254'):
                    phone_number = phone_number[1:]
                elif phone_number.startswith('07'):
                    phone_number = '254' + phone_number[1:]
                elif phone_number.startswith('7') and len(phone_number) == 9:
                    phone_number = '254' + phone_number
                # else: assume already correct
                account_ref = f"Account#{student.admission_no}"
                stk_response = initiate_stk_push(
                    phone_number=phone_number,
                    amount=amount_paid,
                    account_ref=account_ref,
                    transaction_desc=f'Fee payment for {student.full_name}'
                )
                if stk_response.get('ResponseCode') == '0':
                    try:
                        MpesaTransaction.objects.create(
                            student=student,
                            fee_assignment=fee_assignment,
                            phone_number=phone_number,
                            amount=amount_paid,
                            account_reference=account_ref,
                            merchant_request_id=stk_response.get('MerchantRequestID'),
                            checkout_request_id=stk_response.get('CheckoutRequestID'),
                            status='pending'
                        )
                    except Exception as e:
                        print('[M-PESA][WARN] Could not persist MpesaTransaction (student):', e)
                    messages.success(request, 'STK Push sent! Complete payment on your phone.')
                else:
                    messages.error(request, f"STK Push failed: {stk_response.get('errorMessage', stk_response)}")
                    return redirect(request.path)
            # Only create FeePayment immediately for non-Mpesa methods
            if payment_method != 'Mpesa Paybill':
                FeePayment.objects.create(
                    student=student,
                    fee_assignment=fee_assignment,
                    amount_paid=amount_paid,
                    payment_method=payment_method,
                    reference=reference,
                    phone_number=phone_number
                )
            messages.success(request, f'Payment of Ksh. {amount_paid} recorded.')
            return redirect(f'{request.path}?success=1')
        except Exception as e:
            messages.error(request, f'Error submitting payment: {e}')

    if request.GET.get('success') == '1':
        success = True

    from .forms import StudentContactUpdateForm
    contact_form = StudentContactUpdateForm(instance=student, user_instance=student.user)
    # Compute total outstanding balance (arrears from previous terms + current term balance)
    total_outstanding = 0
    try:
        # Current term billed and paid
        billed_current = fee_assignments.aggregate(total=Sum('amount'))['total'] or 0
        paid_current = fee_payments.aggregate(total=Sum('amount_paid'))['total'] or 0

        # Previous terms arrears
        previous_terms = Term.objects.none()
        if current_term:
            previous_terms = Term.objects.filter(start_date__lt=current_term.start_date).order_by('start_date')
        else:
            previous_terms = Term.objects.all().order_by('start_date')

        prev_assignments = FeeAssignment.objects.filter(class_group=student.class_group, term__in=previous_terms)
        billed_prev = prev_assignments.aggregate(total=Sum('amount'))['total'] or 0
        paid_prev = FeePayment.objects.filter(student=student, fee_assignment__term__in=previous_terms).aggregate(total=Sum('amount_paid'))['total'] or 0

        total_outstanding = (billed_prev + billed_current) - (paid_prev + paid_current)
        if total_outstanding < 0:
            total_outstanding = 0
    except Exception:
        # Fail-safe: keep zero if any issue occurs
        total_outstanding = 0
    context = {
        'student': student,
        'current_term': current_term,
        'fee_assignments': fee_assignments,
        'fee_payments': fee_payments,
        'success': success,
        'total_outstanding': total_outstanding,
    }
    return render(request, 'dashboards/student_fees.html', context)

@login_required(login_url='login')
def admin_fees(request):
    from .forms import FeeCategoryForm, FeeAssignmentForm
    from .models import FeeCategory, FeeAssignment, FeePayment, Student, Class, Term
    from django.db.models import Sum, Q, Value, DecimalField
    from decimal import Decimal
    from django.db.models.functions import Coalesce, Cast
    from django.contrib import messages
    import json

    fee_form = FeeCategoryForm()
    assign_form = FeeAssignmentForm()
    
    # --- Edit/Delete Fee Category Logic ---
    edit_category = None
    if request.user.role not in ('admin', 'clerk'):
        return HttpResponseForbidden('You are not authorized to access this page.')

    # Handle Delete (now POST)
    if request.method == 'POST' and 'delete' in request.POST:
        try:
            delete_id = int(request.GET['delete'])
            category = FeeCategory.objects.get(id=delete_id)
            category.delete()
            messages.success(request, 'Fee category deleted successfully!')
            return redirect('admin_fees')
        except FeeCategory.DoesNotExist:
            messages.error(request, 'Fee category not found.')
            return redirect('admin_fees')
        except Exception as e:
            messages.error(request, f'Error deleting fee category: {e}')
            return redirect('admin_fees')

    # Handle Edit (populate form)
    if 'edit' in request.GET:
        try:
            edit_id = int(request.GET['edit'])
            edit_category = FeeCategory.objects.get(id=edit_id)
            fee_form = FeeCategoryForm(instance=edit_category)
        except FeeCategory.DoesNotExist:
            messages.error(request, 'Fee category not found.')
            return redirect('admin_fees')

    # Handle FeeCategoryForm POST (add or update)
    if request.method == 'POST' and 'save_fee_category' in request.POST:
        if request.POST.get('category_id'):
            # Update existing
            try:
                edit_category = FeeCategory.objects.get(id=request.POST['category_id'])
            except FeeCategory.DoesNotExist:
                messages.error(request, 'Fee category not found.')
                return redirect('admin_fees')
            fee_form = FeeCategoryForm(request.POST, instance=edit_category)
        else:
            # New category
            fee_form = FeeCategoryForm(request.POST)
        if fee_form.is_valid():
            try:
                fee_form.save()
                if request.POST.get('category_id'):
                    messages.success(request, 'Fee category updated successfully!')
                else:
                    messages.success(request, 'Fee category saved successfully!')
                return redirect('admin_fees')
            except Exception as e:
                messages.error(request, f'Error saving fee category: {e}')
        else:
            messages.error(request, 'Please correct the errors in the fee category form.')

    # Handle FeeAssignment POST
    if request.method == 'POST' and 'assign_fee' in request.POST:
        assign_form = FeeAssignmentForm(request.POST)
        if assign_form.is_valid():
            fee_category = assign_form.cleaned_data['fee_category']
            class_groups = assign_form.cleaned_data['class_group']
            term = assign_form.cleaned_data['term']
            amount = assign_form.cleaned_data['amount']
            from .models import FeeAssignment
            from django.db import IntegrityError
            assigned_count = 0
            skipped_count = 0
            for class_group in class_groups:
                try:
                    FeeAssignment.objects.create(
                        fee_category=fee_category,
                        class_group=class_group,
                        term=term,
                        amount=amount
                    )
                    assigned_count += 1
                except IntegrityError:
                    skipped_count += 1
            if assigned_count:
                messages.success(request, f'Fee assignment saved for {assigned_count} class(es)!')
            if skipped_count:
                messages.warning(request, f'Skipped {skipped_count} class(es) (already assigned).')
        else:
            messages.error(request, 'Please correct the errors in the fee assignment form.')

    today = timezone.now().date()
    current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()
    all_classes = Class.objects.all()
    all_fee_categories = FeeCategory.objects.all()

    # Compute previous terms ONCE
    if current_term:
        previous_terms = Term.objects.filter(start_date__lt=current_term.start_date)
    else:
        previous_terms = Term.objects.none()

    # Base students queryset (select related to avoid extra queries)
    students_qs = Student.objects.filter(graduated=False).select_related('user', 'class_group')

    selected_class_id = request.GET.get('class_group', '')
    sort_order = request.GET.get('sort', 'largest')
    if selected_class_id:
        students_qs = students_qs.filter(class_group_id=selected_class_id)

    # Precompute fees overview for current term
    fees = FeeAssignment.objects.select_related('fee_category', 'class_group', 'term')
    if current_term:
        fees = fees.filter(term=current_term)

    # Precompute billed amounts for current term, grouped by class and category
    current_billed_qs = FeeAssignment.objects.all()
    if current_term:
        current_billed_qs = current_billed_qs.filter(term=current_term)
    current_billed = {}
    for row in current_billed_qs.values('class_group_id', 'fee_category_id').annotate(
        total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
    ):
        current_billed.setdefault(row['class_group_id'], {})[row['fee_category_id']] = float(row['total'] or 0)

    # Precompute total billed for all previous terms per class
    prev_billed = {}
    if previous_terms.exists():
        for row in (
            FeeAssignment.objects.filter(term__in=previous_terms)
            .values('class_group_id')
            .annotate(total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2)))
        ):
            prev_billed[row['class_group_id']] = float(row['total'] or 0)
    # Precompute total paid (all time) per student
    paid_all_by_student = {}
    for row in FeePayment.objects.values('student_id').annotate(
        total=Sum('amount_paid', output_field=DecimalField(max_digits=12, decimal_places=2))
    ):
        paid_all_by_student[row['student_id']] = float(row['total'] or 0)
    # Precompute total paid in previous terms per student
    paid_prev_by_student = {}
    if previous_terms.exists():
        for row in (
            FeePayment.objects.filter(fee_assignment__term__in=previous_terms)
            .values('student_id')
            .annotate(total=Sum('amount_paid', output_field=DecimalField(max_digits=12, decimal_places=2)))
        ):
            paid_prev_by_student[row['student_id']] = float(row['total'] or 0)

    # Build per-student aggregates for the current page of students (with pagination)
    from django.core.paginator import Paginator
    students_paginator = Paginator(students_qs, 50)
    students_page = students_paginator.get_page(request.GET.get('students_page') or request.GET.get('page'))

    student_category_totals = {}
    student_totals = {}
    student_paid = {}
    student_balances = {}
    student_outstanding = {}

    for student in students_page.object_list:
        class_id = student.class_group_id
        per_cat = {}
        total_billed_current = 0.0
        cat_map = current_billed.get(class_id, {})
        for cat in all_fee_categories:
            amt = float(cat_map.get(cat.id, 0.0))
            per_cat[cat.id] = amt
            total_billed_current += amt

        # Outstanding = billed previous terms for the class - paid by student in previous terms
        outstanding = float(prev_billed.get(class_id, 0.0)) - float(paid_prev_by_student.get(student.id, 0.0))
        paid_all = float(paid_all_by_student.get(student.id, 0.0))
        balance = total_billed_current + outstanding - paid_all

        student_category_totals[student.id] = per_cat
        student_totals[student.id] = total_billed_current
        student_paid[student.id] = paid_all
        student_outstanding[student.id] = outstanding
        student_balances[student.id] = balance

    # Sort students by balance within the page
    if sort_order == 'largest':
        students_sorted = sorted(students_page.object_list, key=lambda s: student_balances.get(s.id, 0), reverse=True)
    else:
        students_sorted = sorted(students_page.object_list, key=lambda s: student_balances.get(s.id, 0))

    # Replace the page's object_list with sorted result while keeping paginator metadata
    students_page.object_list = students_sorted


    category_labels = list(all_fee_categories.values_list('name', flat=True))
    category_data = [
        float(
            (fees.filter(fee_category=cat).aggregate(total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2)))['total']) or 0
        )
        for cat in all_fee_categories
    ]
    payment_labels = []
    payment_data = []
    payments_by_month = (
        FeePayment.objects
        .extra(select={'month': "strftime('%%Y-%%m', payment_date)"})
        .values('month')
        .annotate(total=Sum('amount_paid', output_field=DecimalField(max_digits=12, decimal_places=2)))
        .order_by('month')
    )
    for entry in payments_by_month:
        payment_labels.append(entry['month'])
        payment_data.append(float(entry['total']) if entry['total'] is not None else 0.0)

    # --- Payment History Filtering & Sorting ---
    payment_sort = request.GET.get('payment_sort', 'newest')
    payment_search = request.GET.get('payment_search', '').strip()
    payments_qs = FeePayment.objects.select_related('student__user', 'fee_assignment__fee_category')
    if payment_search:
        payments_qs = payments_qs.filter(
            Q(student__admission_number__icontains=payment_search) |
            Q(student__user__username__icontains=payment_search) |
            Q(student__user__first_name__icontains=payment_search) |
            Q(student__user__last_name__icontains=payment_search)
        )
    if payment_sort == 'oldest':
        payments_qs = payments_qs.order_by('payment_date')
    else:
        payments_qs = payments_qs.order_by('-payment_date')

    # Paginate payments (supports AJAX partial reload)
    payments_paginator = Paginator(payments_qs, 50)
    payments_page = payments_paginator.get_page(request.GET.get('payments_page') or request.GET.get('page'))

    context = {
        'fee_form': fee_form,
        'assign_form': assign_form,
        'fees': fees,
        'all_classes': all_classes,
        'all_fee_categories': all_fee_categories,
        'students_page': students_page,
        'student_category_totals': student_category_totals,
        'student_totals': student_totals,
        'student_paid': student_paid,
        'student_balances': student_balances,
        'selected_class_id': selected_class_id,
        'sort_order': sort_order,
        'current_term': current_term,
        'category_labels': json.dumps(category_labels),
        'category_data': json.dumps(category_data),
        'payment_labels': json.dumps(payment_labels),
        'payment_data': json.dumps(payment_data),
        'student_outstanding': student_outstanding,
        'all_fee_payments': payments_page,
    }
    return render(request, 'dashboards/admin_fees.html', context)

# --- Error handlers ---
def _safe_back_redirect(request, fallback_url='/'):
    """Redirect to HTTP_REFERER if safe and same-origin; otherwise to fallback."""
    from django.conf import settings as _settings
    referer = request.META.get('HTTP_REFERER', '')
    if referer and url_has_allowed_host_and_scheme(referer, allowed_hosts=set(_settings.ALLOWED_HOSTS)):
        return redirect(referer)
    try:
        # Prefer home/dashboard for authenticated users; login for anonymous
        if request.user.is_authenticated:
            return redirect('/')
        else:
            from django.urls import reverse as _reverse
            return redirect(_reverse('login'))
    except Exception:
        return redirect(fallback_url)

def custom_404(request, exception):
    """Project-wide 404 handler: redirect back to previous location if available."""
    # Suppress noisy 404 notifications; redirect back silently
    return _safe_back_redirect(request)

def custom_404_catchall(request, unused_path=None):
    """Catch-all in DEBUG: redirect back rather than rendering the 404 page."""
    # Suppress noisy 404 notifications; redirect back silently
    return _safe_back_redirect(request)

def preview_404(request):
    """Development helper to preview the 404 page while DEBUG=True."""
    from django.conf import settings as _settings
    from django.http import HttpResponseNotFound
    from django.shortcuts import render as _render
    # Only allow when DEBUG=True to avoid exposing in production
    if not getattr(_settings, 'DEBUG', False):
        return HttpResponseNotFound('Not Found')
    # Also clear session in preview to match production behavior
    try:
        from django.contrib.auth import logout as _logout
        _logout(request)
    finally:
        try:
            request.session.flush()
        except Exception:
            pass
    return _render(request, '404.html', status=404)

def custom_403(request, exception=None):
    """Permission denied handler: redirect back to previous safe page."""
    messages.error(request, 'You do not have permission to access that page.')
    return _safe_back_redirect(request)

@login_required(login_url='login')
def admin_payment_history_partial(request):
    """Return only the History of Payments partial for AJAX updates.
    Supports filtering by `payment_search` and sorting by `payment_sort` (newest|oldest).
    Only admins are allowed.
    """
    if getattr(request.user, 'role', None) != 'admin':
        return HttpResponseForbidden('You are not authorized to access this resource.')

    from .models import FeePayment
    from django.db.models import Q

    payment_sort = request.GET.get('payment_sort', 'newest')
    payment_search = request.GET.get('payment_search', '').strip()

    payments_qs = FeePayment.objects.select_related('student__user', 'fee_assignment__fee_category')
    if payment_search:
        payments_qs = payments_qs.filter(
            Q(student__admission_number__icontains=payment_search) |
            Q(student__user__username__icontains=payment_search) |
            Q(student__user__first_name__icontains=payment_search) |
            Q(student__user__last_name__icontains=payment_search)
        )
    if payment_sort == 'oldest':
        payments_qs = payments_qs.order_by('payment_date')
    else:
        payments_qs = payments_qs.order_by('-payment_date')

    context = {
        'all_fee_payments': payments_qs,
    }
    # Render the same partial used by the History tab
    return render(request, 'dashboards/admin_payment_history.html', context)

@login_required(login_url='login')
def admin_users(request):
    # Admin-only access
    if getattr(request.user, 'role', None) != 'admin':
        return redirect('login')
    search_query = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')

    users = User.objects.all()

    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    if role_filter:
        users = users.filter(role=role_filter)

    role_choices = User._meta.get_field('role').choices

    return render(request, 'dashboards/admin_users.html', {
        'users': users,
        'search_query': search_query,
        'role_filter': role_filter,
        'role_choices': role_choices
    })

# API and AJAX placeholders
def exam_events_api(request):
    return JsonResponse([], safe=False)

@require_POST
def add_student_ajax(request):
    return JsonResponse({'status': 'ok'})

@require_POST
def add_teacher_ajax(request):
    return JsonResponse({'status': 'ok'})

@require_POST
def add_class_ajax(request):
    return JsonResponse({'status': 'ok'})

@require_POST
def add_subject_ajax(request):
    return JsonResponse({'status': 'ok'})

from django.core.serializers.json import DjangoJSONEncoder

def events_json(request):
    from .models import Event, Exam, TeacherResponsibility, Teacher
    from django.utils import timezone
    from django.core.serializers.json import DjangoJSONEncoder
    from django.utils.dateparse import parse_datetime
    from datetime import timedelta
    now = timezone.now()
    if not request.user.is_authenticated or request.user.role not in ['admin', 'teacher']:
        return JsonResponse([], safe=False)
    # Optional window from FullCalendar (ISO strings)
    start_param = request.GET.get('start')
    end_param = request.GET.get('end')
    try:
        window_start = parse_datetime(start_param) if start_param else None
        window_end = parse_datetime(end_param) if end_param else None
    except Exception:
        window_start = window_end = None

    # Build base queryset
    events_qs = Event.objects.filter(is_done=False)
    if window_start:
        # Include open-ended events (end is NULL) so they are not excluded
        from django.db.models import Q
        events_qs = events_qs.filter(Q(end__gte=window_start) | Q(end__isnull=True))  # overlapping or open-ended
    if window_end:
        events_qs = events_qs.filter(start__lte=window_end)
    # Fallback window ±1 year to avoid empty feed if params missing
    if not window_start and not window_end:
        one_year = timedelta(days=365)
        events_qs = events_qs.filter(start__gte=now - one_year, start__lte=now + one_year)
    # Ensure start exists
    events_qs = events_qs.exclude(start__isnull=True).order_by('start')
    event_list = []
    for e in events_qs:
        event_list.append({
            'id': f'event-{e.id}',
            'title': e.title,
            'start': e.start.isoformat(),
            'end': e.end.isoformat() if e.end else None,
            'allDay': e.all_day,
            'category': e.category,
        })
    # Exam events (add as allDay events)
    exams_qs = Exam.objects.select_related('term').all()
    # Filter exams into the same window if provided
    if window_start:
        exams_qs = exams_qs.filter(end_date__gte=window_start)  # overlapping
    if window_end:
        exams_qs = exams_qs.filter(start_date__lte=window_end)
    for ex in exams_qs:
        # Use start_date and end_date for calendar events (allDay: end should be exclusive)
        if not ex.start_date:
            continue  # skip exams without a start date (FullCalendar requires start)
        start = ex.start_date.isoformat()
        # FullCalendar expects allDay 'end' to be exclusive
        try:
            end_base = ex.end_date or ex.start_date
            end = (end_base + timedelta(days=1)).isoformat()
        except Exception:
            end = start
        event_list.append({
            'id': f'exam-{ex.id}',
            'title': ex.name,
            'start': start,
            'end': end,
            'allDay': True,
            'category': 'exam',
            'term': str(ex.term),
            'level': ex.level,
            'type': ex.get_type_display() if hasattr(ex, 'get_type_display') else '',
        })

    # Teacher responsibilities for the logged-in teacher
    try:
        # If the current user maps to a Teacher, include their responsibilities
        teacher = getattr(request.user, 'teacher', None)
        if not teacher:
            teacher = Teacher.objects.select_related('user').filter(user=request.user).first()
        if teacher:
            from django.db.models import Q
            resp_qs = TeacherResponsibility.objects.filter(teacher=teacher)
            # Window filtering: treat as allDay from start_date to end_date (inclusive)
            # Include NULLs on the opposite bound so open-ended responsibilities still appear
            if window_start:
                resp_qs = resp_qs.filter(Q(end_date__gte=window_start.date()) | Q(end_date__isnull=True))
            if window_end:
                resp_qs = resp_qs.filter(Q(start_date__lte=window_end.date()) | Q(start_date__isnull=True))
            resp_added = 0
            for r in resp_qs:
                    # derive start/end
                    base_start = r.start_date or (r.assigned_at.date() if getattr(r, 'assigned_at', None) else timezone.now().date())
                    base_end = r.end_date or r.start_date or base_start
                    r_start = base_start.isoformat()
                    # Exclusive end for allDay
                    try:
                        r_end = (base_end + timedelta(days=1)).isoformat()
                    except Exception:
                        r_end = base_start.isoformat()
                    event_list.append({
                        'id': f'resp-{r.id}',
                        'title': r.responsibility,
                        'start': r_start,
                        'end': r_end,
                        'allDay': True,
                        'category': 'responsibility',
                        'details': r.details or '',
                    })
                    resp_added += 1
            try:
                print(f"[events_json] responsibilities added: {resp_added} for teacher={teacher.id}")
            except Exception:
                pass
    except Exception:
        # keep feed resilient
        pass
    try:
        print(f"[events_json] user={request.user.id} role={getattr(request.user,'role',None)} events={len(event_list)} window=({window_start},{window_end})")
    except Exception:
        pass
    return JsonResponse(event_list, safe=False, encoder=DjangoJSONEncoder)


from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from .forms import EventForm
from .models import Term
from django.utils import timezone
import json

@login_required(login_url='login')
@require_POST
def event_create(request):
    print('--- [DEBUG] event_create called ---')
    print('[DEBUG] User:', request.user, 'Role:', getattr(request.user, 'role', None))
    print('[DEBUG] Content-Type:', request.content_type)
    print('[DEBUG] POST:', dict(request.POST))
    # print('[DEBUG] BODY:', request.body)
    
    # Only allow admins and teachers to create events
    if not hasattr(request.user, 'role') or request.user.role not in ['admin', 'teacher']:
        return JsonResponse({'success': False, 'error': 'Unauthorized: Only admins and teachers can add events.'}, status=403)
    try:
        print('[EVENT_CREATE][DEBUG] user:', request.user, 'role:', getattr(request.user, 'role', None))
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.copy()
        print('[EVENT_CREATE][DEBUG] incoming data:', dict(data))
        # If term is not provided, assign current term
        if not data.get('term'):
            today = timezone.now().date()
            current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('-start_date').first()
            if not current_term:
                current_term = Term.objects.order_by('-start_date').first()
            if current_term:
                data['term'] = current_term.id
        print('[EVENT_CREATE][DEBUG] data after term assignment:', dict(data))
        form = EventForm(data)
        print('[EVENT_CREATE][DEBUG] form.is_valid:', form.is_valid())
        print('[EVENT_CREATE][DEBUG] form.errors:', form.errors)
        if form.is_valid():
            event = form.save()
            print('[EVENT_CREATE][DEBUG] event created:', event)
            return JsonResponse({
                'success': True,
                'event': {
                    'id': event.id,
                    'title': event.title,
                    'start': event.start.isoformat(),
                    'end': event.end.isoformat() if event.end else None,
                    'allDay': event.all_day,
                    'category': event.category,
                }
            })
        else:
            print('[EVENT_CREATE][ERRORS]', form.errors)
            error_msg = 'Please correct the following errors: ' + '; '.join([f'{field}: {errs[0]}' for field, errs in form.errors.items()])
            return JsonResponse({'success': False, 'errors': form.errors, 'error': error_msg}, status=400)
    except Exception as e:
        print('[EVENT_CREATE][EXCEPTION]', str(e))
        return JsonResponse({'success': False, 'error': 'An unexpected error occurred while adding the event.'}, status=500)



@require_POST
def event_update(request):
    return JsonResponse({'status': 'ok'})

@require_POST
def event_delete(request):
    return JsonResponse({'status': 'ok'})

def events_feed(request):
    return JsonResponse([], safe=False)

def custom_login_view(request):
    # Always log out any existing user/session on visiting login page
    from django.contrib.auth import logout
    if request.user.is_authenticated:
        logout(request)
        # Optionally, add a message
        # messages.info(request, 'You have been logged out.')

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        role = request.POST['role']

        user = authenticate(request, username=username, password=password)

        if user is not None and user.role == role:
            login(request, user)
            if role == 'admin':
                return redirect('admin_overview')
            elif role == 'teacher':
                # Get teacher id and redirect properly
                if hasattr(user, 'teacher'):
                    return redirect('teacher_dashboard', teacher_id=user.teacher.id)
                else:
                    messages.error(request, 'No teacher profile found for this account.')
                    return redirect('login')
            elif role == 'student':
                if hasattr(user, 'student'):
                    return redirect('student_profile', student_id=user.student.id)
                else:
                    messages.error(request, 'User not a student.')
                    return redirect('login')
            elif role == 'clerk':
                return redirect('clerk_overview')
        else:
            messages.error(request, 'Invalid credentials or role mismatch.')

    return render(request, 'auth/login.html')

# AJAX partials for deferred loading (students, classes, exams)
from django.contrib.auth.decorators import login_required
@login_required(login_url='login')
def admin_students_partial(request):
    if getattr(request.user, 'role', None) != 'admin':
        return redirect('login')
    from .models import Student, Class
    from django.core.paginator import Paginator
    q = request.GET.get('q', '').strip()
    class_id = request.GET.get('class_id')
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 20))

    students = Student.objects.select_related('user', 'class_group').all().order_by('admission_no')
    if class_id:
        try:
            students = students.filter(class_group_id=int(class_id))
        except (TypeError, ValueError):
            pass
    if q:
        from django.db.models import Q
        students = students.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(admission_no__icontains=q)
        )
    paginator = Paginator(students, per_page)
    students_page = paginator.get_page(page)
    context = {
        'students_page': students_page,
        'q': q,
        'class_id': class_id,
    }
    return render(request, 'dashboards/partials/admin_students_table_partial.html', context)


@login_required(login_url='login')
def admin_classes_partial(request):
    if getattr(request.user, 'role', None) != 'admin':
        return redirect('login')
    from .models import Class, TeacherClassAssignment, Teacher, Subject
    from django.core.paginator import Paginator
    level = request.GET.get('level')
    stream = request.GET.get('stream')
    page = request.GET.get('page', 1)
    show_mode = (request.GET.get('show') or '').lower()
    # Default paged view; allow show=all to override
    per_page = int(request.GET.get('per_page', 12))
    if show_mode == 'all':
        per_page = 100000  # effectively all

    classes = Class.objects.all().order_by('name')
    if level:
        try:
            classes = classes.filter(level=int(level))
        except (TypeError, ValueError):
            pass
    if stream:
        classes = classes.filter(stream__iexact=stream)

    # Prefetch related teacher/subjects information if models exist
    # Minimal context to render cards; detailed modals can still be in the main template
    paginator = Paginator(classes, per_page)
    classes_page = paginator.get_page(page)
    # Build a mapping of class -> teacher/subjects summaries (best-effort)
    class_summaries = {}
    class_ids = [c.id for c in classes_page.object_list]
    try:
        assignments = TeacherClassAssignment.objects.select_related('teacher', 'class_group').filter(class_group_id__in=class_ids)
        for a in assignments:
            class_summaries.setdefault(a.class_group_id, {'teachers': set(), 'subjects': set()})
            if a.teacher:
                class_summaries[a.class_group_id]['teachers'].add(str(a.teacher))
            subj = getattr(a, 'subject', None)
            if subj:
                class_summaries[a.class_group_id]['subjects'].add(str(subj))
    except Exception:
        pass
    context = {
        'classes_page': classes_page,
        'class_summaries': class_summaries,
        'level': level,
        'stream': stream,
    }
    return render(request, 'dashboards/partials/admin_classes_grid_partial.html', context)


@login_required(login_url='login')
def admin_exams_available_partial(request):
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    user = request.user
    if not (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser):
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
        return redirect('login')
    from .models import Exam
    from django.core.paginator import Paginator
    q = request.GET.get('q', '').strip()
    term_id = request.GET.get('term_id')
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 20))

    exams = Exam.objects.select_related('term').all().order_by('-start_date')
    if term_id:
        try:
            exams = exams.filter(term_id=int(term_id))
        except (TypeError, ValueError):
            pass
    if q:
        exams = exams.filter(name__icontains=q)
    paginator = Paginator(exams, per_page)
    exams_page = paginator.get_page(page)
    context = {
        'exams_page': exams_page,
        'q': q,
        'term_id': term_id,
    }
    return render(request, 'dashboards/partials/admin_exams_available_partial.html', context)

# Admin Exams View
from django.contrib.auth.decorators import login_required
@login_required(login_url='login')
def admin_exams(request):
    from .forms import ExamForm
    from .models import Exam, Subject, Student, Grade, Class
    from django.contrib import messages
    import pandas as pd
    from django.db import transaction
    # Admin-only access
    if getattr(request.user, 'role', None) != 'admin':
        return redirect('login')

    exam_form = ExamForm(request.POST or None)
    exams = Exam.objects.select_related('term').order_by('-start_date')
    latest_exam = exams.first() if exams.exists() else None
    table_html = None
    summary = {'added': 0, 'removed': 0, 'errors': []}

    # Grades card support: get selected exam, students, and grades
    selected_exam = None
    students = []
    student_grades = {}
    exam_id = request.GET.get('exam_id')
    if exam_id:
        try:
            selected_exam = Exam.objects.get(id=exam_id)
            # Get all students for the exam's level
            students = Student.objects.filter(class_group__level=selected_exam.level).select_related('user')
            # Get all grades for this exam
            grades = Grade.objects.filter(exam=selected_exam)
            for grade in grades:
                student_grades[grade.student_id] = grade
        except Exam.DoesNotExist:
            selected_exam = None

    # Handle exam creation POST
    if request.method == 'POST':
        if 'add_exam' in request.POST:
            form = ExamForm(request.POST)
            if form.is_valid():
                if form.cleaned_data['end_date'] < form.cleaned_data['start_date']:
                    messages.error(request, 'End date cannot be before the start date.')
                else:
                    try:
                        form.save()
                        messages.success(request, 'Exam added successfully!')
                        return redirect('admin_exams')
                    except IntegrityError:
                        messages.error(request, 'An exam with this name already exists.')
            else:
                messages.error(request, 'Error adding exam. Please check the form.')

    # Helper: Build student-subject matrix for the latest exam
    def build_students_subjects_table():
        all_subjects = list(Subject.objects.all().order_by('name'))
        all_classes = list(Class.objects.all().order_by('name'))
        students_subjects_table = []
        for class_obj in all_classes:
            students = Student.objects.filter(class_group=class_obj).order_by('admission_no')
            for student in students:
                student_subjects = set(Grade.objects.filter(student=student).values_list('subject_id', flat=True))
                row = {
                    'admission_no': student.admission_no,
                    'name': student.full_name,
                    'class': class_obj.name,
                    'subjects': [subj.id in student_subjects for subj in all_subjects],
                    'subject_ids': list(student_subjects),
                }
                students_subjects_table.append(row)
        return students_subjects_table, all_subjects

    # Handle grade entry matrix POST
    if request.method == 'POST' and request.POST.get('submit_matrix'):
        updated = 0
        for key, value in request.POST.items():
            if key.startswith('marks-'):
                try:
                    _, admission_no, subject_id = key.split('-', 2)
                    score = value.strip()
                    if score == '':
                        continue
                    score = float(score)
                    student = Student.objects.get(admission_no=admission_no)
                    subject = Subject.objects.get(id=subject_id)
                    exam = latest_exam
                    if not exam:
                        continue
                    Grade.objects.update_or_create(
                        student=student,
                        subject=subject,
                        exam=exam,
                        defaults={'score': score}
                    )
                    updated += 1
                except Exception:
                    continue
        messages.success(request, f"Successfully updated {updated} marks from matrix form.")

    # Handle bulk marksheet upload
    if request.method == 'POST' and request.FILES.get('marksheet'):
        excel_file = request.FILES['marksheet']
        is_preview = 'preview' in request.POST
        is_process = 'process' in request.POST
        try:
            df = pd.read_excel(excel_file)
            table_html = df.to_html(classes='table table-bordered table-striped', index=False)
            required_cols = ['admission_no', 'first_name', 'last_name', 'gender', 'birthdate', 'class_group']
            for col in required_cols:
                if col not in df.columns:
                    summary['errors'].append(f"Missing required column: {col}")
            if summary['errors']:
                messages.error(request, ' '.join(summary['errors']))
                students_subjects_table, all_subjects = build_students_subjects_table()
                return render(request, 'dashboards/admin_exams.html', {
                    'table_html': table_html,
                    'students_subjects_table': students_subjects_table,
                    'all_subjects': all_subjects,
                    'exam_form': exam_form,
                    'exams': exams,
                })
            # Clean and check for duplicates/missing
            seen = set()
            valid_rows = []
            for idx, row in df.iterrows():
                row_errors = []
                admission_no = str(row.get('admission_no', '')).strip()
                first_name = str(row.get('first_name', '')).strip()
                last_name = str(row.get('last_name', '')).strip()
                gender = str(row.get('gender', '')).strip()
                birthdate = row.get('birthdate')
                class_name = str(row.get('class_group', '')).strip()
                if not all([admission_no, first_name, last_name, gender, birthdate, class_name]):
                    row_errors.append(f"Row {idx+2}: Missing required data.")
                if admission_no in seen:
                    row_errors.append(f"Row {idx+2}: Duplicate admission_no {admission_no}.")
                seen.add(admission_no)
                if row_errors:
                    summary['errors'].extend(row_errors)
                else:
                    valid_rows.append({
                        'admission_no': admission_no,
                        'first_name': first_name,
                        'last_name': last_name,
                        'gender': gender,
                        'birthdate': birthdate,
                        'class_group': class_name
                    })
            if not valid_rows:
                messages.error(request, 'No valid students to import.')
                students_subjects_table, all_subjects = build_students_subjects_table()
                return render(request, 'dashboards/admin_exams.html', {
                    'table_html': table_html,
                    'students_subjects_table': students_subjects_table,
                    'all_subjects': all_subjects,
                    'exam_form': exam_form,
                    'exams': exams,
                })
            if is_preview:
                if summary['errors']:
                    messages.error(request, ' '.join(summary['errors']))
                else:
                    messages.info(request, 'Preview successful. No data has been added yet.')
                students_subjects_table, all_subjects = build_students_subjects_table()
                return render(request, 'dashboards/admin_exams.html', {
                    'table_html': table_html,
                    'students_subjects_table': students_subjects_table,
                    'all_subjects': all_subjects,
                    'exam_form': exam_form,
                    'exams': exams,
                })
            if is_process:
                from .models import Exam, Term, AcademicYear, User
                import datetime
                class_name = valid_rows[0]['class_group']
                class_group, _ = Class.objects.get_or_create(name=class_name)
                uploaded_adm_nos = set([row['admission_no'] for row in valid_rows])
                grades_processed = 0
                with transaction.atomic():
                    to_remove = Student.objects.filter(class_group=class_group).exclude(admission_no__in=uploaded_adm_nos)
                    for st in to_remove:
                        if st.user:
                            st.user.delete()
                        st.delete()
                        summary['removed'] += 1
                    for row in valid_rows:
                        if not Student.objects.filter(admission_no=row['admission_no']).exists():
                            user = User.objects.create_user(
                                username=row['admission_no'],
                                first_name=row['first_name'],
                                last_name=row['last_name'],
                                role='student'
                            )
                            user.set_password('student1234')
                            user.save()
                            Student.objects.create(
                                user=user,
                                admission_no=row['admission_no'],
                                class_group=class_group,
                                gender=row['gender'],
                                birthdate=row['birthdate']
                            )
                            summary['added'] += 1
                    year = str(datetime.date.today().year)
                    academic_year, _ = AcademicYear.objects.get_or_create(year=year)
                    term, _ = Term.objects.get_or_create(name='Term 1', academic_year=academic_year)
                    exam_name = f'Opener exams for {class_group.name}'
                    exam, _ = Exam.objects.get_or_create(name=exam_name, term=term, defaults={'date': datetime.date.today()})
                    subj_code_map = {s.code: s for s in Subject.objects.all()}
                    subject_cols = [col for col in df.columns if col not in ['admission_no', 'first_name', 'last_name', 'gender', 'birthdate', 'class_group']]
                    for idx, row in df.iterrows():
                        try:
                            student = Student.objects.get(admission_no=str(row['admission_no']).strip())
                        except Student.DoesNotExist:
                            continue
                        for subj_col in subject_cols:
                            score = row.get(subj_col, None)
                            if pd.isna(score):
                                continue
                            try:
                                score = float(score)
                            except (TypeError, ValueError):
                                continue
                            subj_code = subj_col
                            subject = subj_code_map.get(subj_code)
                            if not subject:
                                subject = Subject.objects.filter(name__iexact=subj_code).first()
                            if not subject:
                                continue
                            Grade.objects.update_or_create(
                                student=student,
                                subject=subject,
                                exam=exam,
                                defaults={'score': score}
                            )
                            grades_processed += 1
                messages.success(request, f"Added: {summary['added']}, Removed: {summary['removed']}, Grades processed: {grades_processed}. Errors: {'; '.join(summary['errors']) if summary['errors'] else 'None'}.")
        except Exception as e:
            messages.error(request, f'Error reading Excel file: {e}')
    students_subjects_table, all_subjects = build_students_subjects_table()
    context = {
        'exam_form': exam_form,
        'exams': exams,
        'table_html': table_html,
        'students_subjects_table': students_subjects_table,
        'all_subjects': all_subjects,
    }
    return render(request, 'dashboards/admin_exams.html', context)

@login_required(login_url='login')
@require_POST
def delete_exam(request, exam_id):
    if not request.user.role == 'admin':
        return HttpResponseForbidden("You are not authorized to delete exams.")
    exam = get_object_or_404(Exam, id=exam_id)
    exam.delete()
    messages.success(request, 'Exam deleted successfully!')
    return redirect('admin_exams')

@login_required(login_url='login')
def exam_calendar_api(request):
    """API endpoint to provide exam data for FullCalendar"""
    try:
        exams = Exam.objects.select_related('term').all()
        events = []
        for exam in exams:
            # Calculate end date for calendar display (add 1 day if same day)
            end_date = exam.end_date
            if exam.start_date == exam.end_date:
                # For single-day events, FullCalendar needs the end date to be the next day
                from datetime import timedelta
                end_date = exam.end_date + timedelta(days=1)
            
            events.append({
                'id': exam.id,
                'title': exam.name,
                'start': exam.start_date.isoformat(),
                'end': end_date.isoformat(),
                'extendedProps': {
                    'term': exam.term.name if exam.term else 'No Term',
                    'type': exam.get_type_display() if hasattr(exam, 'get_type_display') else 'Exam',
                    'description': f"{exam.name} - {exam.term.name if exam.term else 'No Term'}"
                },
                'className': 'fc-event-exam',
                'color': '#0d47a1'
            })
        return JsonResponse(events, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def api_classes(request):
    """API endpoint to get all classes"""
    try:
        classes = Class.objects.all().order_by('name')
        classes_data = [{
            'id': cls.id,
            'name': cls.name,
            'level': cls.level,
            'display_name': f"{cls.name} (Level {cls.level})"
        } for cls in classes]
        return JsonResponse({'classes': classes_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def api_subjects(request):
    """API endpoint to get all subjects"""
    try:
        subjects = Subject.objects.all().order_by('name')
        subjects_data = [{
            'id': subject.id,
            'name': subject.name,
            'code': getattr(subject, 'code', ''),
        } for subject in subjects]
        return JsonResponse({'subjects': subjects_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def api_students_by_class(request, class_id):
    """API endpoint to get students by class ID"""
    try:
        class_obj = get_object_or_404(Class, id=class_id)
        students = Student.objects.filter(class_group=class_obj).select_related('user').order_by('admission_no')
        
        students_data = []
        for student in students:
            name = student.full_name if hasattr(student, 'full_name') else str(student)
            students_data.append({
                'id': student.id,
                'admission_no': student.admission_no,
                'name': name,
                'user_id': student.user.id if student.user else None
            })
        
        return JsonResponse({
            'students': students_data,
            'class_name': str(class_obj)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def api_exam_subjects(request, exam_id):
    """API endpoint to get subjects for a specific exam"""
    try:
        exam = get_object_or_404(Exam, id=exam_id)
        # Get subjects that are associated with classes or all subjects
        subjects = Subject.objects.all().order_by('name')
        
        subjects_data = [{
            'id': subject.id,
            'name': subject.name,
            'code': getattr(subject, 'code', ''),
        } for subject in subjects]
        
        return JsonResponse({
            'subjects': subjects_data,
            'exam_name': exam.name
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def api_bulk_grades(request):
    """API endpoint to save multiple grades at once"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        grades_data = data.get('grades', [])
        
        if not grades_data:
            return JsonResponse({'error': 'No grades data provided'}, status=400)
        
        saved_count = 0
        errors = []
        
        for grade_item in grades_data:
            try:
                student_id = grade_item.get('student_id')
                subject_id = grade_item.get('subject_id')
                exam_id = grade_item.get('exam_id')
                grade_value = grade_item.get('grade')
                
                if not all([student_id, subject_id, exam_id, grade_value is not None]):
                    errors.append(f"Missing required fields for grade entry")
                    continue
                
                # Get the objects
                student = get_object_or_404(Student, id=student_id)
                subject = get_object_or_404(Subject, id=subject_id)
                exam = get_object_or_404(Exam, id=exam_id)
                
                # Create or update grade
                grade, created = Grade.objects.update_or_create(
                    student=student,
                    subject=subject,
                    exam=exam,
                    defaults={'grade': grade_value}
                )
                saved_count += 1
                
            except Exception as e:
                errors.append(f"Error saving grade for student {student_id}: {str(e)}")
                continue
        
        return JsonResponse({
            'success': True,
            'saved_count': saved_count,
            'errors': errors
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def api_upload_bulk_grades(request):
    """API endpoint to upload grades via Excel file"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    
    try:
        import pandas as pd
        from io import BytesIO
        
        excel_file = request.FILES.get('excel_file')
        exam_id = request.POST.get('exam_id')
        
        if not excel_file:
            return JsonResponse({'error': 'No Excel file provided'}, status=400)
        
        if not exam_id:
            return JsonResponse({'error': 'No exam ID provided'}, status=400)
        
        # Read Excel file
        try:
            df = pd.read_excel(BytesIO(excel_file.read()))
        except Exception as e:
            return JsonResponse({'error': f'Error reading Excel file: {str(e)}'}, status=400)
        
        # Validate required columns
        required_columns = ['admission_no', 'subject', 'grade']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'error': f'Missing required columns: {missing_columns}. Required: {required_columns}'
            }, status=400)
        
        exam = get_object_or_404(Exam, id=exam_id)
        processed_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                admission_no = str(row['admission_no']).strip()
                subject_name = str(row['subject']).strip()
                grade_value = float(row['grade'])
                
                # Find student by admission number
                try:
                    student = Student.objects.get(admission_no=admission_no)
                except Student.DoesNotExist:
                    errors.append(f"Row {index + 2}: Student with admission number {admission_no} not found")
                    continue
                
                # Find subject by name
                try:
                    subject = Subject.objects.get(name__iexact=subject_name)
                except Subject.DoesNotExist:
                    errors.append(f"Row {index + 2}: Subject '{subject_name}' not found")
                    continue
                
                # Validate grade value
                if not (0 <= grade_value <= 100):
                    errors.append(f"Row {index + 2}: Grade {grade_value} is not between 0 and 100")
                    continue
                
                # Create or update grade
                grade, created = Grade.objects.update_or_create(
                    student=student,
                    subject=subject,
                    exam=exam,
                    defaults={'grade': grade_value}
                )
                processed_count += 1
                
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                continue
        
        return JsonResponse({
            'success': True,
            'processed_count': processed_count,
            'errors': errors[:10]  # Limit errors to first 10
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def api_download_grade_template(request):
    """API endpoint to download grade template Excel file"""
    try:
        import pandas as pd
        from django.http import HttpResponse
        from io import BytesIO
        
        exam_id = request.GET.get('exam_id')
        class_id = request.GET.get('class_id')
        subject_id = request.GET.get('subject_id')
        
        # Create sample data
        template_data = []
        
        if class_id:
            # Get students from specific class
            students = Student.objects.filter(class_group_id=class_id).select_related('user')
            for student in students:
                template_data.append({
                    'admission_no': student.admission_no,
                    'student_name': student.full_name if hasattr(student, 'full_name') else str(student),
                    'subject': 'Mathematics',  # Example subject
                    'grade': ''  # Empty for user to fill
                })
        else:
            # Create sample template
            template_data = [
                {'admission_no': 'STU001', 'student_name': 'John Doe', 'subject': 'Mathematics', 'grade': ''},
                {'admission_no': 'STU002', 'student_name': 'Jane Smith', 'subject': 'Mathematics', 'grade': ''},
                {'admission_no': 'STU003', 'student_name': 'Bob Johnson', 'subject': 'English', 'grade': ''},
            ]
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Grades', index=False)
            # Insert school header rows using openpyxl
            try:
                from landing.models import SiteSettings
                ss = SiteSettings.objects.first()
                if ss:
                    wb = writer.book
                    ws = wb['Grades']
                    # Insert 3 rows at top
                    ws.insert_rows(1, amount=3)
                    ws['A1'] = getattr(ss, 'school_name', '')
                    ws['A2'] = getattr(ss, 'school_motto', '')
                    ws['A3'] = getattr(ss, 'contact_address', '')
            except Exception:
                pass
            
            # Add instructions sheet
            instructions = pd.DataFrame({
                'Instructions': [
                    '1. Fill in the grade column with values between 0 and 100',
                    '2. Do not modify the admission_no or student_name columns',
                    '3. Subject names must match exactly with subjects in the system',
                    '4. Save the file and upload it using the Excel Upload tab',
                    '5. Empty grade cells will be ignored during upload'
                ]
            })
            instructions.to_excel(writer, sheet_name='Instructions', index=False)
        
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="grade_template_{exam_id or "general"}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def api_download_class_students(request, class_id):
    """API endpoint to download list of students in a class"""
    try:
        import pandas as pd
        from django.http import HttpResponse
        from io import BytesIO
        
        class_obj = get_object_or_404(Class, id=class_id)
        students = Student.objects.filter(class_group=class_obj).select_related('user').order_by('admission_no')
        
        # Create student data
        student_data = []
        for student in students:
            student_data.append({
                'admission_no': student.admission_no,
                'student_name': student.full_name if hasattr(student, 'full_name') else str(student),
                'class': str(class_obj),
                'gender': getattr(student, 'gender', ''),
                'phone': getattr(student, 'phone', ''),
            })
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(student_data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Students', index=False)
            # Insert school header rows using openpyxl
            try:
                from landing.models import SiteSettings
                ss = SiteSettings.objects.first()
                if ss:
                    wb = writer.book
                    ws = wb['Students']
                    # Insert 3 rows at top
                    ws.insert_rows(1, amount=3)
                    ws['A1'] = getattr(ss, 'school_name', '')
                    ws['A2'] = getattr(ss, 'school_motto', '')
                    ws['A3'] = getattr(ss, 'contact_address', '')
            except Exception:
                pass
        
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="students_{class_obj.name}_{class_id}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# --- Messaging Module ---
from .models import Message
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import Q

from .forms import MessagingForm
from django.core.mail import send_mail
from .models import Class, Subject, Teacher, Term
from .services.timetable_scheduler import generate_timetable

@login_required(login_url='login')
def admin_send_message(request):
    category = request.GET.get('user_category')
    form = MessagingForm(request.POST or None, category=category)
    sent = False
    # Admin-only access
    if getattr(request.user, 'role', None) != 'admin':
        return redirect('dashboard')
    if request.method == 'POST' and form.is_valid():
        sender = request.user
        recipient = form.cleaned_data['recipient']
        subject = form.cleaned_data['subject']
        message_body = form.cleaned_data['message']
        send_email = form.cleaned_data['send_email']
        send_sms = form.cleaned_data['send_sms']

        # Save message to DB
        Message.objects.create(sender=sender, recipient=recipient, subject=subject, content=message_body)

        # Send email if requested
        if send_email and recipient.email:
            send_mail(subject, message_body, sender.email, [recipient.email], fail_silently=True)
        # Send SMS if requested (placeholder, implement actual SMS logic)
        if send_sms and hasattr(recipient, 'phone') and recipient.phone:
            pass  # Integrate with SMS API here

        messages.success(request, 'Message sent successfully!')
        sent = True
        form = MessagingForm(category=category)  # Reset form

    # Filter recipients based on selected category
    # Fix: ensure category is lowercased and valid
    valid_roles = [choice[0] for choice in User.ROLE_CHOICES]
    if category and category.lower() in valid_roles:
        form.fields['recipient'].queryset = User.objects.filter(role=category.lower())
    else:
        form.fields['recipient'].queryset = User.objects.none()

    return render(request, 'messaging/send_message.html', {'form': form, 'sent': sent})

@login_required(login_url='login')
def admin_overview(request):
    """Admin landing page with quick stats, upcoming events, and responsibilities."""
    if getattr(request.user, 'role', None) != 'admin':
        return HttpResponseForbidden('Forbidden')

    # Build dashboard context
    from django.utils import timezone
    from .models import (
        Student, Teacher, Class, Subject,
        FeeCategory, Event, Term, TeacherResponsibility,
    )
    from django.db.models import Q, Case, When, Value, IntegerField

    now = timezone.now()

    total_students = Student.objects.filter(graduated=False).count()
    students_boys = Student.objects.filter(graduated=False, gender__iexact='male').count()
    students_girls = Student.objects.filter(graduated=False, gender__iexact='female').count()
    total_teachers = Teacher.objects.count()
    teachers_male = Teacher.objects.filter(gender__iexact='male').count()
    teachers_female = Teacher.objects.filter(gender__iexact='female').count()
    total_classes = Class.objects.count()
    total_subjects = Subject.objects.count()
    total_fees = FeeCategory.objects.count()

    today = now.date()
    current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()

    upcoming_events = Event.objects.filter(is_done=False, start__gte=now).order_by('start')[:5]

    # Toggle to include expired responsibilities
    show_expired = request.GET.get('show_expired') == '1'

    if show_expired:
        # Expired: end_date before today. Show most recently expired first
        responsibilities = (
            TeacherResponsibility.objects
            .filter(end_date__lt=today)
            .select_related('teacher', 'teacher__user', 'assigned_by')
            .order_by('-end_date', '-assigned_at')[:8]
        )
    else:
        # Active: no end date or end_date today/future. Sort by soonest end_date first, with open-ended (null) last
        responsibilities = (
            TeacherResponsibility.objects
            .filter(Q(end_date__isnull=True) | Q(end_date__gte=today))
            .annotate(end_is_null=Case(When(end_date__isnull=True, then=Value(1)), default=Value(0), output_field=IntegerField()))
            .select_related('teacher', 'teacher__user', 'assigned_by')
            .order_by('end_is_null', 'end_date', '-assigned_at')[:8]
        )

    context = {
        'total_students': total_students,
        'students_boys': students_boys,
        'students_girls': students_girls,
        'total_teachers': total_teachers,
        'teachers_male': teachers_male,
        'teachers_female': teachers_female,
        'total_classes': total_classes,
        'total_subjects': total_subjects,
        'total_fees': total_fees,
        'current_term': current_term,
        'upcoming_events': upcoming_events,
        'responsibilities': responsibilities,
    }

    return render(request, 'dashboards/admin_overview.html', context)

@login_required(login_url='login')
def clerk_overview(request):
    # Only admin or clerk can access
    if not is_admin_or_clerk(request.user):
        return HttpResponseForbidden('Forbidden')
    from django.utils import timezone
    from .models import Term, Class, Student, FeeAssignment, FeePayment
    today = timezone.now().date()
    current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()

    # Compute totals for the current term
    fee_assignments = FeeAssignment.objects.all()
    if current_term:
        fee_assignments = fee_assignments.filter(term=current_term)

    # Total assigned = sum(amount * students_in_class) for assignments in term
    total_assigned = 0.0
    for fa in fee_assignments.select_related('class_group'):
        num_students = Student.objects.filter(class_group=fa.class_group, graduated=False).count()
        total_assigned += float(getattr(fa, 'amount', 0)) * num_students

    # Total paid for payments linked to those assignments
    from django.db.models import Sum
    total_paid = FeePayment.objects.filter(fee_assignment__in=fee_assignments).aggregate(total=Sum('amount_paid'))['total'] or 0.0
    total_outstanding = float(total_assigned) - float(total_paid)

    # Recent payments
    recent_payments = (
        FeePayment.objects.select_related('student__user', 'fee_assignment__fee_category')
        .order_by('-payment_date')[:10]
    )

    # Top classes by outstanding
    class_stats = []
    classes = Class.objects.all().order_by('level', 'name')
    for clazz in classes:
        class_assignments = fee_assignments.filter(class_group=clazz)
        assigned = 0.0
        for fa in class_assignments:
            num_students = Student.objects.filter(class_group=clazz, graduated=False).count()
            assigned += float(getattr(fa, 'amount', 0)) * num_students
        paid = FeePayment.objects.filter(fee_assignment__in=class_assignments).aggregate(total=Sum('amount_paid'))['total'] or 0.0
        outstanding = float(assigned) - float(paid)
        if assigned or paid:
            class_stats.append({'class': clazz, 'assigned': assigned, 'paid': paid, 'outstanding': outstanding})
    # Sort by outstanding desc and take top 8
    class_stats.sort(key=lambda r: r['outstanding'], reverse=True)
    class_stats = class_stats[:8]

    context = {
        'current_term': current_term,
        'total_assigned': total_assigned,
        'total_paid': total_paid,
        'total_outstanding': total_outstanding if total_outstanding > 0 else 0,
        'recent_payments': recent_payments,
        'class_stats': class_stats,
    }
    return render(request, 'dashboards/clerk_overview.html', context)

@login_required
def timetable_view(request):
    """Displays the timetable for a selected class."""
    # Handle admin-triggered auto-generate POST from template button
    if request.method == 'POST' and getattr(request.user, 'role', None) == 'admin' and request.POST.get('auto_generate') == '1':
        try:
            report = generate_timetable(overwrite=True)
            # Persist full report to session for display after redirect
            request.session['timetable_report'] = report
            placed = report.get('placed', 0)
            skipped = report.get('skipped', 0)
            messages.success(request, f"Timetable generated. Placed: {placed}, Skipped: {skipped}.")
        except Exception as e:
            messages.error(request, f"Failed to generate timetable: {e}")
        return redirect('timetable_view')
    classes = Class.objects.all().order_by('level', 'name')
    selected_class_id = request.GET.get('class_id')
    selected_class = None
    
    periods = PeriodSlot.objects.filter(is_class_slot=True).order_by('start_time')
    days = [day[0] for day in DefaultTimetable.DAY_CHOICES]
    timetable_grid = {p.id: {d: None for d in days} for p in periods}
    # Per-class capacity guidance
    selected_capacity = None
    selected_required = None
    subjects = Subject.objects.none()
    teachers = Teacher.objects.none()

    if selected_class_id:
        try:
            selected_class = Class.objects.get(id=selected_class_id)

            # Correctly fetch subjects and teachers from the TeacherClassAssignment model
            assignments = TeacherClassAssignment.objects.filter(class_group=selected_class).select_related('subject', 'teacher__user')

            # Build subject and teacher options from assignments; if none, fall back to all
            subject_ids = list(assignments.values_list('subject_id', flat=True).distinct())
            if subject_ids:
                # Exclude component (child) subjects from options
                subjects = Subject.objects.filter(id__in=subject_ids, part_of__isnull=True).order_by('name')
            else:
                # Fallback: all non-child subjects
                subjects = Subject.objects.filter(part_of__isnull=True).order_by('name')

            teacher_ids = list(assignments.values_list('teacher_id', flat=True).distinct())
            if teacher_ids:
                teachers = Teacher.objects.filter(id__in=teacher_ids).select_related('user').order_by('user__first_name', 'user__last_name')
            else:
                teachers = Teacher.objects.select_related('user').all().order_by('user__first_name', 'user__last_name')

            # Populate the grid with existing timetable entries
            entries = DefaultTimetable.objects.filter(class_group=selected_class)
            for entry in entries:
                if entry.period_id in timetable_grid:
                    timetable_grid[entry.period_id][entry.day] = entry

            # Compute capacity (available slots) vs demand (weekly lessons required)
            selected_capacity = periods.count() * len(days)
            # Use same rules as scheduler: exclude component subjects and enforce min weekly lessons
            required = 0
            for a in assignments:
                min_lessons = int(getattr(a.subject, 'min_weekly_lessons', 3) or 3)
                target_lessons = int(getattr(a.subject, 'weekly_lessons', min_lessons) or min_lessons)
                required += max(min_lessons, target_lessons)
            selected_required = required

        except Class.DoesNotExist:
            messages.error(request, "The selected class does not exist.")
            selected_class = None

    # Pull any recent generation report from the session (one-time display)
    timetable_report = request.session.pop('timetable_report', None)

    # Precompute capacity math for template (avoid complex filters)
    cap_diff = (selected_required - selected_capacity) if (selected_capacity is not None and selected_required is not None) else None
    capacity_over = (cap_diff if (cap_diff is not None and cap_diff > 0) else 0)
    capacity_free = ((-cap_diff) if (cap_diff is not None and cap_diff < 0) else 0)

    context = {
        'classes': classes,
        'selected_class': selected_class,
        'periods': periods,
        'days': days,
        'timetable_grid': timetable_grid,
        'subjects': subjects,
        'teachers': teachers,
        'timetable_report': timetable_report,
        'selected_capacity': selected_capacity,
        'selected_required': selected_required,
        'capacity_diff': cap_diff,
        'capacity_over': capacity_over,
        'capacity_free': capacity_free,
    }
    return render(request, 'timetable/timetable.html', context)


@login_required(login_url='login')
def school_timetable_day(request):
    """Render a comprehensive timetable for the entire school for one selected day.
    Grid: rows = PeriodSlot, columns = Class.
    """
    
# Use same day choices as the class timetable
    valid_days = [d[0] for d in DefaultTimetable.DAY_CHOICES]
    selected_day = request.GET.get('day') or (valid_days[0] if valid_days else 'Monday')
    if selected_day not in valid_days and valid_days:
        selected_day = valid_days[0]

    classes = Class.objects.all().order_by('level', 'name')
    periods = PeriodSlot.objects.filter(is_class_slot=True).order_by('start_time')

    # Fetch entries for all days across all classes (for printing the entire week)
    entries_all = (
        DefaultTimetable.objects
        .filter(day__in=valid_days)
        .select_related('class_group', 'subject', 'teacher__user', 'period')
    )

    # Build grid for the selected day: grid[period_id][class_id] = entry or None
    grid = {p.id: {c.id: None for c in classes} for p in periods}

    # Build grids for all days for printing
    grids_by_day = {day: {p.id: {c.id: None for c in classes} for p in periods} for day in valid_days}

    for e in entries_all:
        # Fill the per-day grids
        if e.period_id in grids_by_day.get(e.day, {}):
            if e.class_group_id in grids_by_day[e.day][e.period_id]:
                grids_by_day[e.day][e.period_id][e.class_group_id] = e
        # Also populate the selected-day grid
        if e.day == selected_day and e.period_id in grid and e.class_group_id in grid[e.period_id]:
            grid[e.period_id][e.class_group_id] = e

    context = {
        'selected_day': selected_day,
        'valid_days': valid_days,
        'classes': classes,
        'periods': periods,
        'grid': grid,
        'grids_by_day': grids_by_day,
    }
    return render(request, 'timetable/school_timetable_day.html', context)


@login_required(login_url='login')
def admin_period_slots(request):
    """Admin-only editor for PeriodSlot session times and labels."""
    if getattr(request.user, 'role', None) != 'admin':
        return HttpResponseForbidden('Permission denied.')

    from .models import PeriodSlot
    if request.method == 'POST':
        slot_id = request.POST.get('slot_id')
        label = (request.POST.get('label') or '').strip()
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        is_class_slot = True if request.POST.get('is_class_slot') == 'on' else False

        # Basic validation: start < end
        from datetime import time as dtime
        try:
            # Expecting HH:MM format from input type="time"
            h1, m1 = [int(x) for x in (start_time or '').split(':')]
            h2, m2 = [int(x) for x in (end_time or '').split(':')]
            t1 = dtime(h1, m1)
            t2 = dtime(h2, m2)
            if t1 >= t2:
                messages.error(request, 'Start time must be before end time.')
                return redirect('admin_period_slots')
        except Exception:
            messages.error(request, 'Invalid time format.')
            return redirect('admin_period_slots')

        if slot_id:
            # Update existing slot
            try:
                slot = PeriodSlot.objects.get(id=slot_id)
            except PeriodSlot.DoesNotExist:
                messages.error(request, 'Period not found.')
                return redirect('admin_period_slots')

            slot.label = label or slot.label
            slot.start_time = t1
            slot.end_time = t2
            slot.is_class_slot = is_class_slot
            slot.save()
            messages.success(request, f'Updated {slot.label}.')
        else:
            # Create new slot
            if not label:
                messages.error(request, 'Label is required to add a new period.')
                return redirect('admin_period_slots')
            slot = PeriodSlot.objects.create(
                label=label,
                start_time=t1,
                end_time=t2,
                is_class_slot=is_class_slot,
            )
            messages.success(request, f'Added {slot.label}.')
        return redirect('admin_period_slots')

    periods = PeriodSlot.objects.all().order_by('start_time')
    return render(request, 'timetable/period_slots.html', { 'periods': periods })

@login_required
@require_POST
def timetable_auto_generate(request):
    """Admin-only endpoint: auto-generate the entire school timetable, overwriting existing entries."""
    # Role check
    if getattr(request.user, 'role', None) != 'admin':
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

    try:
        report = generate_timetable(overwrite=True)
        messages.success(request, f"Timetable generated. Placed: {report.get('placed', 0)}, Skipped: {report.get('skipped', 0)}")
        # Optionally return JSON for AJAX; but we redirect back to timetable page
        return redirect('timetable_view')
    except Exception as e:
        messages.error(request, f"Failed to generate timetable: {e}")
        return redirect('timetable_view')
def timetable_edit_api(request):
    try:
        data = json.loads(request.body)
        class_id = data.get('class_id')
        period_id = data.get('period_id')
        day = data.get('day')
        subject_id = data.get('subject')
        teacher_id = data.get('teacher')

        # Permission check: only admins can edit the timetable
        # Assumes custom User model has a 'role' attribute with value 'admin' for admins
        user_role = getattr(request.user, 'role', None)
        if user_role != 'admin':
            return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

        if not all([class_id, period_id, day]):
            return JsonResponse({'success': False, 'error': 'Missing required parameters.'}, status=400)

        class_group = get_object_or_404(Class, id=class_id)
        period = get_object_or_404(PeriodSlot, id=period_id)

        # If subject is empty, delete the entry
        if not subject_id:
            DefaultTimetable.objects.filter(class_group=class_group, period=period, day=day).delete()
            return JsonResponse({'success': True, 'message': 'Slot cleared successfully.'})

        # Reject component (child) subjects for timetable allocation
        try:
            subj_obj = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Subject not found.'}, status=404)
        if subj_obj.part_of_id is not None:
            return JsonResponse({'success': False, 'error': 'Component subjects cannot be allocated on the timetable. Select a parent or atomic subject.'}, status=400)

        # Ensure the subject is actually assigned to this class and get allowed teachers via TeacherClassAssignment
        assignments_qs = TeacherClassAssignment.objects.filter(class_group=class_group, subject=subj_obj)
        allowed_teacher_ids = list(assignments_qs.values_list('teacher_id', flat=True).distinct())
        if not allowed_teacher_ids:
            return JsonResponse({'success': False, 'error': f"{subj_obj.name} is not assigned to {class_group}. Add a TeacherClassAssignment first."}, status=400)

        chosen_teacher_id = None
        teacher_obj = None
        if teacher_id:
            # Validate that the teacher exists and is assigned to this class+subject
            try:
                teacher_obj = Teacher.objects.select_related('user').get(id=teacher_id)
            except Teacher.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Teacher not found.'}, status=404)
            if teacher_obj.id not in allowed_teacher_ids:
                return JsonResponse({'success': False, 'error': f"{teacher_obj.user.get_full_name()} is not assigned to teach {subj_obj.name} for {class_group}."}, status=400)
            # Check teacher availability at this day/period (no double booking across classes)
            conflict = DefaultTimetable.objects.filter(day=day, period=period, teacher=teacher_obj).exclude(class_group=class_group).exists()
            if conflict:
                return JsonResponse({'success': False, 'error': f"{teacher_obj.user.get_full_name()} is already assigned to another class at this time."}, status=400)
            chosen_teacher_id = teacher_obj.id
        else:
            # Auto-pick a free teacher from the allowed list if available
            busy_ids = set(DefaultTimetable.objects.filter(day=day, period=period, teacher_id__in=allowed_teacher_ids)
                           .exclude(class_group=class_group)
                           .values_list('teacher_id', flat=True))
            free_ids = [tid for tid in allowed_teacher_ids if tid not in busy_ids]
            if free_ids:
                chosen_teacher_id = free_ids[0]
                teacher_obj = Teacher.objects.select_related('user').get(id=chosen_teacher_id)
            else:
                return JsonResponse({'success': False, 'error': 'All assigned teachers for this subject are busy at this time.'}, status=400)

        # Get or create the timetable entry
        slot, created = DefaultTimetable.objects.get_or_create(
            class_group=class_group,
            period=period,
            day=day,
            defaults={'subject_id': subject_id, 'teacher_id': chosen_teacher_id}
        )

        if not created:
            # If it already exists, update it
            slot.subject_id = subject_id
            slot.teacher_id = chosen_teacher_id
            slot.save()
        
        # Fetch teacher name for the response
        teacher_name = ''
        if slot.teacher:
            teacher_name = slot.teacher.user.get_full_name()

        return JsonResponse({
            'success': True,
            'message': 'Timetable updated successfully.',
            'entry': {
                'subject_name': slot.subject.name,
                'teacher_name': teacher_name
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# @login_required
# @require_POST
# def add_slot_api(request):
#     # This API will be rewritten for the new model.
#     pass

from datetime import date, timedelta, datetime

@login_required
def timetable_api(request):
    # This API will be rewritten to use the DefaultTimetable model.
    return JsonResponse([], safe=False)


# @login_required
# @require_POST
# def edit_slot_api(request, slot_id):
#     # This API will be rewritten for the new model.
#     pass


@login_required
def get_notifications_api(request):
    notifications = Notification.objects.filter(user=request.user, is_read=False).order_by('-created_at')
    data = [
        {
            'id': n.id,
            'message': n.message,
            'created_at': n.created_at.strftime('%Y-%m-%d %H:%M')
        }
        for n in notifications
    ]
    return JsonResponse({'notifications': data, 'count': notifications.count()})


@login_required
@require_POST
def mark_notification_as_read_api(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.save()
    return JsonResponse({'success': True})

from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.http import JsonResponse
import json

@csrf_exempt
@require_POST
def mpesa_callback(request):
    import logging
    from django.utils import timezone
    from decimal import Decimal
    from .models import Student, FeeAssignment, FeePayment, Term, MpesaTransaction
    from .mpesa_utils import query_stk_status
    logger = logging.getLogger('django')
    try:
        # Basic hardening: validate content-type and size, verify shared secret if configured
        from django.conf import settings
        content_type = (request.META.get('CONTENT_TYPE') or request.content_type or '').lower()
        if 'application/json' not in content_type:
            return JsonResponse({"ResultCode": 1, "ResultDesc": "Invalid content type"}, status=415)
        # Limit body size to 100KB to prevent abuse
        if request.META.get('CONTENT_LENGTH') and int(request.META['CONTENT_LENGTH']) > 100 * 1024:
            return JsonResponse({"ResultCode": 1, "ResultDesc": "Payload too large"}, status=413)
        secret_expected = getattr(settings, 'MPESA_CALLBACK_SECRET', '')
        if secret_expected:
            provided = request.META.get('HTTP_X_MPESA_SECRET', '')
            if not provided or provided != secret_expected:
                return JsonResponse({"ResultCode": 1, "ResultDesc": "Unauthorized"}, status=401)
        data = json.loads(request.body.decode('utf-8'))
        stk_callback = data.get('Body', {}).get('stkCallback', {})
        result_code = stk_callback.get('ResultCode', -1)
        merchant_request_id = stk_callback.get('MerchantRequestID')
        checkout_request_id = stk_callback.get('CheckoutRequestID')
        logger.info(f"[M-PESA CALLBACK] Received: {data}")
        # Log callback to file for admin viewing
        import os
        from datetime import datetime
        log_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'payment_callback_logs.txt')
        try:
            # Ensure the log file exists and append safely
            os.makedirs(os.path.dirname(log_path), exist_ok=True)
            with open(log_path, 'a+') as logf:
                logf.seek(0)
                try:
                    logs = json.load(logf)
                except Exception:
                    logs = []
                logs.append({
                    'timestamp': datetime.now().isoformat(),
                    'status': 'success' if result_code == 0 else 'failed',
                    'details': data
                })
                json.dump(logs, logf, indent=2)
                logf.truncate()
        except Exception as logerr:
            logger.error(f"[M-PESA CALLBACK] Could not write callback log: {logerr}")
        # Helper: normalize phone to 2547XXXXXXXX
        def _normalize(msisdn: str) -> str:
            if not msisdn:
                return msisdn
            s = str(msisdn).strip().replace(' ', '')
            if s.startswith('+254'):
                return s[1:]
            if s.startswith('07'):
                return '254' + s[1:]
            if s.startswith('7') and len(s) == 9:
                return '254' + s
            return s

        if result_code == 0:
            metadata = stk_callback.get('CallbackMetadata', {}).get('Item', [])
            amount = next((item['Value'] for item in metadata if item['Name'] == 'Amount'), None)
            phone = _normalize(next((item['Value'] for item in metadata if item['Name'] == 'PhoneNumber'), None))
            mpesa_receipt = next((item['Value'] for item in metadata if item['Name'] == 'MpesaReceiptNumber'), None)
            trans_time = next((item['Value'] for item in metadata if item['Name'] == 'TransactionDate'), None)
            account_ref_meta = next((item['Value'] for item in metadata if item['Name'] in ['AccountReference','BillRefNumber']), None)
            logger.info(f"[M-PESA CALLBACK] Success: Amount={amount}, Phone={phone}, Receipt={mpesa_receipt}")
            if not amount:
                logger.error(f"[M-PESA CALLBACK] Missing amount in metadata: {metadata}")
                return JsonResponse({"ResultCode": 0, "ResultDesc": "Missing payment amount"})
            # Attach to existing MpesaTransaction if available
            tx = None
            if checkout_request_id:
                tx = MpesaTransaction.objects.filter(checkout_request_id=checkout_request_id).first()
            if not tx and merchant_request_id:
                tx = MpesaTransaction.objects.filter(merchant_request_id=merchant_request_id).first()
            # Determine student priority: AccountReference -> tx.student -> phone
            student = None
            if account_ref_meta and isinstance(account_ref_meta, str) and account_ref_meta.startswith('Account#'):
                adm = account_ref_meta.split('#', 1)[1].strip()
                student = Student.objects.filter(admission_no=adm).first()
            if not student and tx and tx.student_id:
                student = tx.student
            # If still no student, try resolving from tx.account_reference
            if not student and tx and tx.account_reference and isinstance(tx.account_reference, str) and tx.account_reference.startswith('Account#'):
                adm = tx.account_reference.split('#', 1)[1].strip()
                student = Student.objects.filter(admission_no=adm).first()
            if not student:
                # Try to match by phone across common formats: 2547..., +2547..., 07...
                phone_variants = []
                if phone:
                    p = str(phone).strip()
                    phone_variants.append(p)
                    if not p.startswith('+'):
                        phone_variants.append('+' + p)
                    # Convert 2547XXXXXXXX to 07XXXXXXXX
                    if p.startswith('254') and len(p) >= 12:
                        alt = '0' + p[3:]
                        phone_variants.append(alt)
                    # Convert +2547XXXXXXXX to 07XXXXXXXX
                    if p.startswith('+254') and len(p) >= 13:
                        alt = '0' + p[4:]
                        phone_variants.append(alt)
                if phone_variants:
                    student = Student.objects.filter(phone__in=list(set(phone_variants))).first()
            if not student:
                logger.warning(f"[M-PESA CALLBACK] No student found with phone {phone}. Logging unmatched payment.")
                # Log unmatched callback for audit
                unmatched_log_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'payment_callback_logs.txt')
                try:
                    os.makedirs(os.path.dirname(unmatched_log_path), exist_ok=True)
                    with open(unmatched_log_path, 'a+') as logf:
                        logf.seek(0)
                        try:
                            logs = json.load(logf)
                        except Exception:
                            logs = []
                        logs.append({
                            'timestamp': datetime.now().isoformat(),
                            'status': 'unmatched',
                            'details': {
                                'phone': phone,
                                'amount': amount,
                                'reference': mpesa_receipt,
                                'raw': data
                            }
                        })
                        json.dump(logs, logf, indent=2)
                        logf.truncate()
                except Exception as logerr:
                    logger.error(f"[M-PESA CALLBACK] Could not write unmatched callback log: {logerr}")
                return JsonResponse({"ResultCode": 0, "ResultDesc": "Student not found, but callback logged"})
            # Ensure there is at least a stub MpesaTransaction for traceability
            if not tx:
                try:
                    # Generate a unique checkout id if missing
                    if not checkout_request_id:
                        from datetime import datetime
                        checkout_request_id = f"cb-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(phone)[-4:]}"
                    tx = MpesaTransaction.objects.create(
                        student=student,
                        phone_number=phone,
                        amount=Decimal(str(amount)),
                        account_reference=f"Account#{student.admission_no}",
                        merchant_request_id=merchant_request_id,
                        checkout_request_id=checkout_request_id,
                        status='pending'
                    )
                except Exception as e:
                    logger.error(f"[M-PESA CALLBACK] Could not create stub MpesaTransaction: {e}")
            # Find current term
            today = timezone.now().date()
            current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()
            if not current_term:
                logger.warning(f"[M-PESA CALLBACK] No current term found for today {today}; proceeding without linking to a specific term.")
            # Find unpaid FeeAssignment(s) for student's class in current term
            fee_assignments = FeeAssignment.objects.filter(class_group=student.class_group, term=current_term) if current_term else FeeAssignment.objects.none()
            if current_term and not fee_assignments.exists():
                logger.warning(f"[M-PESA CALLBACK] No fee assignments for student {student} in term {current_term}; will record without assignment.")
            # Try to find an assignment with outstanding balance
            from django.db.models import Sum, F
            outstanding_assignment = None
            for assignment in fee_assignments:
                total_paid = FeePayment.objects.filter(student=student, fee_assignment=assignment).aggregate(total=Sum('amount_paid'))['total'] or 0
                if float(total_paid) < float(assignment.amount):
                    outstanding_assignment = assignment
                    break
            if not outstanding_assignment and fee_assignments.exists():
                # If all paid, just link to first assignment
                outstanding_assignment = fee_assignments.first()
            # Idempotency: don't double-record (by receipt)
            if mpesa_receipt and FeePayment.objects.filter(reference=mpesa_receipt).exists():
                logger.warning(f"[M-PESA CALLBACK] Duplicate payment detected: Receipt={mpesa_receipt}, Phone={phone}")
                # Update linked MpesaTransaction if found
                if tx:
                    tx.status = 'success'
                    tx.result_code = int(result_code)
                    tx.result_desc = stk_callback.get('ResultDesc')
                    tx.mpesa_receipt = mpesa_receipt
                    tx.raw_callback = data
                    tx.save(update_fields=['status','result_code','result_desc','mpesa_receipt','raw_callback','updated_at'])
                return JsonResponse({"ResultCode": 0, "ResultDesc": "Duplicate payment ignored"})
            # Verify with Daraja STK Query before approving unless auto-approve is enabled
            auto_approve = getattr(settings, 'MPESA_AUTO_APPROVE_ON_CALLBACK', False)
            logger.info(f"[M-PESA CALLBACK] Auto-approve setting: {auto_approve}")
            verified_ok = False
            query_resp = None
            if auto_approve:
                verified_ok = True
                query_resp = {"auto_approve": True}
                logger.info("[M-PESA CALLBACK] Auto-approving payment without STK query.")
            else:
                query_resp = query_stk_status(checkout_request_id) if checkout_request_id else {"error": "missing CheckoutRequestID"}
                if isinstance(query_resp, dict) and not query_resp.get('error') and str(query_resp.get('ResultCode')) == '0':
                    verified_ok = True
                else:
                    logger.warning(f"[M-PESA CALLBACK] STK query did not confirm success: {query_resp}")
            if not verified_ok:
                # Mark transaction pending and create a visible pending FeePayment (idempotent)
                if tx:
                    tx.status = 'pending'
                    tx.result_code = int(result_code)
                    tx.result_desc = stk_callback.get('ResultDesc')
                    tx.mpesa_receipt = mpesa_receipt
                    tx.raw_callback = data
                    tx.save(update_fields=['status','result_code','result_desc','mpesa_receipt','raw_callback','updated_at'])
                # Create a pending FeePayment if not already created
                try:
                    already = False
                    if mpesa_receipt:
                        already = FeePayment.objects.filter(reference=mpesa_receipt).exists()
                    if not already and tx:
                        already = FeePayment.objects.filter(mpesa_transaction=tx).exists()
                    if not already:
                        FeePayment.objects.create(
                            student=student,
                            fee_assignment=outstanding_assignment,
                            amount_paid=Decimal(str(amount)),
                            payment_method="mpesa",
                            reference=mpesa_receipt,
                            phone_number=phone,
                            status="pending",
                            mpesa_transaction=tx if tx else None
                        )
                except Exception as e:
                    logger.error(f"[M-PESA CALLBACK] Error creating pending FeePayment: {e}")
                logger.warning(f"[M-PESA CALLBACK] Verification failed or pending for CheckoutRequestID={checkout_request_id}, query={query_resp}")
                return JsonResponse({"ResultCode": 0, "ResultDesc": "Verification pending"})
            # Create approved payment
            try:
                # Update in-place if a pending payment exists; otherwise create new
                updated = False
                if mpesa_receipt:
                    pending_fp = FeePayment.objects.filter(reference=mpesa_receipt).first()
                    if pending_fp:
                        pending_fp.fee_assignment = outstanding_assignment
                        pending_fp.amount_paid = Decimal(str(amount))
                        pending_fp.status = "approved"
                        pending_fp.mpesa_transaction = tx if tx else pending_fp.mpesa_transaction
                        pending_fp.save(update_fields=["fee_assignment","amount_paid","status","mpesa_transaction"])
                        updated = True
                if not updated and tx:
                    pending_fp = FeePayment.objects.filter(mpesa_transaction=tx).first()
                    if pending_fp:
                        pending_fp.fee_assignment = outstanding_assignment
                        pending_fp.reference = pending_fp.reference or mpesa_receipt
                        pending_fp.amount_paid = Decimal(str(amount))
                        pending_fp.phone_number = phone
                        pending_fp.status = "approved"
                        pending_fp.save(update_fields=["fee_assignment","reference","amount_paid","phone_number","status"])
                        updated = True
                if not updated:
                    FeePayment.objects.create(
                        student=student,
                        fee_assignment=outstanding_assignment,
                        amount_paid=Decimal(str(amount)),
                        payment_method="mpesa",
                        reference=mpesa_receipt,
                        phone_number=phone,
                        status="approved",
                        mpesa_transaction=tx if tx else None
                    )
            except Exception as e:
                logger.error(f"[M-PESA CALLBACK] Error creating/updating FeePayment: {e}")
                return JsonResponse({"ResultCode": 0, "ResultDesc": "Recorded with issues"})
            if tx:
                tx.status = 'success'
                tx.result_code = int(result_code)
                tx.result_desc = stk_callback.get('ResultDesc')
                tx.mpesa_receipt = mpesa_receipt
                tx.raw_callback = data
                # If we computed outstanding_assignment, link it to tx for completeness
                if outstanding_assignment and not tx.fee_assignment_id:
                    tx.fee_assignment = outstanding_assignment
                tx.save(update_fields=['status','result_code','result_desc','mpesa_receipt','raw_callback','updated_at','fee_assignment'])
            else:
                logger.error(f"[M-PESA CALLBACK] Invalid payment: missing MpesaReceiptNumber for phone {phone}, amount {amount}")
                return JsonResponse({"ResultCode": 0, "ResultDesc": "Invalid payment: missing reference number"})
            logger.info(f"[M-PESA CALLBACK] Payment recorded for student {student} amount {amount} assignment {outstanding_assignment} (pending verification)")
            # Send confirmation SMS to payer and student (best-effort)
            try:
                school_name = getattr(settings, 'SCHOOL_NAME', 'Your School')
                amt_txt = str(amount)
                stud_name = getattr(student, 'name', None) or f"Adm {getattr(student, 'admission_no', '')}"
                receipt_txt = mpesa_receipt or checkout_request_id or ''
                msg = f"{school_name}: Payment received KES {amt_txt} for {stud_name}. Receipt {receipt_txt}. Thank you."
                # SMS to payer (from callback metadata)
                if phone:
                    try:
                        send_sms(phone, msg)
                    except Exception:
                        pass
                # SMS to student's registered number if different
                student_phone = getattr(student, 'phone', None)
                if student_phone and (not phone or str(student_phone).strip() != str(phone).strip()):
                    try:
                        send_sms(student_phone, msg)
                    except Exception:
                        pass
            except Exception:
                # Never fail the callback because of SMS issues
                pass
        else:
            # Extract phone, amount, receipt if present
            metadata = stk_callback.get('CallbackMetadata', {}).get('Item', [])
            amount = next((item['Value'] for item in metadata if item['Name'] == 'Amount'), None)
            phone = _normalize(next((item['Value'] for item in metadata if item['Name'] == 'PhoneNumber'), None))
            mpesa_receipt = next((item['Value'] for item in metadata if item['Name'] == 'MpesaReceiptNumber'), None)
            # Only record if we have at least a phone and student
            if phone:
                student = Student.objects.filter(phone=phone).first()
                if student:
                    today = timezone.now().date()
                    current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()
                    fee_assignments = FeeAssignment.objects.filter(class_group=student.class_group, term=current_term)
                    assignment = fee_assignments.first() if fee_assignments.exists() else None
                    # Idempotency: don't double-record
                    if not (mpesa_receipt and FeePayment.objects.filter(reference=mpesa_receipt).exists()):
                        if mpesa_receipt:
                            FeePayment.objects.create(
                                student=student,
                                fee_assignment=assignment,
                                amount_paid=Decimal(str(amount or 0)),
                                payment_method="mpesa",
                                reference=mpesa_receipt,
                                phone_number=phone,
                                status="rejected"
                            )
                            logger.info(f"[M-PESA CALLBACK] Failed payment recorded for student {student} assignment {assignment} phone {phone}")
                        else:
                            logger.error(f"[M-PESA CALLBACK] Invalid failed payment: missing MpesaReceiptNumber for phone {phone}, amount {amount}")
                            return JsonResponse({"ResultCode": 0, "ResultDesc": "Invalid failed payment: missing reference number"})
            # Update MpesaTransaction if present
            if checkout_request_id:
                tx = MpesaTransaction.objects.filter(checkout_request_id=checkout_request_id).first()
                if tx:
                    tx.status = 'failed'
                    tx.result_code = int(result_code)
                    tx.result_desc = stk_callback.get('ResultDesc')
                    tx.mpesa_receipt = mpesa_receipt
                    tx.raw_callback = data
                    tx.save(update_fields=['status','result_code','result_desc','mpesa_receipt','raw_callback','updated_at'])
            logger.warning(f"[M-PESA CALLBACK] Payment failed or cancelled. ResultCode={result_code}, Data={data}")
    except Exception as e:
        logger.error(f"[M-PESA CALLBACK] Error processing callback: {e}", exc_info=True)
        # DEBUG: return error detail to help diagnose in dev
        return JsonResponse({"ResultCode": 0, "ResultDesc": f"Error processing callback: {e}"})
    return JsonResponse({"ResultCode": 0, "ResultDesc": "Received"})


# --- User-facing PayBill Flow ---
@login_required
def paybill_start(request):
    """Render a simple form for students to initiate an M-Pesa PayBill STK push.

    Fields: amount, phone number. Uses student's admission number for AccountReference.
    """
    user = request.user
    # Only students may initiate from this page (can be relaxed later)
    if getattr(user, 'role', None) != 'student':
        return HttpResponseForbidden("Only students can initiate M-Pesa payments here.")
    student = Student.objects.filter(user=user).first()
    if not student:
        messages.error(request, 'No student profile linked to your account.')
        return redirect('login')

    if request.method == 'POST':
        amount_raw = request.POST.get('amount')
        phone_raw = request.POST.get('phone')
        try:
            amount_val = float(amount_raw)
        except (TypeError, ValueError):
            amount_val = 0
        phone = _normalize_msisdn(phone_raw or '')
        if amount_val <= 0:
            messages.error(request, 'Enter a valid amount greater than 0.')
            return render(request, 'fees/paybill_start.html', {
                'student': student,
                'shortcode': getattr(settings, 'MPESA_SHORTCODE', ''),
                'prefill_phone': phone_raw or '',
            })
        if not phone or not phone.startswith('254'):
            messages.error(request, 'Enter a valid Kenyan phone number (e.g., 07XX or 2547XX...).')
            return render(request, 'fees/paybill_start.html', {
                'student': student,
                'shortcode': getattr(settings, 'MPESA_SHORTCODE', ''),
                'prefill_phone': phone_raw or '',
            })

        account_ref = f"Account#{student.admission_no}"
        tx_desc = f"Fees payment for {student.admission_no}"
        # Call Daraja to initiate STK push
        resp = initiate_stk_push(phone, amount_val, account_ref, tx_desc)
        merchant_id = None
        checkout_id = None
        error_text = None
        if isinstance(resp, dict):
            merchant_id = resp.get('MerchantRequestID')
            checkout_id = resp.get('CheckoutRequestID')
            if not checkout_id:
                # Surface error details from Daraja
                error_text = resp.get('errorMessage') or resp.get('error') or resp.get('ResponseDescription')
        if not checkout_id:
            messages.error(request, f"Failed to initiate STK push: {error_text or 'Unknown error'}")
            return render(request, 'fees/paybill_start.html', {
                'student': student,
                'shortcode': getattr(settings, 'MPESA_SHORTCODE', ''),
                'prefill_phone': phone_raw or '',
            })

        # Record a pending MpesaTransaction for tracking
        tx = MpesaTransaction.objects.create(
            student=student,
            phone_number=phone,
            amount=amount_val,
            account_reference=account_ref,
            merchant_request_id=merchant_id,
            checkout_request_id=checkout_id,
            status='pending'
        )
        # Redirect to status page
        return redirect('paybill_status', checkout_request_id=tx.checkout_request_id)

    return render(request, 'fees/paybill_start.html', {
        'student': student,
        'shortcode': getattr(settings, 'MPESA_SHORTCODE', ''),
        'prefill_phone': student.phone or '',
    })


@login_required
def paybill_status(request, checkout_request_id: str):
    """Show live status for a specific STK push transaction and allow polling/query."""
    user = request.user
    tx = MpesaTransaction.objects.filter(checkout_request_id=checkout_request_id).first()
    if not tx:
        messages.error(request, 'Transaction not found.')
        return redirect('paybill_start')
    # Authorization: only owner student or admin
    if getattr(user, 'role', None) not in ('admin', 'teacher'):
        owner_id = getattr(getattr(tx, 'student', None), 'user_id', None)
        if owner_id and owner_id != user.id:
            return HttpResponseForbidden('Not authorized to view this transaction.')

    # Poll on demand (GET param ?refresh=1) or auto when pending
    if request.GET.get('refresh') == '1' or tx.status == 'pending':
        q = query_stk_status(tx.checkout_request_id)
        # Update local tx based on query if it returns valid data
        if isinstance(q, dict) and not q.get('error'):
            result_code = q.get('ResultCode')
            result_desc = q.get('ResultDesc')
            if result_code is not None:
                try:
                    tx.result_code = int(result_code)
                except Exception:
                    pass
            if result_desc:
                tx.result_desc = result_desc
            # Heuristic: ResultCode 0 means success; others may be in-flight or failed
            if str(q.get('ResultCode')) == '0':
                tx.status = 'success'
            elif str(q.get('ResultCode')) not in (None, '', '0'):
                tx.status = 'failed'
            tx.save()

    context = {
        'tx': tx,
        'shortcode': getattr(settings, 'MPESA_SHORTCODE', ''),
    }
    return render(request, 'fees/paybill_status.html', context)


# --- PayBill C2B Callbacks (Validation & Confirmation) ---
@csrf_exempt
def mpesa_c2b_validation(request):
    """Accept or reject an incoming C2B transaction (PayBill). Minimal validation.

    Expected Safaricom JSON fields include: TransType, TransID, TransTime, TransAmount,
    BusinessShortCode, BillRefNumber, MSISDN, etc.
    """
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception:
        payload = {}
    # Optionally, validate BusinessShortCode matches settings
    shortcode_ok = True
    short_expected = getattr(settings, 'MPESA_SHORTCODE', '')
    if short_expected:
        shortcode_ok = str(payload.get('BusinessShortCode') or payload.get('BusinessShortCode')) == str(short_expected)
    if not shortcode_ok:
        return JsonResponse({"ResultCode": 1, "ResultDesc": "Invalid shortcode"})
    # Accept by default
    return JsonResponse({"ResultCode": 0, "ResultDesc": "Accepted"})


@csrf_exempt
def mpesa_c2b_confirmation(request):
    """Record a confirmed C2B payment. Creates/updates FeePayment and MpesaTransaction.
    Uses TransID as reference and BillRefNumber to identify the student (Account#<adm_no>).
    """
    import logging
    from decimal import Decimal
    logger = logging.getLogger('django')
    try:
        data = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception as e:
        logger.error(f"[C2B CONFIRMATION] Invalid JSON: {e}")
        return JsonResponse({"ResultCode": 0, "ResultDesc": "Received"})

    # Extract fields
    trans_id = data.get('TransID')
    amount = data.get('TransAmount')
    msisdn = data.get('MSISDN') or data.get('MSISDNNumber')
    bill_ref = data.get('BillRefNumber') or data.get('AccountReference')
    phone = _normalize_msisdn(str(msisdn or ''))

    # Persist raw confirmation to the immutable ledger for later reconciliation
    try:
        from .models import MpesaC2BLedger
        ledger_defaults = {
            'trans_time': data.get('TransTime'),
            'amount': Decimal(str(amount or 0)) if amount is not None else None,
            'msisdn': phone or (msisdn or None),
            'bill_ref': bill_ref,
            'business_short_code': data.get('BusinessShortCode') or data.get('ShortCode'),
            'third_party_trans_id': data.get('ThirdPartyTransID'),
            'first_name': data.get('FirstName'),
            'middle_name': data.get('MiddleName'),
            'last_name': data.get('LastName'),
            'org_account_balance': data.get('OrgAccountBalance'),
            'raw': data,
        }
        if trans_id:
            obj, created = MpesaC2BLedger.objects.get_or_create(trans_id=trans_id, defaults=ledger_defaults)
            if not created:
                # Update raw/latest details in case of retries
                for k, v in ledger_defaults.items():
                    setattr(obj, k, v)
                obj.save()
    except Exception as e:
        logger.error(f"[C2B CONFIRMATION] Ledger persist error: {e}")

    # Log to file for audit
    try:
        import os
        from datetime import datetime
        log_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'payment_callback_logs.txt')
        os.makedirs(os.path.dirname(log_path), exist_ok=True)
        with open(log_path, 'a+') as logf:
            logf.seek(0)
            try:
                logs = json.load(logf)
            except Exception:
                logs = []
            logs.append({
                'timestamp': datetime.now().isoformat(),
                'status': 'c2b_confirmation',
                'details': data
            })
            json.dump(logs, logf, indent=2)
            logf.truncate()
    except Exception as logerr:
        logger.error(f"[C2B CONFIRMATION] Could not write log: {logerr}")

    if not trans_id:
        return JsonResponse({"ResultCode": 0, "ResultDesc": "Missing TransID"})

    # Resolve student by bill reference if in expected pattern
    student = None
    if isinstance(bill_ref, str) and bill_ref.startswith('Account#'):
        adm = bill_ref.split('#', 1)[1].strip()
        student = Student.objects.filter(admission_no=adm).first()
    if not student and phone:
        student = Student.objects.filter(phone=phone).first()

    # Optional: link to an existing pending MpesaTransaction by account_reference
    tx = MpesaTransaction.objects.filter(mpesa_receipt=trans_id).first()
    if not tx:
        tx = MpesaTransaction.objects.filter(account_reference=bill_ref).order_by('-created_at').first()

    # Find current term and an outstanding assignment
    from django.db.models import Sum as _Sum
    today = timezone.now().date()
    current_term = Term.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date').first()
    outstanding_assignment = None
    if student and current_term:
        fee_assignments = FeeAssignment.objects.filter(class_group=student.class_group, term=current_term)
        for fa in fee_assignments:
            total_paid = FeePayment.objects.filter(student=student, fee_assignment=fa).aggregate(total=_Sum('amount_paid'))['total'] or 0
            if float(total_paid) < float(fa.amount):
                outstanding_assignment = fa
                break
        if not outstanding_assignment and fee_assignments.exists():
            outstanding_assignment = fee_assignments.first()

    # Idempotency: don't double-create FeePayment
    existing = FeePayment.objects.filter(reference=trans_id).first()
    if existing:
        # Ensure it's approved and linked
        if tx and not existing.mpesa_transaction_id:
            existing.mpesa_transaction = tx
            existing.save(update_fields=['mpesa_transaction'])
        return JsonResponse({"ResultCode": 0, "ResultDesc": "Recorded"})

    # Create/Update MpesaTransaction
    try:
        if not tx:
            # Create a lightweight transaction record for traceability
            tx = MpesaTransaction.objects.create(
                student=student,
                phone_number=phone or '',
                amount=Decimal(str(amount or 0)),
                account_reference=bill_ref,
                checkout_request_id=f"c2b-{trans_id}",
                status='success',
                mpesa_receipt=trans_id,
                result_desc='C2B Confirmation'
            )
        else:
            tx.mpesa_receipt = tx.mpesa_receipt or trans_id
            tx.status = 'success'
            tx.result_desc = 'C2B Confirmation'
            tx.save(update_fields=['mpesa_receipt','status','result_desc','updated_at'])
    except Exception as e:
        logger.error(f"[C2B CONFIRMATION] Error updating MpesaTransaction: {e}")

    # Create approved FeePayment
    try:
        if student:
            FeePayment.objects.create(
                student=student,
                fee_assignment=outstanding_assignment,
                amount_paid=Decimal(str(amount or 0)),
                payment_method='mpesa',
                reference=trans_id,
                phone_number=phone,
                status='approved',
                mpesa_transaction=tx
            )
            # Send confirmation SMS to payer and student (best-effort)
            try:
                school_name = getattr(settings, 'SCHOOL_NAME', 'Your School')
                amt_txt = str(amount or 0)
                stud_name = getattr(student, 'name', None) or f"Adm {getattr(student, 'admission_no', '')}"
                receipt_txt = trans_id
                msg = f"{school_name}: Payment received KES {amt_txt} for {stud_name}. Receipt {receipt_txt}. Thank you."
                if phone:
                    try:
                        send_sms(phone, msg)
                    except Exception:
                        pass
                student_phone = getattr(student, 'phone', None)
                if student_phone and (not phone or str(student_phone).strip() != str(phone).strip()):
                    try:
                        send_sms(student_phone, msg)
                    except Exception:
                        pass
            except Exception:
                pass
        else:
            # If no student resolved, record as pending without student link (optional policy)
            logger.warning(f"[C2B CONFIRMATION] No student matched for BillRef={bill_ref} phone={phone}")
    except Exception as e:
        logger.error(f"[C2B CONFIRMATION] Error creating FeePayment: {e}")

    return JsonResponse({"ResultCode": 0, "ResultDesc": "Received"})